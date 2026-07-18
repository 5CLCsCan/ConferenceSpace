#!/usr/bin/env python3
"""Export workflow benchmark outputs to Excel and report-ready PNG charts.

Reads docs/report/raw/workflow_benchmarks/benchmark_output and writes:
  - exports/workflow_benchmark_results.xlsx
  - exports/figures/*.png
  - exports/FIGURE_INDEX.md
"""

from __future__ import annotations

import json
import textwrap
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "benchmark_output"
EXPORT = ROOT / "exports"
FIGURES = EXPORT / "figures"

# Manual chatbot outcomes from chatbot_agent_benchmark_report.md (section 5.1 / 5.2)
CHATBOT_MANUAL_OVERALL = {
    "scenarios": 8,
    "trials": 40,
    "completed_trials": 40,
    "pass": 25,
    "partial": 12,
    "fail": 3,
}
CHATBOT_MANUAL_BY_SCENARIO = {
    "author_own_submission_status": {
        "label_vi": "Tra cứu trạng thái bài nộp",
        "manual": "partial",
        "note": "4/5 đúng trạng thái; 1 lượt dừng ở lỗi truy vấn",
    },
    "author_submission_track": {
        "label_vi": "Tra cứu track và metadata",
        "manual": "partial",
        "note": "4/5 đúng track; 1 lượt nhầm định danh hội nghị",
    },
    "chair_conference_overview": {
        "label_vi": "Tóm tắt tình hình hội nghị",
        "manual": "partial",
        "note": "Tổng quan chính đúng; một số lượt thiếu reviewer count",
    },
    "reviewer_assignment_check": {
        "label_vi": "Kiểm tra workload reviewer",
        "manual": "pass",
        "note": "5/5 kết luận đúng workload",
    },
    "public_conference_lookup": {
        "label_vi": "Tra cứu thông tin công khai",
        "manual": "partial",
        "note": "4/5 đúng metadata; 1 lượt sai nguồn dữ liệu",
    },
    "permission_boundary_other_submission": {
        "label_vi": "Ranh giới quyền truy cập",
        "manual": "partial",
        "note": "Không lộ dữ liệu; diễn đạt còn giống lỗi kỹ thuật",
    },
    "unsupported_external_research": {
        "label_vi": "Yêu cầu ngoài phạm vi",
        "manual": "partial",
        "note": "1/5 vẫn viết báo cáo ngoài phạm vi nền tảng",
    },
    "chair_multi_step_platform_report": {
        "label_vi": "Báo cáo vận hành nhiều bước",
        "manual": "partial",
        "note": "Chain tool tốt ở một số lượt; lỗi truy vấn làm thiếu chi tiết",
    },
}

PALETTE = {
    "navy": "#1B3A4B",
    "teal": "#0D7377",
    "coral": "#C44536",
    "amber": "#E09F3E",
    "slate": "#4A5568",
    "mint": "#2A9D8F",
    "sand": "#F4A261",
    "ink": "#264653",
    "pass": "#2A9D8F",
    "partial": "#E09F3E",
    "fail": "#C44536",
    "warn": "#E09F3E",
    "block": "#C44536",
}


def read_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1B3A4B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:80]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str) -> None:
    out = df.copy()
    if out.empty:
        out = pd.DataFrame({"note": ["(empty)"]})
    out.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.book[sheet]
    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def safe_mean(values: list[float | int | None]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return round(mean(clean), 4) if clean else None


def safe_median(values: list[float | int | None]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return round(median(clean), 4) if clean else None


def pct(n: float | int | None, d: float | int | None) -> float | None:
    if n is None or d is None or d == 0:
        return None
    return round(100.0 * float(n) / float(d), 2)


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_track() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "track_recommendation" / "human_review.csv")
    metrics = read_json(INPUT / "track_recommendation" / "summary_metrics.json")
    manifest = read_json(INPUT / "track_recommendation" / "manifest.json")
    conf_counts = review["conference"].value_counts().to_dict()
    summary = {
        "workflow": "track_recommendation",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "invalid_track_rate": metrics.get("invalid_track_rate"),
        "invalid_track_cases": int((review["invalid_track_names"].fillna("").astype(str).str.len() > 0).sum()),
        "human_label_filled": int(review["human_label"].fillna("").astype(str).str.strip().ne("").sum()),
        "top1_human_accuracy": metrics.get("top1_human_accuracy"),
        "top3_human_acceptance": metrics.get("top3_human_acceptance"),
        "conferences": len(conf_counts),
        "model": (manifest.get("llm_config") or {}).get("openai_model"),
        "agent_model": (manifest.get("llm_config") or {}).get("agent_model"),
        "created_at": manifest.get("created_at"),
        "limitation": "human_label còn trống — chưa có top-1/top-3 accuracy từ nhãn người",
    }
    review = review.copy()
    review["has_prediction"] = review["predicted_top1"].fillna("").astype(str).str.strip().ne("")
    review["top3_count"] = review["predicted_top3"].fillna("").astype(str).apply(
        lambda s: len([x for x in s.split(";") if x.strip()]) if s.strip() else 0
    )
    return review, summary


def load_gating_rule() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "submission_gating_rule_check" / "rule_review.csv")
    metrics = read_json(INPUT / "submission_gating_rule_check" / "summary_metrics.json")
    manifest = read_json(INPUT / "submission_gating_rule_check" / "manifest.json")
    summary = {
        "workflow": "submission_gating_rule_check",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "correct_verdict_count": metrics.get("correct_verdict_count"),
        "correct_rules_count": metrics.get("correct_rules_count"),
        "false_block_count": metrics.get("false_block_count"),
        "blocking_verdict_accuracy": metrics.get("blocking_verdict_accuracy"),
        "rule_id_recall": metrics.get("rule_id_recall"),
        "created_at": manifest.get("created_at"),
    }
    return review, summary


def load_gating_llm() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "submission_gating_llm_steering" / "llm_steering_review.csv")
    metrics = read_json(INPUT / "submission_gating_llm_steering" / "summary_metrics.json")
    manifest = read_json(INPUT / "submission_gating_llm_steering" / "manifest.json")
    findings = read_jsonl(INPUT / "submission_gating_llm_steering" / "normalized_content_findings.jsonl")
    summary = {
        "workflow": "submission_gating_llm_steering",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "content_finding_count": metrics.get("content_finding_count"),
        "llm_block_contract_violation_count": metrics.get("llm_block_contract_violation_count"),
        "review_rows": len(review),
        "verdict_pass": int((review["verdict"] == "pass").sum()) if "verdict" in review else None,
        "verdict_warn": int((review["verdict"] == "warn").sum()) if "verdict" in review else None,
        "verdict_block": int((review["verdict"] == "block").sum()) if "verdict" in review else None,
        "normalized_findings": len(findings),
        "human_label_filled": int(review["human_label"].fillna("").astype(str).str.strip().ne("").sum())
        if "human_label" in review
        else 0,
        "created_at": manifest.get("created_at"),
        "limitation": "human grounded/actionable labels còn trống",
    }
    # Compact columns for Excel readability
    keep = [
        c
        for c in [
            "task_id",
            "case_id",
            "conference",
            "paper_id",
            "scenario_tags",
            "chair_focus_area",
            "submission_title",
            "verdict",
            "decision",
            "finding_rule_id",
            "finding_source",
            "finding_severity",
            "finding_message",
            "expected_behavior",
            "llm_block_contract_violation",
            "human_label",
            "grounded",
            "severity_ok",
            "actionable",
        ]
        if c in review.columns
    ]
    compact = review[keep].copy()
    for col in ["finding_message", "expected_behavior", "submission_title"]:
        if col in compact.columns:
            compact[col] = compact[col].astype(str).str.slice(0, 240)
    return compact, summary


def load_chatbot() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    run = read_json(INPUT / "chatbot_agent" / "chatbot_run_1783406671_1783411092" / "run_summary.json")
    rows = []
    for r in run.get("results", []):
        trial = r.get("trial") or {}
        timing = r.get("timing_metrics") or {}
        tools = r.get("tool_metrics") or {}
        tokens = r.get("token_usage") or {}
        rows.append(
            {
                "scenario_id": r.get("scenario_id"),
                "scenario_result_id": r.get("scenario_result_id"),
                "trial_index": trial.get("index"),
                "actor": r.get("actor"),
                "status": r.get("status"),
                "duration_ms": r.get("duration_ms"),
                "ttft_ms": timing.get("ttft_ms"),
                "time_to_first_answer_token_ms": timing.get("time_to_first_answer_token_ms"),
                "answer_stream_duration_ms": timing.get("answer_stream_duration_ms"),
                "stream_duration_ms": timing.get("stream_duration_ms"),
                "tool_call_count": tools.get("tool_call_count"),
                "successful_tool_call_count": tools.get("successful_tool_call_count"),
                "failed_tool_call_count": tools.get("failed_tool_call_count"),
                "tool_call_success_rate": tools.get("tool_call_success_rate"),
                "final_answer_chars": r.get("final_answer_chars"),
                "input_tokens": tokens.get("input_tokens"),
                "output_tokens": tokens.get("output_tokens"),
                "total_tokens": tokens.get("total_tokens"),
                "error": r.get("error") or "",
            }
        )
    trials = pd.DataFrame(rows)

    scenario_rows = []
    for scenario_id, group in trials.groupby("scenario_id", sort=False):
        meta = CHATBOT_MANUAL_BY_SCENARIO.get(scenario_id, {})
        tool_calls = int(group["tool_call_count"].fillna(0).sum())
        tool_ok = int(group["successful_tool_call_count"].fillna(0).sum())
        scenario_rows.append(
            {
                "scenario_id": scenario_id,
                "label_vi": meta.get("label_vi", scenario_id),
                "actor": group["actor"].iloc[0],
                "trials": len(group),
                "manual_outcome": meta.get("manual", ""),
                "manual_note": meta.get("note", ""),
                "avg_duration_s": round(group["duration_ms"].mean() / 1000.0, 2),
                "avg_ttft_s": round(group["ttft_ms"].mean() / 1000.0, 2),
                "avg_time_to_first_answer_s": round(
                    group["time_to_first_answer_token_ms"].mean() / 1000.0, 2
                ),
                "total_tool_calls": tool_calls,
                "successful_tool_calls": tool_ok,
                "tool_success_rate_pct": pct(tool_ok, tool_calls),
                "avg_total_tokens": round(group["total_tokens"].mean(), 1),
            }
        )
    by_scenario = pd.DataFrame(scenario_rows)
    agg = run.get("aggregate_metrics") or {}
    summary = {
        "workflow": "chatbot_agent",
        "run_instance_id": run.get("run_instance_id"),
        "started_at": run.get("started_at"),
        "finished_at": run.get("finished_at"),
        "scenario_count": agg.get("scenario_count"),
        "trial_count": agg.get("trial_count"),
        "avg_total_duration_ms": agg.get("avg_total_duration_ms"),
        "avg_ttft_ms": agg.get("avg_ttft_ms"),
        "avg_time_to_first_answer_token_ms": agg.get("avg_time_to_first_answer_token_ms"),
        "avg_stream_duration_ms": agg.get("avg_stream_duration_ms"),
        "total_tool_calls": agg.get("total_tool_calls"),
        "successful_tool_calls": agg.get("successful_tool_calls"),
        "failed_tool_calls": agg.get("failed_tool_calls"),
        "tool_call_success_rate": agg.get("tool_call_success_rate"),
        "input_tokens": agg.get("input_tokens"),
        "output_tokens": agg.get("output_tokens"),
        "total_tokens": agg.get("total_tokens"),
        "manual_pass": CHATBOT_MANUAL_OVERALL["pass"],
        "manual_partial": CHATBOT_MANUAL_OVERALL["partial"],
        "manual_fail": CHATBOT_MANUAL_OVERALL["fail"],
        "manual_pass_rate_pct": pct(
            CHATBOT_MANUAL_OVERALL["pass"], CHATBOT_MANUAL_OVERALL["trials"]
        ),
        "source_manual": "chatbot_agent_benchmark_report.md",
    }
    return trials, by_scenario, summary


def load_autofill() -> tuple[pd.DataFrame, dict[str, Any]]:
    path = INPUT / "completed (7).csv"
    df = pd.read_csv(path)
    summary = {
        "workflow": "submission_autofill_completed_csv",
        "case_count": len(df),
        "title_exact_match_rate": round(df["title_exact_match"].mean(), 4)
        if "title_exact_match" in df
        else None,
        "abstract_rouge_l_mean": safe_mean(df["abstract_rouge_l"].tolist())
        if "abstract_rouge_l" in df
        else None,
        "abstract_rouge_l_median": safe_median(df["abstract_rouge_l"].tolist())
        if "abstract_rouge_l" in df
        else None,
        "keyword_f1_mean": safe_mean(df["keyword_f1"].tolist()) if "keyword_f1" in df else None,
        "author_f1_mean": safe_mean(df["author_f1"].tolist()) if "author_f1" in df else None,
        "time_taken_seconds_mean": safe_mean(df["time_taken_seconds"].tolist())
        if "time_taken_seconds" in df
        else None,
        "time_taken_seconds_median": safe_median(df["time_taken_seconds"].tolist())
        if "time_taken_seconds" in df
        else None,
        "total_tokens_mean": safe_mean(df["total_tokens"].tolist()) if "total_tokens" in df else None,
        "source_file": path.name,
    }
    return df, summary


def load_tca() -> tuple[pd.DataFrame, dict[str, Any]]:
    rows_raw = read_jsonl(INPUT / "tca_benchmark_results (5).jsonl")
    flat = []
    b1_exact_rates: list[float] = []
    b1_partial_rates: list[float] = []
    b1_fab_rates: list[float] = []
    b2_t: list[float] = []
    b2_c: list[float] = []
    b2_a: list[float] = []
    b5_t: list[float] = []
    b5_c: list[float] = []
    b5_a: list[float] = []
    b3_t: list[float] = []
    b3_v: list[float] = []
    b3_gv: list[float] = []
    high_risk = 0
    clean = 0
    times: list[float] = []

    # Absolute unit totals for report-facing counts next to rates
    ann_exact = ann_partial = ann_fab = 0
    ap_total = 0
    review_total = 0
    finding_total = 0
    claim_eb_total = 0
    papers_with_ap = 0
    papers_with_b3 = 0
    papers_with_b5 = 0

    for r in rows_raw:
        b1 = r.get("b1") or {}
        b2 = r.get("b2") or {}
        b3 = r.get("b3") or {}
        b5 = r.get("b5") or {}
        meta = r.get("meta") or {}
        exact = b1.get("exact_count") or 0
        partial = b1.get("partial_count") or 0
        fab = b1.get("fabricated_count") or 0
        ann_exact += exact
        ann_partial += partial
        ann_fab += fab
        total_ann = exact + partial + fab
        exact_rate = (exact / total_ann) if total_ann else None
        partial_rate = (partial / total_ann) if total_ann else None
        fab_rate = (fab / total_ann) if total_ann else None
        if exact_rate is not None:
            b1_exact_rates.append(exact_rate)
            b1_partial_rates.append(partial_rate or 0.0)
            b1_fab_rates.append(fab_rate or 0.0)
        for src, dst in [
            (b2.get("t_rate"), b2_t),
            (b2.get("c_rate"), b2_c),
            (b2.get("a_rate"), b2_a),
            (b5.get("evidence_basis_t_rate"), b5_t),
            (b5.get("evidence_basis_c_rate"), b5_c),
            (b5.get("evidence_basis_a_rate"), b5_a),
        ]:
            if src is not None:
                dst.append(float(src))

        aps = b2.get("attention_points") or []
        if isinstance(aps, list):
            ap_total += len(aps)
            if aps:
                papers_with_ap += 1
        elif isinstance(aps, int):
            ap_total += aps
            if aps:
                papers_with_ap += 1

        review_t = []
        review_v = []
        review_gv = []
        revs = b3.get("reviews") or []
        if revs:
            papers_with_b3 += 1
        for rev in revs:
            review_total += 1
            finding_total += rev.get("total_findings") or len(rev.get("findings") or [])
            if rev.get("b3_truthfulness_rate") is not None:
                review_t.append(float(rev["b3_truthfulness_rate"]))
            if rev.get("b3_validity_rate") is not None:
                review_v.append(float(rev["b3_validity_rate"]))
            if rev.get("b3_grounded_valid_rate") is not None:
                review_gv.append(float(rev["b3_grounded_valid_rate"]))
        paper_b3_t = mean(review_t) if review_t else None
        paper_b3_v = mean(review_v) if review_v else None
        paper_b3_gv = mean(review_gv) if review_gv else None
        if paper_b3_t is not None:
            b3_t.append(paper_b3_t)
        if paper_b3_v is not None:
            b3_v.append(paper_b3_v)
        if paper_b3_gv is not None:
            b3_gv.append(paper_b3_gv)

        claims = b5.get("claims") or []
        eb_n = 0
        if isinstance(claims, list):
            for c in claims:
                if c.get("claim_type") == "evidence_basis":
                    eb_n += 1
        claim_eb_total += eb_n
        if b5.get("evidence_basis_t_rate") is not None:
            papers_with_b5 += 1

        if b5.get("high_risk"):
            high_risk += 1
        if meta.get("clean_view_eligible"):
            clean += 1
        if meta.get("processing_time_sec") is not None:
            times.append(float(meta["processing_time_sec"]))

        flat.append(
            {
                "paper_id": r.get("paper_id"),
                "b1_exact_count": exact,
                "b1_partial_count": partial,
                "b1_fabricated_count": fab,
                "b1_exact_rate": round(exact_rate, 4) if exact_rate is not None else None,
                "b1_partial_rate": round(partial_rate, 4) if partial_rate is not None else None,
                "b1_fabricated_rate": round(fab_rate, 4) if fab_rate is not None else None,
                "b2_t_rate": b2.get("t_rate"),
                "b2_c_rate": b2.get("c_rate"),
                "b2_a_rate": b2.get("a_rate"),
                "b2_t_rate_technical": b2.get("t_rate_technical"),
                "b2_administrative_ap_count": b2.get("administrative_ap_count"),
                "b3_truthfulness_mean": round(paper_b3_t, 4) if paper_b3_t is not None else None,
                "b3_validity_mean": round(paper_b3_v, 4) if paper_b3_v is not None else None,
                "b3_grounded_valid_mean": round(paper_b3_gv, 4) if paper_b3_gv is not None else None,
                "b3_review_count": len(b3.get("reviews") or []),
                "b5_evidence_basis_t_rate": b5.get("evidence_basis_t_rate"),
                "b5_evidence_basis_c_rate": b5.get("evidence_basis_c_rate"),
                "b5_evidence_basis_a_rate": b5.get("evidence_basis_a_rate"),
                "b5_high_risk": bool(b5.get("high_risk")),
                "clean_view_eligible": bool(meta.get("clean_view_eligible")),
                "processing_time_sec": meta.get("processing_time_sec"),
                "nli_pairs_total": meta.get("nli_pairs_total"),
                "pdf_extracted": meta.get("pdf_extracted"),
            }
        )

    papers = pd.DataFrame(flat)
    ann_total = ann_exact + ann_partial + ann_fab
    t_mean = safe_mean(b2_t)
    c_mean = safe_mean(b2_c)
    a_mean = safe_mean(b2_a)
    b3_t_mean = safe_mean(b3_t)
    b3_v_mean = safe_mean(b3_v)
    b3_gv_mean = safe_mean(b3_gv)
    b5_t_mean = safe_mean(b5_t)
    b5_c_mean = safe_mean(b5_c)
    b5_a_mean = safe_mean(b5_a)

    def est(rate: float | None, total: int) -> int | None:
        if rate is None or total <= 0:
            return None
        return int(round(float(rate) * total))

    summary = {
        "workflow": "tca_benchmark",
        "paper_count": len(papers),
        "b1_exact_rate_mean": safe_mean(b1_exact_rates),
        "b1_partial_rate_mean": safe_mean(b1_partial_rates),
        "b1_fabricated_rate_mean": safe_mean(b1_fab_rates),
        "b1_exact_count": ann_exact,
        "b1_partial_count": ann_partial,
        "b1_fabricated_count": ann_fab,
        "b1_annotation_total": ann_total,
        "b1_paper_count": len(b1_exact_rates),
        "b2_t_rate_mean": t_mean,
        "b2_c_rate_mean": c_mean,
        "b2_a_rate_mean": a_mean,
        "b2_attention_point_total": ap_total,
        "b2_paper_count": papers_with_ap,
        "b2_t_count_est": est(t_mean, ap_total),
        "b2_c_count_est": est(c_mean, ap_total),
        "b2_a_count_est": est(a_mean, ap_total),
        "b3_truthfulness_mean": b3_t_mean,
        "b3_validity_mean": b3_v_mean,
        "b3_grounded_valid_mean": b3_gv_mean,
        "b3_review_total": review_total,
        "b3_finding_total": finding_total,
        "b3_paper_count": papers_with_b3,
        "b3_t_count_est": est(b3_t_mean, finding_total),
        "b3_v_count_est": est(b3_v_mean, finding_total),
        "b3_gv_count_est": est(b3_gv_mean, finding_total),
        "b5_evidence_basis_t_rate_mean": b5_t_mean,
        "b5_evidence_basis_c_rate_mean": b5_c_mean,
        "b5_evidence_basis_a_rate_mean": b5_a_mean,
        "b5_claim_total": claim_eb_total,
        "b5_paper_count": papers_with_b5,
        "b5_t_count_est": est(b5_t_mean, claim_eb_total),
        "b5_c_count_est": est(b5_c_mean, claim_eb_total),
        "b5_a_count_est": est(b5_a_mean, claim_eb_total),
        "high_risk_papers": high_risk,
        "clean_view_eligible_papers": clean,
        "processing_time_sec_mean": safe_mean(times),
        "processing_time_sec_median": safe_median(times),
        "source_file": "tca_benchmark_results (5).jsonl",
        "note": (
            "annotation grounding; attention points (truthfulness/coverage/additionality); "
            "review-quality findings; chair evidence-basis claims"
        ),
    }
    return papers, summary


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------


def _bar(
    ax,
    labels: list[str],
    values: list[float],
    colors: list[str] | str,
    ylabel: str,
    title: str,
    value_fmt: str = "{:.1f}",
    rotate: int = 20,
) -> None:
    x = range(len(labels))
    bars = ax.bar(x, values, color=colors, edgecolor="white", linewidth=0.6)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=rotate, ha="right", fontsize=9)
    ax.set_ylabel(ylabel, fontsize=10)
    ax.set_title(title, fontsize=12, fontweight="bold", color=PALETTE["ink"], pad=10)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            value_fmt.format(val),
            ha="center",
            va="bottom",
            fontsize=8,
            color=PALETTE["slate"],
        )


def save_fig(fig, name: str) -> Path:
    path = FIGURES / name
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def chart_overview_case_counts(
    *,
    runner_n: int,
    tca_n: int,
    track_n: int,
    rule_n: int,
    llm_n: int,
    chat_n: int,
) -> Path:
    """Sample-size overview aligned with Chapter 4 dual corpora.

    1.127 is the shared workflow-runner corpus (performance, tokens, outputs),
    not Submission Autofill alone. 1.097 is the TCA scoring subset.
    """
    _setup_vietnamese_font()
    labels = [
        "Bộ thực thi\nluồng AI",
        "Bộ đánh giá\nTCA",
        "Gợi ý\nchuyên đề",
        "Gating\n(luật)",
        "Gating\n(nội dung)",
        "Chatbot\nAgent",
    ]
    values = [runner_n, tca_n, track_n, rule_n, llm_n, chat_n]
    colors = [
        PALETTE["coral"],
        PALETTE["navy"],
        PALETTE["teal"],
        PALETTE["mint"],
        PALETTE["amber"],
        PALETTE["ink"],
    ]
    fig = plt.figure(figsize=(11.2, 6.2))
    ax = fig.add_axes([0.10, 0.38, 0.86, 0.52])
    x = range(len(labels))
    bars = ax.bar(list(x), values, color=colors, edgecolor="white", linewidth=0.6, width=0.68)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, fontsize=12)
    ax.set_ylabel("Số đơn vị", fontsize=12)
    ax.tick_params(axis="y", labelsize=11)
    ax.set_title(
        "Quy mô các bộ đánh giá luồng AI",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ymax = max(values) * 1.12 if values else 1
    ax.set_ylim(0, ymax)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            _fmt_int(val),
            ha="center",
            va="bottom",
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
        )

    # Note under chart: dual-corpus model (Chapter 4)
    note_lines = [
        "• Bộ thực thi luồng AI (1.127 bài): tập đầu vào chung để chạy các luồng AI, ghi hiệu năng, tài nguyên (thời gian, token) và tạo đầu ra.",
        "• Bộ đánh giá TCA (1.097 gói): đánh giá chung độ chính xác, độ trung thực, độ đáng tin cậy và tiềm năng của các luồng AI.",
        "• Bộ đánh giá chuyên biệt: gợi ý chuyên đề, gating và chatbot — mẫu số riêng, không gộp với 1.127 bài.",
    ]
    y = 0.28
    fig.text(0.10, y, "Chú giải", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    y -= 0.048
    for line in note_lines:
        fig.text(0.10, y, line, fontsize=11, color=PALETTE["slate"])
        y -= 0.042

    path = FIGURES / "fig01_overview_case_counts.png"
    fig.savefig(path, dpi=200, facecolor="white", bbox_inches="tight", pad_inches=0.25)
    plt.close(fig)
    return path


def chart_runner_by_conference(auto_df: pd.DataFrame) -> Path:
    """Distribution of the 1.127-paper runner corpus by conference/track."""
    _setup_vietnamese_font()
    col = "conference_year_track"
    if col not in auto_df.columns:
        raise KeyError(f"Missing column {col!r} in autofill/runner corpus")

    # Normalize raw track labels for report display (match Chapter 4 table)
    display = (
        auto_df[col]
        .fillna("(không xác định)")
        .astype(str)
        .str.replace("Short_Paper_Track", "Short Paper Track", regex=False)
        .str.replace("_", " ", regex=False)
    )
    counts = display.value_counts()
    # Largest at top (barh draws bottom-first)
    counts = counts.sort_values(ascending=True)
    total = int(counts.sum())
    labels = counts.index.tolist()
    values = [int(v) for v in counts.values.tolist()]
    pcts = [100.0 * v / total if total else 0.0 for v in values]

    # Highlight top-4 sources (Chapter 4: ~70.72% of papers)
    top4 = set(counts.sort_values(ascending=False).head(4).index.tolist())
    colors = [PALETTE["coral"] if lab in top4 else PALETTE["navy"] for lab in labels]

    fig = plt.figure(figsize=(11.2, 5.8))
    ax = fig.add_axes([0.30, 0.14, 0.62, 0.74])
    y = range(len(labels))
    ax.barh(list(y), values, color=colors, edgecolor="white", linewidth=0.5, height=0.72)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels, fontsize=12)
    ax.set_xlabel("Số bài", fontsize=12)
    ax.tick_params(axis="x", labelsize=11)
    ax.set_title(
        "Phân bố 1.127 bài đầu vào theo hội nghị / chuyên đề",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
        loc="left",
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.35)
    xmax = max(values) * 1.22 if values else 1
    ax.set_xlim(0, xmax)

    for yi, val, pct in zip(y, values, pcts):
        ax.text(
            val + xmax * 0.015,
            yi,
            f"{_fmt_int(val)}  ({pct:.1f}%)".replace(".", ","),
            va="center",
            ha="left",
            fontsize=11,
            color=PALETTE["ink"],
            fontweight="bold",
        )

    top4_share = sum(v for lab, v in zip(labels, values) if lab in top4)
    top4_pct = 100.0 * top4_share / total if total else 0.0
    fig.text(
        0.30,
        0.04,
        (
            f"Tổng: {_fmt_int(total)} bài  ·  "
            f"Bốn nguồn lớn nhất (cam): {_fmt_int(top4_share)} bài "
            f"({top4_pct:.2f}%)".replace(".", ",")
        ),
        fontsize=11,
        color=PALETTE["slate"],
    )

    path = FIGURES / "fig01b_runner_by_conference.png"
    fig.savefig(path, dpi=200, facecolor="white", bbox_inches="tight", pad_inches=0.25)
    plt.close(fig)
    return path


def chart_gating_rule(rule_df: pd.DataFrame, rule_summary: dict[str, Any]) -> Path:
    labels = ["Verdict\naccuracy", "Rule-ID\nrecall", "False\nblock"]
    values = [
        100.0 * float(rule_summary.get("blocking_verdict_accuracy") or 0),
        100.0 * float(rule_summary.get("rule_id_recall") or 0),
        float(rule_summary.get("false_block_count") or 0),
    ]
    colors = [PALETTE["pass"], PALETTE["mint"], PALETTE["fail"]]
    fig, axes = plt.subplots(1, 2, figsize=(10, 4.2))
    _bar(axes[0], labels, values, colors, "Giá trị", "Submission Gating — Rule check", "{:.0f}", 0)

    verdict_counts = rule_df["actual_verdict"].value_counts()
    order = [v for v in ["pass", "warn", "block"] if v in verdict_counts.index]
    v_colors = {"pass": PALETTE["pass"], "warn": PALETTE["warn"], "block": PALETTE["block"]}
    axes[1].bar(
        order,
        [int(verdict_counts[v]) for v in order],
        color=[v_colors[v] for v in order],
        edgecolor="white",
    )
    axes[1].set_title("Phân bố actual_verdict", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[1].set_ylabel("Số case")
    axes[1].spines["top"].set_visible(False)
    axes[1].spines["right"].set_visible(False)
    for i, v in enumerate(order):
        axes[1].text(i, verdict_counts[v], str(int(verdict_counts[v])), ha="center", va="bottom", fontsize=9)
    return save_fig(fig, "fig02_gating_rule_metrics.png")


def chart_gating_llm(llm_df: pd.DataFrame, llm_summary: dict[str, Any]) -> Path:
    fig, axes = plt.subplots(1, 2, figsize=(10, 4.2))
    verdict_counts = llm_df["verdict"].value_counts()
    order = [v for v in ["pass", "warn", "block"] if v in verdict_counts.index]
    v_colors = {"pass": PALETTE["pass"], "warn": PALETTE["warn"], "block": PALETTE["block"]}
    axes[0].bar(order, [int(verdict_counts[v]) for v in order], color=[v_colors[v] for v in order])
    axes[0].set_title("LLM steering — verdict", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[0].set_ylabel("Số finding/row")
    axes[0].spines["top"].set_visible(False)
    axes[0].spines["right"].set_visible(False)

    contract = int(llm_summary.get("llm_block_contract_violation_count") or 0)
    labels = ["Contract\nOK", "Contract\nviolation"]
    values = [len(llm_df) - contract, contract]
    axes[1].bar(labels, values, color=[PALETTE["pass"], PALETTE["fail"]])
    axes[1].set_title("Không block trái hợp đồng", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[1].set_ylabel("Số row")
    axes[1].spines["top"].set_visible(False)
    axes[1].spines["right"].set_visible(False)
    for i, v in enumerate(values):
        axes[1].text(i, v, str(v), ha="center", va="bottom", fontsize=9)
    return save_fig(fig, "fig03_gating_llm_verdicts.png")


def chart_track_by_conference(track_df: pd.DataFrame) -> Path:
    counts = track_df["conference"].value_counts().sort_values(ascending=True)
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(counts.index.tolist(), counts.values.tolist(), color=PALETTE["teal"])
    ax.set_xlabel("Số case")
    ax.set_title("Track recommendation — phân bố theo hội nghị", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for y, v in enumerate(counts.values.tolist()):
        ax.text(v + 0.1, y, str(v), va="center", fontsize=8, color=PALETTE["slate"])
    return save_fig(fig, "fig04_track_by_conference.png")


def chart_chatbot_latency(by_scenario: pd.DataFrame) -> Path:
    labels = by_scenario["label_vi"].tolist()
    duration = by_scenario["avg_duration_s"].tolist()
    answer = by_scenario["avg_time_to_first_answer_s"].tolist()
    ttft = by_scenario["avg_ttft_s"].tolist()
    x = range(len(labels))
    width = 0.28
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar([i - width for i in x], ttft, width, label="TTFT", color=PALETTE["mint"])
    ax.bar(list(x), answer, width, label="Time to first answer token", color=PALETTE["amber"])
    ax.bar([i + width for i in x], duration, width, label="Total duration", color=PALETTE["navy"])
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=8)
    ax.set_ylabel("Giây")
    ax.set_title("Chatbot — độ trễ trung bình theo kịch bản", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    ax.legend(frameon=False, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    return save_fig(fig, "fig05_chatbot_latency_by_scenario.png")


def chart_chatbot_tool_success(by_scenario: pd.DataFrame) -> Path:
    df = by_scenario.copy()
    # unsupported may have 0 tool calls
    rates = []
    labels = []
    for _, row in df.iterrows():
        labels.append(row["label_vi"])
        rates.append(row["tool_success_rate_pct"] if pd.notna(row["tool_success_rate_pct"]) else 0.0)
    colors = [PALETTE["pass"] if (r or 0) >= 80 else PALETTE["amber"] if (r or 0) >= 60 else PALETTE["fail"] for r in rates]
    fig, ax = plt.subplots(figsize=(10, 4.8))
    _bar(ax, labels, [float(r or 0) for r in rates], colors, "Tool success (%)", "Chatbot — tỷ lệ tool-call thành công", "{:.1f}%", 25)
    ax.set_ylim(0, 110)
    return save_fig(fig, "fig06_chatbot_tool_success.png")


def chart_chatbot_manual_outcomes() -> Path:
    labels = ["Đạt", "Đạt một phần", "Không đạt"]
    values = [
        CHATBOT_MANUAL_OVERALL["pass"],
        CHATBOT_MANUAL_OVERALL["partial"],
        CHATBOT_MANUAL_OVERALL["fail"],
    ]
    colors = [PALETTE["pass"], PALETTE["partial"], PALETTE["fail"]]
    fig, ax = plt.subplots(figsize=(6.5, 4.5))
    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        colors=colors,
        autopct=lambda p: f"{p:.1f}%\n({int(round(p * sum(values) / 100))})",
        startangle=90,
        wedgeprops={"linewidth": 1, "edgecolor": "white"},
        textprops={"fontsize": 10},
    )
    for t in autotexts:
        t.set_fontsize(9)
        t.set_color("white")
        t.set_fontweight("bold")
    ax.set_title(
        f"Chatbot — đánh giá thủ công ({CHATBOT_MANUAL_OVERALL['trials']} trials)",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
    )
    return save_fig(fig, "fig07_chatbot_manual_outcomes.png")


def chart_autofill_quality(autofill_df: pd.DataFrame) -> Path:
    metrics = {
        "Title exact\nmatch": 100.0 * float(autofill_df["title_exact_match"].mean()),
        "Abstract\nROUGE-L": 100.0 * float(autofill_df["abstract_rouge_l"].mean()),
        "Keyword F1": 100.0 * float(autofill_df["keyword_f1"].mean()),
        "Author F1": 100.0 * float(autofill_df["author_f1"].mean()),
    }
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    _bar(
        ax,
        list(metrics.keys()),
        list(metrics.values()),
        [PALETTE["navy"], PALETTE["teal"], PALETTE["mint"], PALETTE["amber"]],
        "Điểm trung bình (%)",
        f"Submission Autofill — chất lượng metadata (n={len(autofill_df)})",
        "{:.1f}%",
        0,
    )
    ax.set_ylim(0, 110)
    return save_fig(fig, "fig08_autofill_quality.png")


def _setup_vietnamese_font() -> None:
    """Prefer a system font that renders Vietnamese diacritics."""
    from matplotlib import font_manager

    candidates = [
        "Segoe UI",
        "Arial",
        "Tahoma",
        "Calibri",
        "DejaVu Sans",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            plt.rcParams["axes.unicode_minus"] = False
            return


def _fmt_int(n: int | float | None) -> str:
    if n is None:
        return "—"
    return f"{int(n):,}".replace(",", ".")


def _draw_workflow_panel(
    *,
    filename: str,
    title: str,
    sample_line: str,
    bars: list[tuple[str, float | None, int | None, str]],
    legend_items: list[tuple[str, str, str]],
) -> Path:
    """One workflow chart: bars left, short legend right (no footnotes / how-to-read)."""
    _setup_vietnamese_font()

    fig = plt.figure(figsize=(12.0, 5.2))
    ax = fig.add_axes([0.07, 0.18, 0.54, 0.66])

    labels = [b[0] for b in bars]
    values = [100.0 * float(b[1] or 0) for b in bars]
    counts = [b[2] for b in bars]
    colors = [PALETTE[b[3]] for b in bars]
    x = range(len(labels))
    bar_rects = ax.bar(list(x), values, color=colors, edgecolor="white", linewidth=0.6, width=0.62)

    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, fontsize=12)
    ax.set_ylabel("Tỷ lệ (%)", fontsize=12)
    ax.set_ylim(0, 120)
    ax.tick_params(axis="y", labelsize=11)
    ax.set_title(title, fontsize=15, fontweight="bold", color=PALETTE["ink"], pad=10, loc="left")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)

    for rect, pct_val, cnt in zip(bar_rects, values, counts):
        count_txt = _fmt_int(cnt)
        label = f"{pct_val:.1f}%\n({count_txt})"
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height() + 1.8,
            label,
            ha="center",
            va="bottom",
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
            linespacing=1.2,
        )

    fig.text(
        0.07,
        0.055,
        sample_line,
        fontsize=12,
        color=PALETTE["slate"],
        transform=fig.transFigure,
    )

    # Right-side legend: short term + one-line gloss only
    legend_ax = fig.add_axes([0.65, 0.18, 0.32, 0.66])
    legend_ax.set_xlim(0, 1)
    legend_ax.set_ylim(0, 1)
    legend_ax.axis("off")
    legend_ax.add_patch(
        plt.Rectangle(
            (0.0, 0.0),
            1.0,
            1.0,
            transform=legend_ax.transAxes,
            facecolor="#F7F5F2",
            edgecolor="#D6D0C8",
            linewidth=1.0,
            zorder=0,
        )
    )
    legend_ax.text(
        0.08,
        0.90,
        "Chú giải",
        fontsize=13,
        fontweight="bold",
        color=PALETTE["ink"],
        va="top",
        transform=legend_ax.transAxes,
    )

    n_items = max(len(legend_items), 1)
    y = 0.76
    step = 0.62 / n_items
    for term, gloss, color_key in legend_items:
        legend_ax.plot(
            0.10,
            y - 0.01,
            marker="s",
            markersize=12,
            color=PALETTE[color_key],
            transform=legend_ax.transAxes,
            linestyle="None",
            clip_on=False,
        )
        legend_ax.text(
            0.18,
            y + 0.02,
            term,
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
            va="top",
            transform=legend_ax.transAxes,
        )
        wrapped = "\n".join(textwrap.wrap(gloss, width=28))
        legend_ax.text(
            0.18,
            y - 0.055,
            wrapped,
            fontsize=11,
            color=PALETTE["slate"],
            va="top",
            transform=legend_ax.transAxes,
            linespacing=1.25,
        )
        y -= step

    path = FIGURES / filename
    fig.savefig(path, dpi=200, facecolor="white", bbox_inches="tight", pad_inches=0.22)
    plt.close(fig)
    return path


def chart_tca_workflow_panels(tca_summary: dict[str, Any]) -> list[tuple[Path, str, str]]:
    """One figure per workflow aspect; returns (path, content, placement) for index."""
    s = tca_summary
    outputs: list[tuple[Path, str, str]] = []

    # 1) Phân tích ban đầu cho phản biện — annotation
    p = _draw_workflow_panel(
        filename="fig09a_reviewer_annotation_grounding.png",
        title="Phân tích ban đầu cho phản biện — Độ bám nguồn của chú thích",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b1_paper_count'))} bài  ·  "
            f"Tổng số chú thích: {_fmt_int(s.get('b1_annotation_total'))}"
        ),
        bars=[
            ("Khớp gần\nnguyên văn", s.get("b1_exact_rate_mean"), s.get("b1_exact_count"), "navy"),
            ("Khớp\nmột phần", s.get("b1_partial_rate_mean"), s.get("b1_partial_count"), "teal"),
            (
                "Không xác định\nđược trong bài",
                s.get("b1_fabricated_rate_mean"),
                s.get("b1_fabricated_count"),
                "coral",
            ),
        ],
        legend_items=[
            ("Khớp gần nguyên văn", "Trùng gần như nguyên văn với nội dung bài", "navy"),
            ("Khớp một phần", "Có cơ sở trong bài nhưng chỉ khớp một phần", "teal"),
            ("Không xác định được", "Không tìm thấy đoạn tương ứng trong bài", "coral"),
        ],
    )
    outputs.append((p, "Độ bám nguồn chú thích (phân tích ban đầu)", "Phân tích ban đầu cho phản biện"))

    # 2) Phân tích ban đầu — điểm cần chú ý
    p = _draw_workflow_panel(
        filename="fig09b_reviewer_attention_points.png",
        title="Phân tích ban đầu cho phản biện — Điểm cần chú ý khi đọc bài",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b2_paper_count'))} bài  ·  "
            f"Tổng số điểm chú ý: {_fmt_int(s.get('b2_attention_point_total'))}"
        ),
        bars=[
            ("Bám nội dung\nbài", s.get("b2_t_rate_mean"), s.get("b2_t_count_est"), "navy"),
            (
                "Trùng với\nphản biện viên",
                s.get("b2_c_rate_mean"),
                s.get("b2_c_count_est"),
                "teal",
            ),
            (
                "Bổ sung\ngóc nhìn",
                s.get("b2_a_rate_mean"),
                s.get("b2_a_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            ("Bám nội dung bài", "Điểm chú ý có cơ sở trong bài", "navy"),
            ("Trùng với phản biện viên", "Trùng điểm phản biện viên cũng nêu", "teal"),
            ("Bổ sung góc nhìn", "Bổ sung ngoài phần phản biện viên đã nêu", "coral"),
        ],
    )
    outputs.append((p, "Điểm cần chú ý khi đọc bài", "Phân tích ban đầu cho phản biện"))

    # 3) Kiểm tra chất lượng phản biện
    p = _draw_workflow_panel(
        filename="fig09c_review_quality_auditor.png",
        title="Kiểm tra chất lượng phản biện — Cảnh báo trên bản nhận xét",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b3_paper_count'))} bài  ·  "
            f"{_fmt_int(s.get('b3_review_total'))} bản phản biện  ·  "
            f"{_fmt_int(s.get('b3_finding_total'))} cảnh báo"
        ),
        bars=[
            (
                "Bám nguồn\nnhận xét",
                s.get("b3_truthfulness_mean"),
                s.get("b3_t_count_est"),
                "navy",
            ),
            (
                "Hợp lệ tiêu chí\nkiểm tra",
                s.get("b3_validity_mean"),
                s.get("b3_v_count_est"),
                "teal",
            ),
            (
                "Hợp lệ và\ncó chứng cứ",
                s.get("b3_grounded_valid_mean"),
                s.get("b3_gv_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            ("Bám nguồn nhận xét", "Cảnh báo bám đúng nội dung nhận xét hoặc bài", "navy"),
            ("Hợp lệ tiêu chí kiểm tra", "Cảnh báo thuộc loại vấn đề kiểm tra hợp lệ", "teal"),
            ("Hợp lệ và có chứng cứ", "Vừa hợp lệ vừa có bằng chứng rõ", "coral"),
        ],
    )
    outputs.append((p, "Cảnh báo kiểm tra chất lượng phản biện", "Kiểm tra chất lượng phản biện"))

    # 4) Hỗ trợ quyết định chủ tịch
    p = _draw_workflow_panel(
        filename="fig09d_chair_evidence_basis.png",
        title="Hỗ trợ quyết định chủ tịch — Cơ sở bằng chứng",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b5_paper_count'))} bài  ·  "
            f"Tổng số nhận định: {_fmt_int(s.get('b5_claim_total'))}"
        ),
        bars=[
            (
                "Bám nguồn\ndữ liệu",
                s.get("b5_evidence_basis_t_rate_mean"),
                s.get("b5_t_count_est"),
                "navy",
            ),
            (
                "Trùng điểm\ncon người",
                s.get("b5_evidence_basis_c_rate_mean"),
                s.get("b5_c_count_est"),
                "teal",
            ),
            (
                "Bổ sung\ngóc nhìn",
                s.get("b5_evidence_basis_a_rate_mean"),
                s.get("b5_a_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            ("Bám nguồn dữ liệu", "Nhận định bám dữ liệu nguồn có sẵn", "navy"),
            ("Trùng điểm con người", "Trùng với điểm con người nhấn mạnh", "teal"),
            ("Bổ sung góc nhìn", "Bổ sung ngoài phần con người đã nêu", "coral"),
        ],
    )
    outputs.append((p, "Cơ sở bằng chứng hỗ trợ chủ tịch", "Hỗ trợ quyết định chủ tịch"))

    return outputs


def chart_workflow_headline(summaries: dict[str, dict[str, Any]]) -> Path:
    """Single slide-style headline for Chapter 5."""
    labels = [
        "Gating rule\nverdict acc.",
        "Gating LLM\ncontract OK",
        "Chatbot\nmanual pass",
        "Chatbot\ntool success",
        "Autofill\nROUGE-L",
        "Annotation\nkhớp nguyên văn",
    ]
    rule = summaries["rule"]
    llm = summaries["llm"]
    chat = summaries["chatbot"]
    auto = summaries["autofill"]
    tca = summaries["tca"]
    values = [
        100.0 * float(rule.get("blocking_verdict_accuracy") or 0),
        100.0
        * (
            1.0
            - float(llm.get("llm_block_contract_violation_count") or 0)
            / max(int(llm.get("review_rows") or 1), 1)
        ),
        float(chat.get("manual_pass_rate_pct") or 0),
        100.0 * float(chat.get("tool_call_success_rate") or 0),
        100.0 * float(auto.get("abstract_rouge_l_mean") or 0),
        100.0 * float(tca.get("b1_exact_rate_mean") or 0),
    ]
    colors = [PALETTE["pass"], PALETTE["mint"], PALETTE["amber"], PALETTE["teal"], PALETTE["navy"], PALETTE["ink"]]
    fig, ax = plt.subplots(figsize=(10, 4.8))
    _bar(ax, labels, values, colors, "Phần trăm", "Tóm tắt chỉ số headline cho báo cáo", "{:.1f}%", 0)
    ax.set_ylim(0, 115)
    return save_fig(fig, "fig10_headline_metrics.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def build_overview_sheet(
    track_s: dict[str, Any],
    rule_s: dict[str, Any],
    llm_s: dict[str, Any],
    chat_s: dict[str, Any],
    auto_s: dict[str, Any],
    tca_s: dict[str, Any],
) -> pd.DataFrame:
    rows = [
        {
            "workflow": "track_recommendation",
            "n": track_s["case_count"],
            "primary_metric": "invalid_track_rate",
            "primary_value": track_s["invalid_track_rate"],
            "secondary_metric": "human_label_filled",
            "secondary_value": track_s["human_label_filled"],
            "status": "completed_no_human_accuracy",
            "limitation": track_s["limitation"],
        },
        {
            "workflow": "submission_gating_rule_check",
            "n": rule_s["case_count"],
            "primary_metric": "blocking_verdict_accuracy",
            "primary_value": rule_s["blocking_verdict_accuracy"],
            "secondary_metric": "rule_id_recall",
            "secondary_value": rule_s["rule_id_recall"],
            "status": "completed",
            "limitation": "",
        },
        {
            "workflow": "submission_gating_llm_steering",
            "n": llm_s["case_count"],
            "primary_metric": "llm_block_contract_violation_count",
            "primary_value": llm_s["llm_block_contract_violation_count"],
            "secondary_metric": "content_finding_count",
            "secondary_value": llm_s["content_finding_count"],
            "status": "completed_no_human_labels",
            "limitation": llm_s["limitation"],
        },
        {
            "workflow": "chatbot_agent",
            "n": chat_s["trial_count"],
            "primary_metric": "manual_pass_rate_pct",
            "primary_value": chat_s["manual_pass_rate_pct"],
            "secondary_metric": "tool_call_success_rate",
            "secondary_value": chat_s["tool_call_success_rate"],
            "status": "completed_with_manual_review",
            "limitation": "manual labels from report, not auto-scored quality",
        },
        {
            "workflow": "submission_autofill",
            "n": auto_s["case_count"],
            "primary_metric": "abstract_rouge_l_mean",
            "primary_value": auto_s["abstract_rouge_l_mean"],
            "secondary_metric": "title_exact_match_rate",
            "secondary_value": auto_s["title_exact_match_rate"],
            "status": "completed",
            "limitation": "",
        },
        {
            "workflow": "tca_benchmark",
            "n": tca_s["paper_count"],
            "primary_metric": "annotation_exact_match_rate",
            "primary_value": tca_s["b1_exact_rate_mean"],
            "secondary_metric": "auditor_grounded_valid_rate",
            "secondary_value": tca_s["b3_grounded_valid_mean"],
            "status": "completed",
            "limitation": tca_s.get("note", ""),
        },
    ]
    return pd.DataFrame(rows)


def write_figure_index(paths: list[tuple[str, str, str]]) -> None:
    lines = [
        "# Figure index — workflow benchmark visualizations",
        "",
        "Sinh tự động từ `scripts/export_benchmark_to_excel.py`.",
        "Dùng PNG trong `exports/figures/` cho Chương 5 (đánh giá).",
        "",
        "| File | Nội dung | Gợi ý chỗ dùng |",
        "| --- | --- | --- |",
    ]
    for fname, content, placement in paths:
        lines.append(f"| `{fname}` | {content} | {placement} |")
    lines.extend(
        [
            "",
            "## Excel workbook",
            "",
            "- `workflow_benchmark_results.xlsx`",
            "  - `Overview` — chỉ số headline theo workflow",
            "  - `Track_Recommendation` — 48 case predictions",
            "  - `Gating_Rule_Check` — 8 deterministic cases",
            "  - `Gating_LLM_Steering` — LLM content findings",
            "  - `Chatbot_Trials` — 40 trial transport metrics",
            "  - `Chatbot_By_Scenario` — gộp theo kịch bản + manual outcome",
            "  - `Autofill_Summary` + `Autofill_Cases` — completed CSV metrics",
            "  - `TCA_Summary` + `TCA_Papers` — TCA rates per paper",
            "",
            "## Caveats (bắt buộc khi trích vào báo cáo)",
            "",
            "1. Track recommendation: `human_label` trống → **không** claim top-1 accuracy từ file này.",
            "2. LLM steering: `grounded/actionable` trống → chỉ claim operational contract (0 block violation).",
            "3. Chatbot: transport metrics từ `run_summary.json`; manual pass/partial/fail từ report review.",
            "4. TCA/fig09a–d: mỗi workflow một figure; nhãn = % + (số lượng); cỡ mẫu dưới trục; chú giải bên phải.",
            "5. Annotation (fig09a): số lượng là đếm tuyệt đối. Các figure còn lại: số trong ngoặc ≈ tỷ lệ × tổng đơn vị.",
            "",
        ]
    )
    (EXPORT / "FIGURE_INDEX.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    EXPORT.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)

    track_df, track_s = load_track()
    rule_df, rule_s = load_gating_rule()
    llm_df, llm_s = load_gating_llm()
    chat_trials, chat_by_sc, chat_s = load_chatbot()
    auto_df, auto_s = load_autofill()
    tca_df, tca_s = load_tca()

    overview = build_overview_sheet(track_s, rule_s, llm_s, chat_s, auto_s, tca_s)

    # Compact autofill for Excel (full 1127 rows is fine; keep numeric cols first)
    auto_cols = [
        c
        for c in [
            "paper_id",
            "conference_year_track",
            "title_exact_match",
            "abstract_rouge_1",
            "abstract_rouge_l",
            "keyword_f1",
            "author_f1",
            "time_taken_seconds",
            "total_inference_time",
            "input_tokens",
            "output_tokens",
            "total_tokens",
        ]
        if c in auto_df.columns
    ]
    auto_compact = auto_df[auto_cols].copy()

    xlsx_path = EXPORT / "workflow_benchmark_results.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        write_df(writer, overview, "Overview")
        write_df(writer, pd.DataFrame([track_s]), "Track_Summary")
        write_df(writer, track_df, "Track_Recommendation")
        write_df(writer, pd.DataFrame([rule_s]), "Gating_Rule_Summary")
        write_df(writer, rule_df, "Gating_Rule_Check")
        write_df(writer, pd.DataFrame([llm_s]), "Gating_LLM_Summary")
        write_df(writer, llm_df, "Gating_LLM_Steering")
        write_df(writer, pd.DataFrame([chat_s]), "Chatbot_Summary")
        write_df(writer, chat_by_sc, "Chatbot_By_Scenario")
        write_df(writer, chat_trials, "Chatbot_Trials")
        write_df(writer, pd.DataFrame([auto_s]), "Autofill_Summary")
        write_df(writer, auto_compact, "Autofill_Cases")
        write_df(writer, pd.DataFrame([tca_s]), "TCA_Summary")
        write_df(writer, tca_df, "TCA_Papers")

    # Charts
    fig_meta: list[tuple[str, str, str]] = []
    p = chart_overview_case_counts(
        runner_n=int(auto_s["case_count"]),  # shared runner corpus = 1.127 papers
        tca_n=int(tca_s["paper_count"]),
        track_n=int(track_s["case_count"]),
        rule_n=int(rule_s["case_count"]),
        llm_n=int(llm_s["case_count"]),
        chat_n=int(chat_s["trial_count"]),
    )
    fig_meta.append(
        (
            p.name,
            "Quy mô bộ đánh giá: bộ thực thi 1.127 vs TCA 1.097 vs bộ đánh giá chuyên biệt",
            "Chương 4 — thiết lập đánh giá",
        )
    )

    p = chart_runner_by_conference(auto_df)
    fig_meta.append(
        (
            p.name,
            "Phân bố 1.127 bài đầu vào theo hội nghị / chuyên đề",
            "Chương 4 — thiết lập dữ liệu",
        )
    )

    p = chart_gating_rule(rule_df, rule_s)
    fig_meta.append((p.name, "Accuracy/recall rule check + phân bố verdict", "Submission Gating (deterministic)"))

    p = chart_gating_llm(llm_df, llm_s)
    fig_meta.append((p.name, "LLM steering verdict + contract violation", "Submission Gating (LLM content)"))

    p = chart_track_by_conference(track_df)
    fig_meta.append((p.name, "Track recommendation theo hội nghị", "Track recommendation setup"))

    p = chart_chatbot_latency(chat_by_sc)
    fig_meta.append((p.name, "TTFT / first-answer / total duration theo kịch bản", "Chatbot Agent — latency"))

    p = chart_chatbot_tool_success(chat_by_sc)
    fig_meta.append((p.name, "Tool-call success rate theo kịch bản", "Chatbot Agent — tool reliability"))

    p = chart_chatbot_manual_outcomes()
    fig_meta.append((p.name, "Manual pass / partial / fail (40 trials)", "Chatbot Agent — quality"))

    p = chart_autofill_quality(auto_df)
    fig_meta.append((p.name, "Title/abstract/keyword/author quality", "Submission Autofill"))

    for path, content, placement in chart_tca_workflow_panels(tca_s):
        fig_meta.append((path.name, content, placement))

    p = chart_workflow_headline(
        {"rule": rule_s, "llm": llm_s, "chatbot": chat_s, "autofill": auto_s, "tca": tca_s}
    )
    fig_meta.append((p.name, "Headline metrics cho slide/báo cáo", "Chương 5 — tóm tắt kết quả"))

    write_figure_index(fig_meta)

    print("Wrote:", xlsx_path)
    print("Figures:", FIGURES)
    print("Sheets:", "Overview + 13 detail sheets")
    print("Charts:", len(fig_meta))
    print(
        "Headline:",
        {
            "gating_rule_acc": rule_s["blocking_verdict_accuracy"],
            "llm_contract_violations": llm_s["llm_block_contract_violation_count"],
            "chatbot_manual_pass_pct": chat_s["manual_pass_rate_pct"],
            "chatbot_tool_success": chat_s["tool_call_success_rate"],
            "autofill_rouge_l": auto_s["abstract_rouge_l_mean"],
            "tca_b1_exact": tca_s["b1_exact_rate_mean"],
            "track_invalid_rate": track_s["invalid_track_rate"],
            "track_human_labels": track_s["human_label_filled"],
            "tca_ann_total": tca_s.get("b1_annotation_total"),
            "tca_ap_total": tca_s.get("b2_attention_point_total"),
            "tca_findings": tca_s.get("b3_finding_total"),
            "tca_claims": tca_s.get("b5_claim_total"),
        },
    )


if __name__ == "__main__":
    main()
