#!/usr/bin/env python3
"""Export deterministic-workflow quality benchmarks for thesis Chapter 4.

Source of truth (copied into this package):
  benchmark_output/quality-results.csv
  benchmark_output/quality-results.md
  ranking_case_counts.xlsx   (case-level Hit@k counts, already in package)

Backend provenance:
  backend/benchmarks/quality/results/quality-results.{csv,md}
  backend/benchmarks/quality/  (ranking + assignment + COI quality suite)

Writes:
  exports/deterministic_workflow_results.xlsx
  exports/figures/*.png
  exports/FIGURE_INDEX.md
  exports/SUMMARY.md
  exports/summary_metrics.json
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "benchmark_output"
EXPORT = ROOT / "exports"
FIGURES = EXPORT / "figures"

QUALITY_CSV = INPUT / "quality-results.csv"
QUALITY_MD = INPUT / "quality-results.md"
CASE_COUNTS_XLSX = ROOT / "ranking_case_counts.xlsx"

# Same palette as ai_workflow_benchmarks / system_performance for visual consistency.
PALETTE = {
    "navy": "#1B3A4B",
    "teal": "#0D7377",
    "coral": "#C44536",
    "amber": "#E09F3E",
    "slate": "#4A5568",
    "mint": "#2A9D8F",
    "sand": "#F4A261",
    "ink": "#264653",
    "pass": "#2A9D8F",
    "partial": "#E09F3E",
    "fail": "#C44536",
    "jaccard": "#0D7377",
    "overlap": "#E09F3E",
    "random": "#4A5568",
    "greedy": "#1B3A4B",
    "round_robin": "#2A9D8F",
}

METHOD_VI = {
    "jaccard": "Jaccard (production)",
    "overlap_count": "Overlap count (baseline)",
    "random": "Random (baseline)",
    "greedy": "Greedy (production)",
    "round_robin": "Round-robin (baseline)",
}

METHOD_ROLE = {
    "jaccard": "production",
    "overlap_count": "baseline",
    "random": "baseline",
    "greedy": "production",
    "round_robin": "baseline",
}

METHOD_COLOR = {
    "jaccard": PALETTE["jaccard"],
    "overlap_count": PALETTE["overlap"],
    "random": PALETTE["random"],
    "greedy": PALETTE["greedy"],
    "round_robin": PALETTE["mint"],
}

RUN_META = {
    "run_id": "matching-quality-benchmark",
    "package": "deterministic_workflow",
    "source_csv": "backend/benchmarks/quality/results/quality-results.csv",
    "source_md": "backend/benchmarks/quality/results/quality-results.md",
    "source_case_counts": "docs/report/statistics/deterministic_workflow/ranking_case_counts.xlsx",
    "suite_path": "backend/benchmarks/quality/",
    "algorithm_family": "deterministic (Jaccard + greedy assignment + COI hard constraint)",
    "proxy_ground_truth": "leave-one-out authorship (author is relevant reviewer for own paper topics)",
    "note": "Quality (not speed). Speed lives in system_performance / backend micro+k6.",
}


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1B3A4B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:100]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str) -> None:
    out = df.copy()
    if out.empty:
        out = pd.DataFrame({"note": ["(empty)"]})
    sheet = sheet[:31]
    out.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.book[sheet]
    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "A2"
    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions


def apply_style() -> None:
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "axes.titlesize": 12,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": PALETTE["slate"],
            "axes.grid": False,
        }
    )


def save_fig(fig, name: str, dpi: int = 200) -> Path:
    path = FIGURES / name
    fig.savefig(path, dpi=dpi, bbox_inches="tight", facecolor="white", pad_inches=0.25)
    plt.close(fig)
    return path


def _finish_ax(ax, title: str, ylabel: str | None = None) -> None:
    ax.set_title(title, fontsize=12, fontweight="bold", color=PALETTE["ink"], pad=10)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=10, color=PALETTE["slate"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.tick_params(colors=PALETTE["slate"])


def _label_bars(ax, bars, fmt: str = "{:.0%}", dy: float = 0.015, fontsize: float = 9) -> None:
    for bar in bars:
        h = bar.get_height()
        if np.isnan(h):
            continue
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            h + dy,
            fmt.format(h) if isinstance(h, float) and h <= 1.5 and "%" in fmt else (
                f"{h:g}" if not isinstance(h, float) else fmt.format(h)
            ),
            ha="center",
            va="bottom",
            fontsize=fontsize,
            color=PALETTE["ink"],
            fontweight="bold",
        )


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def parse_fixture_from_md(path: Path) -> dict[str, Any]:
    """Parse fixture stats from quality-results.md header."""
    text = path.read_text(encoding="utf-8") if path.exists() else ""
    stats: dict[str, Any] = {
        "authors": None,
        "papers": None,
        "reviewers": None,
        "loo_queries": None,
        "topic_vocab": None,
        "mean_topics_per_paper": None,
    }
    # "- Authors: 60 | Papers: 2565 | Reviewers: 60 | LOO queries: 60"
    m = re.search(
        r"Authors:\s*(\d+)\s*\|\s*Papers:\s*(\d+)\s*\|\s*Reviewers:\s*(\d+)\s*\|\s*LOO queries:\s*(\d+)",
        text,
    )
    if m:
        stats["authors"] = int(m.group(1))
        stats["papers"] = int(m.group(2))
        stats["reviewers"] = int(m.group(3))
        stats["loo_queries"] = int(m.group(4))
    m2 = re.search(r"Topic vocabulary:\s*(\d+)\s*\|\s*Mean topics/paper:\s*([0-9.]+)", text)
    if m2:
        stats["topic_vocab"] = int(m2.group(1))
        stats["mean_topics_per_paper"] = float(m2.group(2))
    return stats


def load_quality_long() -> pd.DataFrame:
    if not QUALITY_CSV.exists():
        raise FileNotFoundError(f"Missing quality CSV: {QUALITY_CSV}")
    df = pd.read_csv(QUALITY_CSV)
    expected = {"section", "method", "metric", "value"}
    if not expected.issubset(set(df.columns)):
        raise ValueError(f"quality-results.csv missing columns: {expected - set(df.columns)}")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df["method_vi"] = df["method"].map(METHOD_VI).fillna(df["method"])
    df["role"] = df["method"].map(METHOD_ROLE).fillna("other")
    return df


def load_case_counts() -> dict[str, pd.DataFrame]:
    """Load pre-built case-count workbook (Hit@k correct/incorrect counts)."""
    if not CASE_COUNTS_XLSX.exists():
        return {}
    xls = pd.ExcelFile(CASE_COUNTS_XLSX)
    out: dict[str, pd.DataFrame] = {}
    for sheet in xls.sheet_names:
        out[sheet] = pd.read_excel(CASE_COUNTS_XLSX, sheet_name=sheet)
    return out


def ranking_wide(long_df: pd.DataFrame, fixture: dict[str, Any]) -> pd.DataFrame:
    sub = long_df[long_df["section"] == "ranking"].copy()
    if sub.empty:
        return pd.DataFrame()
    wide = sub.pivot(index="method", columns="metric", values="value").reset_index()
    wide.columns.name = None
    for col in ["hit_at_1", "hit_at_5", "hit_at_10", "mrr", "ndcg_at_10"]:
        if col not in wide.columns:
            wide[col] = np.nan
    n = fixture.get("loo_queries") or 60
    wide["n_queries"] = n
    wide["method_vi"] = wide["method"].map(METHOD_VI).fillna(wide["method"])
    wide["role"] = wide["method"].map(METHOD_ROLE).fillna("other")
    # Case counts from rates (rounded) — keep consistent with ranking_case_counts.xlsx
    for k in (1, 5, 10):
        col = f"hit_at_{k}"
        wide[f"correct_hit@{k}"] = (wide[col] * n).round().astype(int)
        wide[f"incorrect_hit@{k}"] = n - wide[f"correct_hit@{k}"]
    order = {"jaccard": 0, "overlap_count": 1, "random": 2}
    wide["_ord"] = wide["method"].map(order).fillna(99)
    wide = wide.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)
    cols = [
        "method",
        "method_vi",
        "role",
        "n_queries",
        "hit_at_1",
        "hit_at_5",
        "hit_at_10",
        "mrr",
        "ndcg_at_10",
        "correct_hit@1",
        "incorrect_hit@1",
        "correct_hit@5",
        "incorrect_hit@5",
        "correct_hit@10",
        "incorrect_hit@10",
    ]
    return wide[[c for c in cols if c in wide.columns]]


def assignment_wide(long_df: pd.DataFrame) -> pd.DataFrame:
    sub = long_df[long_df["section"] == "assignment"].copy()
    if sub.empty:
        return pd.DataFrame()
    wide = sub.pivot(index="method", columns="metric", values="value").reset_index()
    wide.columns.name = None
    for col in [
        "coverage",
        "load_stddev",
        "load_gini",
        "coi_violations",
        "mean_score",
        "min_score",
        "fallback_rate",
    ]:
        if col not in wide.columns:
            wide[col] = np.nan
    wide["method_vi"] = wide["method"].map(METHOD_VI).fillna(wide["method"])
    wide["role"] = wide["method"].map(METHOD_ROLE).fillna("other")
    wide["coi_clean"] = wide["coi_violations"].fillna(0).astype(int) == 0
    order = {"greedy": 0, "round_robin": 1, "random": 2}
    wide["_ord"] = wide["method"].map(order).fillna(99)
    wide = wide.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)
    cols = [
        "method",
        "method_vi",
        "role",
        "coverage",
        "load_stddev",
        "load_gini",
        "coi_violations",
        "coi_clean",
        "mean_score",
        "min_score",
        "fallback_rate",
    ]
    return wide[[c for c in cols if c in wide.columns]]


def build_ranking_case_detail(rank_df: pd.DataFrame) -> pd.DataFrame:
    """Long form: method × Hit@k case counts (for report tables)."""
    rows: list[dict[str, Any]] = []
    for _, r in rank_df.iterrows():
        for k, metric in [(1, "hit_at_1"), (5, "hit_at_5"), (10, "hit_at_10")]:
            correct = int(r.get(f"correct_hit@{k}", 0))
            incorrect = int(r.get(f"incorrect_hit@{k}", 0))
            rate = float(r[metric]) if pd.notna(r[metric]) else float("nan")
            rows.append(
                {
                    "method": r["method"],
                    "method_vi": r["method_vi"],
                    "role": r["role"],
                    "criterion": f"Hit@{k}",
                    "n_queries": int(r["n_queries"]),
                    "correct": correct,
                    "incorrect": incorrect,
                    "rate": rate,
                    "mrr": float(r["mrr"]) if pd.notna(r["mrr"]) else None,
                    "ndcg_at_10": float(r["ndcg_at_10"]) if pd.notna(r["ndcg_at_10"]) else None,
                    "interpretation_vi": (
                        f"{correct}/{int(r['n_queries'])} lần author thật nằm trong top {k}"
                        if k > 1
                        else f"{correct}/{int(r['n_queries'])} lần gợi ý đúng ngay vị trí #1"
                    ),
                }
            )
    return pd.DataFrame(rows)


def build_overview(
    fixture: dict[str, Any],
    rank_df: pd.DataFrame,
    assign_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = [
        {
            "area": "môi trường",
            "metric": "suite",
            "value": "matching quality (deterministic)",
            "note": RUN_META["suite_path"],
        },
        {
            "area": "môi trường",
            "metric": "fixture",
            "value": (
                f"{fixture.get('authors')} authors · {fixture.get('papers')} papers · "
                f"{fixture.get('loo_queries')} LOO queries"
            ),
            "note": (
                f"topic_vocab={fixture.get('topic_vocab')}; "
                f"mean_topics/paper={fixture.get('mean_topics_per_paper')}"
            ),
        },
        {
            "area": "môi trường",
            "metric": "proxy_ground_truth",
            "value": "leave-one-out authorship",
            "note": RUN_META["proxy_ground_truth"],
        },
        {
            "area": "môi trường",
            "metric": "algorithm",
            "value": "Jaccard + greedy two-pass + COI hard filter",
            "note": RUN_META["algorithm_family"],
        },
    ]

    if not rank_df.empty:
        jac = rank_df[rank_df["method"] == "jaccard"]
        rnd = rank_df[rank_df["method"] == "random"]
        if not jac.empty:
            j = jac.iloc[0]
            rows.append(
                {
                    "area": "ranking",
                    "metric": "jaccard_hit@1",
                    "value": f"{j['hit_at_1']:.1%}",
                    "note": f"{int(j['correct_hit@1'])}/{int(j['n_queries'])} queries",
                }
            )
            rows.append(
                {
                    "area": "ranking",
                    "metric": "jaccard_hit@5",
                    "value": f"{j['hit_at_5']:.1%}",
                    "note": f"{int(j['correct_hit@5'])}/{int(j['n_queries'])} queries",
                }
            )
            rows.append(
                {
                    "area": "ranking",
                    "metric": "jaccard_hit@10",
                    "value": f"{j['hit_at_10']:.1%}",
                    "note": f"{int(j['correct_hit@10'])}/{int(j['n_queries'])} queries",
                }
            )
            rows.append(
                {
                    "area": "ranking",
                    "metric": "jaccard_mrr",
                    "value": f"{j['mrr']:.3f}",
                    "note": f"nDCG@10={j['ndcg_at_10']:.3f}",
                }
            )
        if not jac.empty and not rnd.empty:
            lift = float(jac.iloc[0]["mrr"]) / float(rnd.iloc[0]["mrr"]) if float(rnd.iloc[0]["mrr"]) else None
            rows.append(
                {
                    "area": "ranking",
                    "metric": "mrr_lift_vs_random",
                    "value": f"{lift:.1f}×" if lift else "n/a",
                    "note": f"random MRR={rnd.iloc[0]['mrr']:.3f}",
                }
            )

    if not assign_df.empty:
        gre = assign_df[assign_df["method"] == "greedy"]
        if not gre.empty:
            g = gre.iloc[0]
            rows.append(
                {
                    "area": "assignment",
                    "metric": "greedy_coverage",
                    "value": f"{g['coverage']:.1%}",
                    "note": f"fallback_rate={g['fallback_rate']:.1%}",
                }
            )
            rows.append(
                {
                    "area": "assignment",
                    "metric": "greedy_mean_score",
                    "value": f"{g['mean_score']:.4f}",
                    "note": f"min_score={g['min_score']:.4f}",
                }
            )
            rows.append(
                {
                    "area": "assignment",
                    "metric": "greedy_load_gini",
                    "value": f"{g['load_gini']:.4f}",
                    "note": f"load_stddev={g['load_stddev']:.3f}",
                }
            )
        # COI hard constraint across all assigners
        coi_total = int(assign_df["coi_violations"].fillna(0).sum())
        rows.append(
            {
                "area": "coi",
                "metric": "coi_violations_all_methods",
                "value": str(coi_total),
                "note": "hard constraint: must be 0 for production greedy",
            }
        )
        rows.append(
            {
                "area": "coi",
                "metric": "coi_clean_methods",
                "value": f"{int(assign_df['coi_clean'].sum())}/{len(assign_df)}",
                "note": "greedy / round_robin / random all respect COI map",
            }
        )

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------


def chart_fixture_setup(fixture: dict[str, Any]) -> Path:
    labels = [
        "Authors\n(reviewers)",
        "Papers",
        "LOO queries",
        "Topic vocab\n(×100)",
    ]
    vals = [
        fixture.get("authors") or 0,
        fixture.get("papers") or 0,
        fixture.get("loo_queries") or 0,
        (fixture.get("topic_vocab") or 0) / 100.0,
    ]
    colors = [PALETTE["navy"], PALETTE["teal"], PALETTE["mint"], PALETTE["amber"]]
    fig, ax = plt.subplots(figsize=(9.2, 5.0))
    bars = ax.bar(labels, vals, color=colors, edgecolor="white", width=0.62)
    display = [
        str(fixture.get("authors") or 0),
        f"{fixture.get('papers') or 0:,}",
        str(fixture.get("loo_queries") or 0),
        f"{fixture.get('topic_vocab') or 0:,}",
    ]
    for bar, lab in zip(bars, display):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() * 1.02 + max(vals) * 0.01,
            lab,
            ha="center",
            va="bottom",
            fontsize=10,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    _finish_ax(ax, "Quy mô fixture đánh giá chất lượng deterministic", "Giá trị")
    ax.set_ylim(0, max(vals) * 1.25 if max(vals) else 1)
    mean_t = fixture.get("mean_topics_per_paper")
    fig.text(
        0.02,
        -0.02,
        f"Proxy ground truth: leave-one-out authorship · mean topics/paper = {mean_t} · "
        f"thuật toán production: Jaccard + greedy two-pass + COI hard filter",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig01_fixture_setup.png")


def chart_ranking_hit_at_k(rank_df: pd.DataFrame) -> Path:
    metrics = ["hit_at_1", "hit_at_5", "hit_at_10"]
    metric_labels = ["Hit@1", "Hit@5", "Hit@10"]
    methods = list(rank_df["method"])
    x = np.arange(len(metric_labels))
    width = 0.25
    fig, ax = plt.subplots(figsize=(10.2, 5.2))
    for i, method in enumerate(methods):
        row = rank_df[rank_df["method"] == method].iloc[0]
        vals = [float(row[m]) for m in metrics]
        offset = (i - (len(methods) - 1) / 2) * width
        bars = ax.bar(
            x + offset,
            vals,
            width,
            label=METHOD_VI.get(method, method),
            color=METHOD_COLOR.get(method, PALETTE["slate"]),
            edgecolor="white",
        )
        for bar, v in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.015,
                f"{v:.0%}",
                ha="center",
                fontsize=8,
                color=PALETTE["slate"],
            )
    ax.set_xticks(x)
    ax.set_xticklabels(metric_labels)
    ax.set_ylim(0, 1.05)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    _finish_ax(ax, "Reviewer ranking — Hit@k (leave-one-out authorship)", "Tỷ lệ")
    ax.legend(frameon=False, fontsize=8.5, loc="upper left")
    n = int(rank_df.iloc[0]["n_queries"]) if not rank_df.empty else 0
    fig.text(0.02, -0.02, f"n = {n} LOO queries · production = jaccard", fontsize=8.5, color=PALETTE["slate"])
    return save_fig(fig, "fig02_ranking_hit_at_k.png")


def chart_ranking_mrr_ndcg(rank_df: pd.DataFrame) -> Path:
    methods = list(rank_df["method"])
    labels = [METHOD_VI.get(m, m) for m in methods]
    mrr = [float(rank_df[rank_df["method"] == m].iloc[0]["mrr"]) for m in methods]
    ndcg = [float(rank_df[rank_df["method"] == m].iloc[0]["ndcg_at_10"]) for m in methods]
    x = np.arange(len(methods))
    width = 0.36
    fig, ax = plt.subplots(figsize=(10.0, 5.0))
    b1 = ax.bar(x - width / 2, mrr, width, label="MRR", color=PALETTE["teal"], edgecolor="white")
    b2 = ax.bar(x + width / 2, ndcg, width, label="nDCG@10", color=PALETTE["navy"], edgecolor="white")
    for bars in (b1, b2):
        for bar in bars:
            h = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                h + 0.01,
                f"{h:.3f}",
                ha="center",
                fontsize=8.5,
                color=PALETTE["ink"],
            )
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=8, ha="right")
    ax.set_ylim(0, max(mrr + ndcg) * 1.25 if mrr or ndcg else 1)
    _finish_ax(ax, "Reviewer ranking — MRR và nDCG@10", "Giá trị")
    ax.legend(frameon=False, fontsize=9)
    return save_fig(fig, "fig03_ranking_mrr_ndcg.png")


def chart_ranking_case_counts(rank_df: pd.DataFrame) -> Path:
    """Stacked correct/incorrect for production jaccard across Hit@k."""
    jac = rank_df[rank_df["method"] == "jaccard"]
    if jac.empty:
        jac = rank_df.iloc[[0]]
    row = jac.iloc[0]
    ks = [1, 5, 10]
    correct = [int(row[f"correct_hit@{k}"]) for k in ks]
    incorrect = [int(row[f"incorrect_hit@{k}"]) for k in ks]
    labels = [f"Hit@{k}" for k in ks]
    x = np.arange(len(ks))
    fig, ax = plt.subplots(figsize=(9.0, 5.0))
    b1 = ax.bar(x, correct, color=PALETTE["pass"], edgecolor="white", label="Đúng")
    b2 = ax.bar(x, incorrect, bottom=correct, color=PALETTE["fail"], edgecolor="white", label="Sai", alpha=0.85)
    n = int(row["n_queries"])
    for i, (c, inc) in enumerate(zip(correct, incorrect)):
        ax.text(i, c / 2 if c else 1, f"{c}", ha="center", va="center", color="white", fontweight="bold", fontsize=11)
        ax.text(
            i,
            c + inc / 2 if inc else c + 1,
            f"{inc}",
            ha="center",
            va="center",
            color="white",
            fontweight="bold",
            fontsize=10,
        )
        rate = c / n if n else 0
        ax.text(i, n + 1.5, f"{rate:.0%}", ha="center", fontsize=10, fontweight="bold", color=PALETTE["ink"])
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylim(0, n * 1.18)
    _finish_ax(
        ax,
        f"Jaccard (production) — số case đúng/sai theo Hit@k (n={n})",
        "Số case",
    )
    ax.legend(frameon=False, fontsize=9, loc="upper right")
    # silence unused
    _ = (b1, b2)
    return save_fig(fig, "fig04_ranking_case_counts.png")


def chart_ranking_vs_baselines(rank_df: pd.DataFrame) -> Path:
    """Lift of production vs random on key metrics."""
    jac = rank_df[rank_df["method"] == "jaccard"]
    rnd = rank_df[rank_df["method"] == "random"]
    ovl = rank_df[rank_df["method"] == "overlap_count"]
    if jac.empty or rnd.empty:
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.text(0.5, 0.5, "Missing jaccard/random rows", ha="center")
        return save_fig(fig, "fig05_ranking_vs_baselines.png")

    metrics = ["hit_at_1", "hit_at_5", "hit_at_10", "mrr", "ndcg_at_10"]
    metric_labels = ["Hit@1", "Hit@5", "Hit@10", "MRR", "nDCG@10"]
    j = jac.iloc[0]
    r = rnd.iloc[0]
    o = ovl.iloc[0] if not ovl.empty else None

    x = np.arange(len(metrics))
    width = 0.28
    fig, ax = plt.subplots(figsize=(11.0, 5.2))
    series = [
        ("jaccard", [float(j[m]) for m in metrics], PALETTE["jaccard"]),
    ]
    if o is not None:
        series.append(("overlap_count", [float(o[m]) for m in metrics], PALETTE["overlap"]))
    series.append(("random", [float(r[m]) for m in metrics], PALETTE["random"]))

    for i, (name, vals, color) in enumerate(series):
        offset = (i - (len(series) - 1) / 2) * width
        bars = ax.bar(
            x + offset,
            vals,
            width,
            label=METHOD_VI.get(name, name),
            color=color,
            edgecolor="white",
        )
        for bar, v in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.012,
                f"{v:.2f}" if v < 1 else f"{v:g}",
                ha="center",
                fontsize=7.5,
                color=PALETTE["slate"],
            )

    ax.set_xticks(x)
    ax.set_xticklabels(metric_labels)
    ax.set_ylim(0, 1.08)
    _finish_ax(ax, "Production Jaccard so với baseline (overlap_count, random)", "Giá trị")
    ax.legend(frameon=False, fontsize=8.5, ncol=3, loc="upper left")
    return save_fig(fig, "fig05_ranking_vs_baselines.png")


def chart_assignment_coverage_score(assign_df: pd.DataFrame) -> Path:
    methods = list(assign_df["method"])
    labels = [METHOD_VI.get(m, m) for m in methods]
    coverage = [float(assign_df[assign_df["method"] == m].iloc[0]["coverage"]) for m in methods]
    mean_score = [float(assign_df[assign_df["method"] == m].iloc[0]["mean_score"]) for m in methods]
    # Normalize mean_score to 0–1 display relative to max for dual readability
    x = np.arange(len(methods))
    width = 0.36
    fig, ax1 = plt.subplots(figsize=(10.2, 5.2))
    b1 = ax1.bar(
        x - width / 2,
        coverage,
        width,
        label="Coverage (≥ min reviewers)",
        color=PALETTE["navy"],
        edgecolor="white",
    )
    ax1.set_ylim(0, 1.15)
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    ax1.set_ylabel("Coverage", color=PALETTE["navy"])
    for bar, v in zip(b1, coverage):
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.02,
            f"{v:.0%}",
            ha="center",
            fontsize=9,
            fontweight="bold",
            color=PALETTE["ink"],
        )

    ax2 = ax1.twinx()
    b2 = ax2.bar(
        x + width / 2,
        mean_score,
        width,
        label="Mean assigned-pair score",
        color=PALETTE["teal"],
        edgecolor="white",
    )
    ax2.set_ylabel("Mean score (Jaccard)", color=PALETTE["teal"])
    ymax = max(mean_score) * 1.45 if max(mean_score) > 0 else 0.02
    ax2.set_ylim(0, ymax)
    for bar, v in zip(b2, mean_score):
        ax2.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + ymax * 0.03,
            f"{v:.4f}",
            ha="center",
            fontsize=8.5,
            color=PALETTE["ink"],
        )

    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=8, ha="right")
    ax1.set_title(
        "Assignment — coverage vs mean relevance score",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
    )
    ax1.spines["top"].set_visible(False)
    ax2.spines["top"].set_visible(False)
    ax1.grid(axis="y", linestyle="--", alpha=0.35)
    # Combined legend
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, frameon=False, fontsize=8.5, loc="upper right")
    fig.text(
        0.02,
        -0.02,
        "Greedy ưu tiên điểm phù hợp; baselines đạt coverage 100% nhưng mean score ~3–4× thấp hơn.",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig06_assignment_coverage_score.png")


def chart_assignment_load_balance(assign_df: pd.DataFrame) -> Path:
    methods = list(assign_df["method"])
    labels = [METHOD_VI.get(m, m) for m in methods]
    gini = [float(assign_df[assign_df["method"] == m].iloc[0]["load_gini"]) for m in methods]
    stddev = [float(assign_df[assign_df["method"] == m].iloc[0]["load_stddev"]) for m in methods]
    x = np.arange(len(methods))
    width = 0.36
    fig, axes = plt.subplots(1, 2, figsize=(11.0, 4.8))

    colors = [METHOD_COLOR.get(m, PALETTE["slate"]) for m in methods]
    b1 = axes[0].bar(x, gini, color=colors, edgecolor="white", width=0.62)
    for bar, v in zip(b1, gini):
        axes[0].text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(gini) * 0.03,
            f"{v:.4f}",
            ha="center",
            fontsize=9,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(labels, rotation=10, ha="right")
    _finish_ax(axes[0], "Load Gini (thấp hơn = cân hơn)", "Gini")

    b2 = axes[1].bar(x, stddev, color=colors, edgecolor="white", width=0.62)
    for bar, v in zip(b2, stddev):
        axes[1].text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max(stddev) * 0.03,
            f"{v:.2f}",
            ha="center",
            fontsize=9,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(labels, rotation=10, ha="right")
    _finish_ax(axes[1], "Load StdDev", "StdDev")

    fig.suptitle(
        "Cân bằng tải phân công (assignment)",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        y=1.02,
    )
    fig.tight_layout()
    return save_fig(fig, "fig07_assignment_load_balance.png")


def chart_coi_violations(assign_df: pd.DataFrame) -> Path:
    methods = list(assign_df["method"])
    labels = [METHOD_VI.get(m, m) for m in methods]
    viol = [int(assign_df[assign_df["method"] == m].iloc[0]["coi_violations"]) for m in methods]
    colors = [PALETTE["pass"] if v == 0 else PALETTE["fail"] for v in viol]
    fig, ax = plt.subplots(figsize=(9.0, 4.8))
    bars = ax.bar(labels, viol, color=colors, edgecolor="white", width=0.55)
    for bar, v in zip(bars, viol):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            max(0.08, v + 0.05),
            str(v),
            ha="center",
            fontsize=14,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    ax.set_ylim(0, max(1.0, max(viol) * 1.4 if viol else 1))
    _finish_ax(ax, "COI violations sau phân công (ràng buộc cứng = 0)", "Số vi phạm")
    fig.text(
        0.02,
        -0.02,
        "Cả greedy (production) và baselines đều nhận cùng bản đồ COI (self-authorship) và không gán cặp xung đột.",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig08_coi_violations.png")


def chart_assignment_fallback(assign_df: pd.DataFrame) -> Path:
    methods = list(assign_df["method"])
    labels = [METHOD_VI.get(m, m) for m in methods]
    fb = [float(assign_df[assign_df["method"] == m].iloc[0]["fallback_rate"]) for m in methods]
    colors = [METHOD_COLOR.get(m, PALETTE["slate"]) for m in methods]
    fig, ax = plt.subplots(figsize=(9.0, 4.8))
    bars = ax.bar(labels, fb, color=colors, edgecolor="white", width=0.55)
    for bar, v in zip(bars, fb):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.01,
            f"{v:.1%}",
            ha="center",
            fontsize=11,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    ax.set_ylim(0, max(0.35, max(fb) * 1.35 if fb else 0.35))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.0%}"))
    _finish_ax(ax, "Tỷ lệ fallback (pass-2) khi greedy không đủ reviewer", "Fallback rate")
    fig.text(
        0.02,
        -0.02,
        "Baselines không có pass-2 (fallback=0) vì luôn gán đủ slot, bất chấp điểm phù hợp thấp.",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig09_assignment_fallback.png")


def chart_headline(rank_df: pd.DataFrame, assign_df: pd.DataFrame) -> Path:
    """Slide-ready headline metrics for Chapter 4."""
    cards: list[tuple[str, str, str]] = []
    if not rank_df.empty:
        j = rank_df[rank_df["method"] == "jaccard"]
        if not j.empty:
            row = j.iloc[0]
            cards.append(("Hit@1", f"{row['hit_at_1']:.0%}", "Jaccard production"))
            cards.append(("Hit@5", f"{row['hit_at_5']:.0%}", "Jaccard production"))
            cards.append(("MRR", f"{row['mrr']:.3f}", "Jaccard production"))
    if not assign_df.empty:
        g = assign_df[assign_df["method"] == "greedy"]
        if not g.empty:
            row = g.iloc[0]
            cards.append(("Coverage", f"{row['coverage']:.0%}", "Greedy production"))
            cards.append(("Mean score", f"{row['mean_score']:.3f}", "Greedy production"))
            cards.append(("COI viol.", f"{int(row['coi_violations'])}", "Hard constraint"))

    n = len(cards) or 1
    fig, axes = plt.subplots(1, n, figsize=(2.2 * n + 1.5, 3.6))
    if n == 1:
        axes = [axes]
    colors = [
        PALETTE["teal"],
        PALETTE["mint"],
        PALETTE["navy"],
        PALETTE["amber"],
        PALETTE["ink"],
        PALETTE["pass"],
    ]
    for ax, (title, value, note), color in zip(axes, cards, colors):
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")
        ax.add_patch(
            plt.Rectangle((0.05, 0.08), 0.9, 0.84, facecolor=color, edgecolor="white", linewidth=0, alpha=0.92)
        )
        ax.text(0.5, 0.72, title, ha="center", va="center", fontsize=11, color="white", fontweight="bold")
        ax.text(0.5, 0.42, value, ha="center", va="center", fontsize=20, color="white", fontweight="bold")
        ax.text(0.5, 0.18, note, ha="center", va="center", fontsize=8, color="white", alpha=0.9)
    fig.suptitle(
        "Headline — deterministic matching quality",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        y=1.02,
    )
    fig.tight_layout()
    return save_fig(fig, "fig10_headline_metrics.png")


# ---------------------------------------------------------------------------
# Docs
# ---------------------------------------------------------------------------


def write_figure_index(fig_meta: list[tuple[str, str, str]]) -> None:
    lines = [
        "# Figure index — deterministic workflow (Chương 4)",
        "",
        "Sinh tự động từ `scripts/export_benchmark_to_excel.py`.",
        "Nguồn: `benchmark_output/quality-results.{csv,md}` + `ranking_case_counts.xlsx`.",
        "Dùng PNG trong `exports/figures/` cho **Chương 4 — đánh giá thuật toán deterministic** "
        "(đối sánh phản biện, phân công, COI).",
        "",
        "| File | Nội dung | Gợi ý chỗ dùng |",
        "| --- | --- | --- |",
    ]
    for fname, content, placement in fig_meta:
        lines.append(f"| `{fname}` | {content} | {placement} |")
    lines.extend(
        [
            "",
            "## Excel workbook",
            "",
            "- `deterministic_workflow_results.xlsx` — workbook báo cáo",
            "  - `Overview` — headline + fixture + COI",
            "  - `Ranking_Metrics` — Hit@k / MRR / nDCG + case counts",
            "  - `Ranking_Case_Detail` — long form Hit@k đúng/sai theo method",
            "  - `Assignment_Metrics` — coverage / load / score / fallback / COI",
            "  - `Quality_Long` — CSV gốc long-form (section, method, metric, value)",
            "  - `Fixture` — thống kê dataset",
            "  - `Case_Counts_*` — sheet copy từ `ranking_case_counts.xlsx`",
            "  - `Run_Meta` — provenance",
            "- `benchmark_output/quality-results.{csv,md}` — raw từ backend quality suite",
            "- `ranking_case_counts.xlsx` — bảng case-level đã soạn cho báo cáo",
            "",
            "## Caveats (bắt buộc khi trích vào báo cáo)",
            "",
            "1. **Chất lượng ≠ tốc độ.** Gói này đo *độ đúng* ranking/assignment/COI. "
            "Tốc độ HTTP/micro nằm ở `system_performance/`.",
            "2. **Proxy ground truth:** leave-one-out authorship — author được coi là reviewer "
            "phù hợp cho paper của chính họ. Không phải gold assignment từ hội nghị thật.",
            "3. **Lexical Jaccard:** khớp chuỗi exact trên topic/keyword — synonym (`NLP` ≠ "
            "`Natural Language Processing`) bị tính miss.",
            "4. **COI trong quality suite** = self-authorship map trong scenario synthetic; "
            "không thay đánh giá full 3-layer COI (declared + Neo4j multi-hop) dưới tải production "
            "(xem k6 COI trong system_performance).",
            "5. Greedy coverage < 100% là trade-off có chủ đích: ưu tiên điểm phù hợp; pass-2 "
            "chỉ cứu paper 0 reviewer, không lấp paper dưới min reviewers.",
            "6. `overlap_count` có thể gần hoặc hơi hơn Jaccard ở một số metric (union-normalization "
            "phạt profile rộng) — vẫn giữ Jaccard production vì ổn định hơn trên profile không đều.",
            "7. Số case đúng/sai = `round(rate × n_queries)`; khớp `ranking_case_counts.xlsx`.",
            "",
        ]
    )
    (EXPORT / "FIGURE_INDEX.md").write_text("\n".join(lines), encoding="utf-8")


def write_summary_md(
    fixture: dict[str, Any],
    rank_df: pd.DataFrame,
    assign_df: pd.DataFrame,
) -> None:
    lines = [
        "# Tóm tắt benchmark deterministic workflow (Chương 4)",
        "",
        f"**Run:** `{RUN_META['run_id']}`  ",
        f"**Nguồn raw:** `{RUN_META['source_csv']}`  ",
        f"**Suite:** `{RUN_META['suite_path']}`  ",
        f"**Fixture:** {fixture.get('authors')} authors · {fixture.get('papers')} papers · "
        f"{fixture.get('loo_queries')} LOO queries · vocab {fixture.get('topic_vocab')}  ",
        f"**Thuật toán:** {RUN_META['algorithm_family']}",
        "",
        "## 1. Reviewer ranking (gợi ý phản biện)",
        "",
        "Proxy ground truth: leave-one-out authorship — author thật phải xuất hiện trong top-k.",
        "",
        "| Method | Role | Hit@1 | Hit@5 | Hit@10 | MRR | nDCG@10 |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for _, r in rank_df.iterrows():
        lines.append(
            f"| {r['method']} | {r['role']} | {r['hit_at_1']:.3f} | {r['hit_at_5']:.3f} | "
            f"{r['hit_at_10']:.3f} | {r['mrr']:.3f} | {r['ndcg_at_10']:.3f} |"
        )

    if not rank_df.empty:
        j = rank_df[rank_df["method"] == "jaccard"].iloc[0]
        rnd = rank_df[rank_df["method"] == "random"].iloc[0]
        lift = float(j["mrr"]) / float(rnd["mrr"]) if float(rnd["mrr"]) else float("nan")
        lines.extend(
            [
                "",
                "**Nhận xét ngắn:**",
                "",
                f"- Jaccard production: Hit@1 **{j['hit_at_1']:.0%}** "
                f"({int(j['correct_hit@1'])}/{int(j['n_queries'])}), "
                f"Hit@5 **{j['hit_at_5']:.0%}**, Hit@10 **{j['hit_at_10']:.0%}**, "
                f"MRR **{j['mrr']:.3f}**.",
                f"- Vượt random rõ: MRR lift ≈ **{lift:.1f}×** "
                f"(random MRR {rnd['mrr']:.3f} ≈ sàn lý thuyết leave-one-out).",
                "- `overlap_count` gần Jaccard (thậm chí Hit@10 cao hơn một chút) — "
                "union-normalization của Jaccard phạt profile topic rộng.",
                "",
            ]
        )

    lines.extend(
        [
            "## 2. Assignment (phân công tối ưu ràng buộc)",
            "",
            "| Method | Role | Coverage | Load StdDev | Load Gini | COI | Mean score | Fallback |",
            "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for _, r in assign_df.iterrows():
        lines.append(
            f"| {r['method']} | {r['role']} | {r['coverage']:.3f} | {r['load_stddev']:.3f} | "
            f"{r['load_gini']:.3f} | {int(r['coi_violations'])} | {r['mean_score']:.4f} | "
            f"{r['fallback_rate']:.3f} |"
        )

    if not assign_df.empty:
        g = assign_df[assign_df["method"] == "greedy"].iloc[0]
        lines.extend(
            [
                "",
                "**Nhận xét ngắn:**",
                "",
                f"- Greedy production: coverage **{g['coverage']:.1%}**, mean score **{g['mean_score']:.4f}**, "
                f"fallback **{g['fallback_rate']:.1%}**, **0 COI violations**.",
                "- Baselines (round_robin / random) đạt coverage 100% nhưng mean score ~3× thấp hơn — "
                "chứng tỏ greedy đổi coverage lấy relevance.",
                "- Load Gini greedy cao hơn baselines một chút (tập trung reviewer giỏi); "
                "vẫn trong mức chấp nhận được cho lượt gợi ý.",
                "",
            ]
        )

    lines.extend(
        [
            "## 3. COI (ràng buộc cứng trong assignment)",
            "",
            f"- Tổng COI violations trên mọi method: "
            f"**{int(assign_df['coi_violations'].fillna(0).sum()) if not assign_df.empty else 0}**.",
            "- Quality suite gắn self-authorship COI vào scenario; greedy/round_robin/random "
            "đều nhận cùng conflict map và **không** gán cặp xung đột.",
            "- Đánh giá latency/throughput endpoint COI production nằm ở gói `system_performance` "
            "(k6 scenario `coi`, ~16.7k request, median ~9.5 ms).",
            "",
            "## 4. File dùng trong báo cáo",
            "",
            "| Nhu cầu | File |",
            "| --- | --- |",
            "| Bảng ranking / assignment / COI | `exports/deterministic_workflow_results.xlsx` |",
            "| Biểu đồ chèn LaTeX/Word | `exports/figures/fig01_…` → `fig10_…` |",
            "| Mục lục figure + caveats | `exports/FIGURE_INDEX.md` |",
            "| Case counts Hit@k (đã soạn) | `ranking_case_counts.xlsx` |",
            "| Raw backend quality | `benchmark_output/quality-results.{csv,md}` |",
            "",
            "## 5. Caveats (bắt buộc)",
            "",
            "1. Proxy ground truth leave-one-out — đo tính nhất quán fingerprint topic, không phải gold assignment hội nghị.",
            "2. Jaccard lexical exact-match — synonym là miss.",
            "3. COI quality = self-authorship trong scenario; full graph COI xem system_performance + Neo4j.",
            "4. Không trộn số liệu quality với latency k6/micro.",
            "",
        ]
    )
    (EXPORT / "SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    EXPORT.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    apply_style()

    print("Loading quality results…")
    long_df = load_quality_long()
    fixture = parse_fixture_from_md(QUALITY_MD)
    rank_df = ranking_wide(long_df, fixture)
    assign_df = assignment_wide(long_df)
    case_detail = build_ranking_case_detail(rank_df)
    overview = build_overview(fixture, rank_df, assign_df)
    case_sheets = load_case_counts()

    fixture_df = pd.DataFrame(
        [
            {"field": k, "value": v}
            for k, v in {
                **fixture,
                "algorithm_family": RUN_META["algorithm_family"],
                "proxy_ground_truth": RUN_META["proxy_ground_truth"],
            }.items()
        ]
    )
    run_meta_df = pd.DataFrame(
        [{"field": k, "value": json.dumps(v) if isinstance(v, (list, dict)) else v} for k, v in RUN_META.items()]
    )

    xlsx_path = EXPORT / "deterministic_workflow_results.xlsx"
    print(f"Writing {xlsx_path}…")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        write_df(writer, overview, "Overview")
        write_df(writer, rank_df, "Ranking_Metrics")
        write_df(writer, case_detail, "Ranking_Case_Detail")
        write_df(writer, assign_df, "Assignment_Metrics")
        write_df(writer, long_df, "Quality_Long")
        write_df(writer, fixture_df, "Fixture")
        for sheet_name, df in case_sheets.items():
            # Preserve original Vietnamese sheet names, Excel-safe
            safe = re.sub(r"[\\/*?:\[\]]", "_", str(sheet_name))[:31]
            write_df(writer, df, f"Case_{safe}" if not safe.startswith("Case") else safe)
        write_df(writer, run_meta_df, "Run_Meta")

    print("Rendering figures…")
    fig_meta: list[tuple[str, str, str]] = []

    p = chart_fixture_setup(fixture)
    fig_meta.append((p.name, "Quy mô fixture: authors / papers / LOO / vocab", "Ch4 — thiết lập đánh giá quality"))

    p = chart_ranking_hit_at_k(rank_df)
    fig_meta.append((p.name, "Hit@1 / Hit@5 / Hit@10 theo method", "Ch4 — reviewer ranking"))

    p = chart_ranking_mrr_ndcg(rank_df)
    fig_meta.append((p.name, "MRR và nDCG@10 theo method", "Ch4 — reviewer ranking"))

    p = chart_ranking_case_counts(rank_df)
    fig_meta.append((p.name, "Jaccard: số case đúng/sai theo Hit@k", "Ch4 — bảng case counts"))

    p = chart_ranking_vs_baselines(rank_df)
    fig_meta.append((p.name, "Jaccard vs overlap_count vs random", "Ch4 — so baseline ranking"))

    p = chart_assignment_coverage_score(assign_df)
    fig_meta.append((p.name, "Coverage vs mean relevance score", "Ch4 — assignment quality"))

    p = chart_assignment_load_balance(assign_df)
    fig_meta.append((p.name, "Load Gini + StdDev theo method", "Ch4 — cân bằng tải"))

    p = chart_coi_violations(assign_df)
    fig_meta.append((p.name, "COI violations = 0 trên mọi method", "Ch4 — ràng buộc COI"))

    p = chart_assignment_fallback(assign_df)
    fig_meta.append((p.name, "Fallback rate (pass-2) greedy", "Ch4 — trade-off coverage"))

    p = chart_headline(rank_df, assign_df)
    fig_meta.append((p.name, "Headline metrics cho slide/báo cáo", "Ch4 — tóm tắt kết quả"))

    write_figure_index(fig_meta)
    write_summary_md(fixture, rank_df, assign_df)

    summary_json = {
        "run_meta": RUN_META,
        "fixture": fixture,
        "ranking": rank_df.to_dict(orient="records"),
        "assignment": assign_df.to_dict(orient="records"),
        "headline": overview.to_dict(orient="records"),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "figures": [f[0] for f in fig_meta],
    }
    (EXPORT / "summary_metrics.json").write_text(
        json.dumps(summary_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Wrote:", xlsx_path)
    print("Figures:", FIGURES, f"({len(fig_meta)} charts)")
    print("Docs: FIGURE_INDEX.md, SUMMARY.md, summary_metrics.json")
    if not rank_df.empty and not assign_df.empty:
        j = rank_df[rank_df["method"] == "jaccard"].iloc[0]
        g = assign_df[assign_df["method"] == "greedy"].iloc[0]
        print(
            "Headline:",
            {
                "jaccard_hit@1": float(j["hit_at_1"]),
                "jaccard_hit@5": float(j["hit_at_5"]),
                "jaccard_mrr": float(j["mrr"]),
                "greedy_coverage": float(g["coverage"]),
                "greedy_mean_score": float(g["mean_score"]),
                "coi_violations": int(g["coi_violations"]),
            },
        )


if __name__ == "__main__":
    main()
