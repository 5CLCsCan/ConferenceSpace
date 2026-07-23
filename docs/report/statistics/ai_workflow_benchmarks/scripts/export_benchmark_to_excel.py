#!/usr/bin/env python3
"""Export workflow benchmark outputs to Excel and report-ready PNG charts.

Reads docs/report/raw/workflow_benchmarks/benchmark_output and writes:
  - exports/workflow_benchmark_results.xlsx
  - exports/figures/*.png
  - exports/FIGURE_INDEX.md
"""

from __future__ import annotations

import json
import textwrap
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.patches import FancyBboxPatch, Polygon
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "benchmark_output"
EXPORT = ROOT / "exports"
FIGURES = EXPORT / "figures"
REPORT_FIGSIZE = (8.0, 16.0 / 3.0)
# Slightly larger 3:2 canvas for dense multi-category charts.
REPORT_FIGSIZE_DENSE = (10.5, 7.0)
REPORT_DPI = 375

# Manual chatbot outcomes from chatbot_agent_benchmark_report.md (section 5.1 / 5.2)
CHATBOT_MANUAL_OVERALL = {
    "scenarios": 8,
    "trials": 40,
    "completed_trials": 40,
    "pass": 25,
    "partial": 12,
    "fail": 3,
}
CHATBOT_MANUAL_BY_SCENARIO = {
    "author_own_submission_status": {
        "label_vi": "Tra cứu trạng thái bài nộp",
        "manual": "partial",
        "note": "4/5 đúng trạng thái; 1 lượt dừng ở lỗi truy vấn",
    },
    "author_submission_track": {
        "label_vi": "Tra cứu track và metadata của bài nộp",
        "manual": "partial",
        "note": "4/5 lượt xác định đúng track; 1 lượt nhầm định danh hội nghị",
    },
    "chair_conference_overview": {
        "label_vi": "Tóm tắt tình hình hội nghị",
        "manual": "partial",
        "note": "Tổng quan chính đúng; một số lượt thiếu số lượng phản biện viên",
    },
    "reviewer_assignment_check": {
        "label_vi": "Kiểm tra khối lượng công việc phản biện",
        "manual": "pass",
        "note": "5/5 lượt kết luận đúng khối lượng công việc phản biện",
    },
    "public_conference_lookup": {
        "label_vi": "Tra cứu thông tin công khai",
        "manual": "partial",
        "note": "4/5 lượt trả đúng thông tin; 1 lượt dùng sai nguồn dữ liệu",
    },
    "permission_boundary_other_submission": {
        "label_vi": "Ranh giới quyền truy cập",
        "manual": "partial",
        "note": "Không lộ dữ liệu; diễn đạt còn giống lỗi kỹ thuật",
    },
    "unsupported_external_research": {
        "label_vi": "Yêu cầu ngoài phạm vi",
        "manual": "partial",
        "note": "1/5 vẫn viết báo cáo ngoài phạm vi nền tảng",
    },
    "chair_multi_step_platform_report": {
        "label_vi": "Báo cáo vận hành nhiều bước",
        "manual": "partial",
        "note": "Một số lượt phối hợp tốt nhiều công cụ; lỗi truy vấn làm thiếu chi tiết",
    },
}

PALETTE = {
    "navy": "#1B3A4B",
    "teal": "#0D7377",
    "coral": "#C44536",
    "amber": "#E09F3E",
    "slate": "#4A5568",
    "mint": "#2A9D8F",
    "sand": "#F4A261",
    "ink": "#264653",
    "pass": "#2A9D8F",
    "partial": "#E09F3E",
    "fail": "#C44536",
    "warn": "#E09F3E",
    "block": "#C44536",
}

# Canonical operational resource table (Chapter 4 — Bảng độ trễ & token).
# Source: workflow benchmark reports + chapter4.tex tab:ch4-ai-latency-token.
# Token values are per work unit (paper / audit / conversation) — not comparable across rows.
OPERATIONAL_RESOURCE_ROWS: list[dict[str, Any]] = [
    {
        "workflow_id": "autofill_metadata",
        "label_vi": "Submission Autofill — trích xuất metadata",
        "label_short": "Submission Autofill:\ntrích xuất metadata",
        "label_chart": "Submission Autofill\nmetadata",
        "unit": "bài",
        "latency_mean_s": 10.64,
        "latency_median_s": 9.32,
        "latency_max_s": 102.20,
        "token_mean": 4094.0,
        "ops_mode": "interactive_with_timeout",
        "ops_mode_vi": "Tương tác có giới hạn chờ và trạng thái xử lý",
        "source": "submission_autofill_benchmark_report.md",
    },
    {
        "workflow_id": "autofill_track",
        "label_vi": "Submission Autofill — Track Recommendation",
        "label_short": "Submission Autofill:\nTrack Recommendation",
        "label_chart": "Submission Autofill\nTrack Recommendation",
        "unit": "bài",
        "latency_mean_s": 18.19,
        "latency_median_s": 17.54,
        "latency_max_s": 37.42,
        "token_mean": None,
        "ops_mode": "interactive_with_timeout",
        "ops_mode_vi": "Tương tác có giới hạn chờ và trạng thái xử lý",
        "source": "track_recommendation / chapter4.tex",
    },
    {
        "workflow_id": "gating_rule",
        "label_vi": "Submission Gating — Rule Check",
        "label_short": "Submission Gating:\nRule Check",
        "label_chart": "Submission Gating\nRule Check",
        "unit": "lượt kiểm tra",
        "latency_mean_s": 0.08,
        "latency_median_s": 0.09,
        "latency_max_s": 0.14,
        "token_mean": None,
        "ops_mode": "sync",
        "ops_mode_vi": "Đồng bộ",
        "source": "submission_gating_benchmark_report.md",
    },
    {
        "workflow_id": "gating_llm",
        "label_vi": "Submission Gating — LLM Steering",
        "label_short": "Submission Gating:\nLLM Steering",
        "label_chart": "Submission Gating\nLLM Steering",
        "unit": "lượt kiểm tra",
        "latency_mean_s": 11.83,
        "latency_median_s": 11.47,
        "latency_max_s": 19.64,
        "token_mean": None,
        "ops_mode": "non_blocking",
        "ops_mode_vi": "Xử lý song song, không chặn thao tác",
        "source": "submission_gating_benchmark_report.md",
    },
    {
        "workflow_id": "reviewer_initial_analysis",
        "label_vi": "Reviewer Initial Analysis",
        "label_short": "Reviewer Initial\nAnalysis",
        "label_chart": "Reviewer Initial\nAnalysis",
        "unit": "bài",
        "latency_mean_s": 39.18,
        "latency_median_s": 37.53,
        "latency_max_s": 126.36,
        "token_mean": 11575.0,
        "ops_mode": "background",
        "ops_mode_vi": "Xử lý nền hoặc chạy trước",
        "source": "reviewer_initial_analysis_benchmark_report.md",
    },
    {
        "workflow_id": "review_quality_auditor",
        "label_vi": "Review Quality Auditor",
        "label_short": "Review Quality\nAuditor",
        "label_chart": "Review Quality\nAuditor",
        "unit": "lượt kiểm tra",
        "latency_mean_s": 15.55,
        "latency_median_s": 14.63,
        "latency_max_s": 123.67,
        "token_mean": 7874.0,
        "ops_mode": "background",
        "ops_mode_vi": "Xử lý nền và hiển thị trạng thái",
        "source": "review_quality_auditor_benchmark_report.md",
    },
    {
        "workflow_id": "chair_decision_copilot",
        "label_vi": "Chair Decision Copilot",
        "label_short": "Chair Decision\nCopilot",
        "label_chart": "Chair Decision\nCopilot",
        "unit": "bài",
        "latency_mean_s": 21.68,
        "latency_median_s": 20.59,
        "latency_max_s": 116.74,
        "token_mean": 6242.0,
        "ops_mode": "background",
        "ops_mode_vi": "Xử lý nền hoặc chạy trước",
        "source": "chair_decision_copilot_benchmark_report.md",
    },
    {
        "workflow_id": "chatbot_agent",
        "label_vi": "Chatbot Agent",
        "label_short": "Chatbot\nAgent",
        "label_chart": "Chatbot\nAgent",
        "unit": "hội thoại",
        "latency_mean_s": 26.53,
        "latency_median_s": None,
        "latency_max_s": 57.89,
        "token_mean": 360.5,
        "ops_mode": "streaming",
        "ops_mode_vi": "Phản hồi từng phần và hiển thị trạng thái tra cứu",
        "source": "chatbot_agent_benchmark_report.md",
        "note": "Độ trễ cao nhất là giá trị của nhóm kịch bản chậm nhất, không phải giá trị cao nhất của một hội thoại riêng lẻ.",
    },
]

# Chatbot response-stage metrics (same report as Table 4.7 chatbot row).
CHATBOT_RESPONSE_STAGES: dict[str, float] = {
    "ttft_s": 2.36,
    "first_answer_token_s": 23.02,
    "total_duration_s": 26.53,
    "stream_duration_s": 24.17,
    "tool_calls_total": 128,
    "tool_calls_failed": 31,
    "tool_success_rate_pct": 75.78,
}


def read_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1B3A4B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:80]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str) -> None:
    out = df.copy()
    if out.empty:
        out = pd.DataFrame({"note": ["(empty)"]})
    out.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.book[sheet]
    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def safe_mean(values: list[float | int | None]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return round(mean(clean), 4) if clean else None


def safe_median(values: list[float | int | None]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return round(median(clean), 4) if clean else None


def pct(n: float | int | None, d: float | int | None) -> float | None:
    if n is None or d is None or d == 0:
        return None
    return round(100.0 * float(n) / float(d), 2)


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_track() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "track_recommendation" / "human_review.csv")
    metrics = read_json(INPUT / "track_recommendation" / "summary_metrics.json")
    manifest = read_json(INPUT / "track_recommendation" / "manifest.json")
    conf_counts = review["conference"].value_counts().to_dict()
    summary = {
        "workflow": "track_recommendation",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "invalid_track_rate": metrics.get("invalid_track_rate"),
        "invalid_track_cases": int((review["invalid_track_names"].fillna("").astype(str).str.len() > 0).sum()),
        "human_label_filled": int(review["human_label"].fillna("").astype(str).str.strip().ne("").sum()),
        "top1_human_accuracy": metrics.get("top1_human_accuracy"),
        "top3_human_acceptance": metrics.get("top3_human_acceptance"),
        "conferences": len(conf_counts),
        "model": (manifest.get("llm_config") or {}).get("openai_model"),
        "agent_model": (manifest.get("llm_config") or {}).get("agent_model"),
        "created_at": manifest.get("created_at"),
        "limitation": "human_label còn trống — chưa có top-1/top-3 accuracy từ nhãn người",
    }
    review = review.copy()
    review["has_prediction"] = review["predicted_top1"].fillna("").astype(str).str.strip().ne("")
    review["top3_count"] = review["predicted_top3"].fillna("").astype(str).apply(
        lambda s: len([x for x in s.split(";") if x.strip()]) if s.strip() else 0
    )
    return review, summary


def load_gating_rule() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "submission_gating_rule_check" / "rule_review.csv")
    metrics = read_json(INPUT / "submission_gating_rule_check" / "summary_metrics.json")
    manifest = read_json(INPUT / "submission_gating_rule_check" / "manifest.json")
    summary = {
        "workflow": "submission_gating_rule_check",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "correct_verdict_count": metrics.get("correct_verdict_count"),
        "correct_rules_count": metrics.get("correct_rules_count"),
        "false_block_count": metrics.get("false_block_count"),
        "blocking_verdict_accuracy": metrics.get("blocking_verdict_accuracy"),
        "rule_id_recall": metrics.get("rule_id_recall"),
        "created_at": manifest.get("created_at"),
    }
    return review, summary


def load_gating_llm() -> tuple[pd.DataFrame, dict[str, Any]]:
    review = pd.read_csv(INPUT / "submission_gating_llm_steering" / "llm_steering_review.csv")
    metrics = read_json(INPUT / "submission_gating_llm_steering" / "summary_metrics.json")
    manifest = read_json(INPUT / "submission_gating_llm_steering" / "manifest.json")
    findings = read_jsonl(INPUT / "submission_gating_llm_steering" / "normalized_content_findings.jsonl")
    summary = {
        "workflow": "submission_gating_llm_steering",
        "case_count": metrics.get("case_count"),
        "completed_count": metrics.get("completed_count"),
        "failed_count": metrics.get("failed_count"),
        "content_finding_count": metrics.get("content_finding_count"),
        "llm_block_contract_violation_count": metrics.get("llm_block_contract_violation_count"),
        "review_rows": len(review),
        "verdict_pass": int((review["verdict"] == "pass").sum()) if "verdict" in review else None,
        "verdict_warn": int((review["verdict"] == "warn").sum()) if "verdict" in review else None,
        "verdict_block": int((review["verdict"] == "block").sum()) if "verdict" in review else None,
        "normalized_findings": len(findings),
        "human_label_filled": int(review["human_label"].fillna("").astype(str).str.strip().ne("").sum())
        if "human_label" in review
        else 0,
        "created_at": manifest.get("created_at"),
        "limitation": "human grounded/actionable labels còn trống",
    }
    # Compact columns for Excel readability
    keep = [
        c
        for c in [
            "task_id",
            "case_id",
            "conference",
            "paper_id",
            "scenario_tags",
            "chair_focus_area",
            "submission_title",
            "verdict",
            "decision",
            "finding_rule_id",
            "finding_source",
            "finding_severity",
            "finding_message",
            "expected_behavior",
            "llm_block_contract_violation",
            "human_label",
            "grounded",
            "severity_ok",
            "actionable",
        ]
        if c in review.columns
    ]
    compact = review[keep].copy()
    for col in ["finding_message", "expected_behavior", "submission_title"]:
        if col in compact.columns:
            compact[col] = compact[col].astype(str).str.slice(0, 240)
    return compact, summary


def load_chatbot() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    run = read_json(INPUT / "chatbot_agent" / "chatbot_run_1783406671_1783411092" / "run_summary.json")
    rows = []
    for r in run.get("results", []):
        trial = r.get("trial") or {}
        timing = r.get("timing_metrics") or {}
        tools = r.get("tool_metrics") or {}
        tokens = r.get("token_usage") or {}
        rows.append(
            {
                "scenario_id": r.get("scenario_id"),
                "scenario_result_id": r.get("scenario_result_id"),
                "trial_index": trial.get("index"),
                "actor": r.get("actor"),
                "status": r.get("status"),
                "duration_ms": r.get("duration_ms"),
                "ttft_ms": timing.get("ttft_ms"),
                "time_to_first_answer_token_ms": timing.get("time_to_first_answer_token_ms"),
                "answer_stream_duration_ms": timing.get("answer_stream_duration_ms"),
                "stream_duration_ms": timing.get("stream_duration_ms"),
                "tool_call_count": tools.get("tool_call_count"),
                "successful_tool_call_count": tools.get("successful_tool_call_count"),
                "failed_tool_call_count": tools.get("failed_tool_call_count"),
                "tool_call_success_rate": tools.get("tool_call_success_rate"),
                "final_answer_chars": r.get("final_answer_chars"),
                "input_tokens": tokens.get("input_tokens"),
                "output_tokens": tokens.get("output_tokens"),
                "total_tokens": tokens.get("total_tokens"),
                "error": r.get("error") or "",
            }
        )
    trials = pd.DataFrame(rows)

    scenario_rows = []
    for scenario_id, group in trials.groupby("scenario_id", sort=False):
        meta = CHATBOT_MANUAL_BY_SCENARIO.get(scenario_id, {})
        tool_calls = int(group["tool_call_count"].fillna(0).sum())
        tool_ok = int(group["successful_tool_call_count"].fillna(0).sum())
        scenario_rows.append(
            {
                "scenario_id": scenario_id,
                "label_vi": meta.get("label_vi", scenario_id),
                "actor": group["actor"].iloc[0],
                "trials": len(group),
                "manual_outcome": meta.get("manual", ""),
                "manual_note": meta.get("note", ""),
                "avg_duration_s": round(group["duration_ms"].mean() / 1000.0, 2),
                "avg_ttft_s": round(group["ttft_ms"].mean() / 1000.0, 2),
                "avg_time_to_first_answer_s": round(
                    group["time_to_first_answer_token_ms"].mean() / 1000.0, 2
                ),
                "total_tool_calls": tool_calls,
                "successful_tool_calls": tool_ok,
                "tool_success_rate_pct": pct(tool_ok, tool_calls),
                "avg_total_tokens": round(group["total_tokens"].mean(), 1),
            }
        )
    by_scenario = pd.DataFrame(scenario_rows)
    agg = run.get("aggregate_metrics") or {}
    summary = {
        "workflow": "chatbot_agent",
        "run_instance_id": run.get("run_instance_id"),
        "started_at": run.get("started_at"),
        "finished_at": run.get("finished_at"),
        "scenario_count": agg.get("scenario_count"),
        "trial_count": agg.get("trial_count"),
        "avg_total_duration_ms": agg.get("avg_total_duration_ms"),
        "avg_ttft_ms": agg.get("avg_ttft_ms"),
        "avg_time_to_first_answer_token_ms": agg.get("avg_time_to_first_answer_token_ms"),
        "avg_stream_duration_ms": agg.get("avg_stream_duration_ms"),
        "total_tool_calls": agg.get("total_tool_calls"),
        "successful_tool_calls": agg.get("successful_tool_calls"),
        "failed_tool_calls": agg.get("failed_tool_calls"),
        "tool_call_success_rate": agg.get("tool_call_success_rate"),
        "input_tokens": agg.get("input_tokens"),
        "output_tokens": agg.get("output_tokens"),
        "total_tokens": agg.get("total_tokens"),
        "manual_pass": CHATBOT_MANUAL_OVERALL["pass"],
        "manual_partial": CHATBOT_MANUAL_OVERALL["partial"],
        "manual_fail": CHATBOT_MANUAL_OVERALL["fail"],
        "manual_pass_rate_pct": pct(
            CHATBOT_MANUAL_OVERALL["pass"], CHATBOT_MANUAL_OVERALL["trials"]
        ),
        "source_manual": "chatbot_agent_benchmark_report.md",
    }
    return trials, by_scenario, summary


def load_autofill() -> tuple[pd.DataFrame, dict[str, Any]]:
    path = INPUT / "completed (7).csv"
    df = pd.read_csv(path)
    summary = {
        "workflow": "submission_autofill_completed_csv",
        "case_count": len(df),
        "title_exact_match_rate": round(df["title_exact_match"].mean(), 4)
        if "title_exact_match" in df
        else None,
        "abstract_rouge_l_mean": safe_mean(df["abstract_rouge_l"].tolist())
        if "abstract_rouge_l" in df
        else None,
        "abstract_rouge_l_median": safe_median(df["abstract_rouge_l"].tolist())
        if "abstract_rouge_l" in df
        else None,
        "keyword_f1_mean": safe_mean(df["keyword_f1"].tolist()) if "keyword_f1" in df else None,
        "author_f1_mean": safe_mean(df["author_f1"].tolist()) if "author_f1" in df else None,
        "time_taken_seconds_mean": safe_mean(df["time_taken_seconds"].tolist())
        if "time_taken_seconds" in df
        else None,
        "time_taken_seconds_median": safe_median(df["time_taken_seconds"].tolist())
        if "time_taken_seconds" in df
        else None,
        "total_tokens_mean": safe_mean(df["total_tokens"].tolist()) if "total_tokens" in df else None,
        "source_file": path.name,
    }
    return df, summary


def load_tca() -> tuple[pd.DataFrame, dict[str, Any]]:
    rows_raw = read_jsonl(INPUT / "tca_benchmark_results (5).jsonl")
    flat = []
    b1_exact_rates: list[float] = []
    b1_partial_rates: list[float] = []
    b1_fab_rates: list[float] = []
    b2_t: list[float] = []
    b2_c: list[float] = []
    b2_a: list[float] = []
    b5_t: list[float] = []
    b5_c: list[float] = []
    b5_a: list[float] = []
    b3_t: list[float] = []
    b3_v: list[float] = []
    b3_gv: list[float] = []
    high_risk = 0
    clean = 0
    times: list[float] = []

    # Absolute unit totals for report-facing counts next to rates
    ann_exact = ann_partial = ann_fab = 0
    ap_total = 0
    review_total = 0
    finding_total = 0
    claim_eb_total = 0
    papers_with_ap = 0
    papers_with_b3 = 0
    papers_with_b5 = 0

    for r in rows_raw:
        b1 = r.get("b1") or {}
        b2 = r.get("b2") or {}
        b3 = r.get("b3") or {}
        b5 = r.get("b5") or {}
        meta = r.get("meta") or {}
        exact = b1.get("exact_count") or 0
        partial = b1.get("partial_count") or 0
        fab = b1.get("fabricated_count") or 0
        ann_exact += exact
        ann_partial += partial
        ann_fab += fab
        total_ann = exact + partial + fab
        exact_rate = (exact / total_ann) if total_ann else None
        partial_rate = (partial / total_ann) if total_ann else None
        fab_rate = (fab / total_ann) if total_ann else None
        if exact_rate is not None:
            b1_exact_rates.append(exact_rate)
            b1_partial_rates.append(partial_rate or 0.0)
            b1_fab_rates.append(fab_rate or 0.0)
        for src, dst in [
            (b2.get("t_rate"), b2_t),
            (b5.get("evidence_basis_t_rate"), b5_t),
        ]:
            if src is not None:
                dst.append(float(src))

        b2_c_value = b2.get("c_rate")
        b2_a_value = b2.get("a_rate")
        if (
            b2_c_value is not None
            and b2_a_value is not None
            and float(b2_c_value) + float(b2_a_value) > 0
        ):
            b2_c.append(float(b2_c_value))
            b2_a.append(float(b2_a_value))

        b5_c_value = b5.get("evidence_basis_c_rate")
        b5_a_value = b5.get("evidence_basis_a_rate")
        if (
            b5_c_value is not None
            and b5_a_value is not None
            and float(b5_c_value) + float(b5_a_value) > 0
        ):
            b5_c.append(float(b5_c_value))
            b5_a.append(float(b5_a_value))

        aps = b2.get("attention_points") or []
        if isinstance(aps, list):
            ap_total += len(aps)
            if aps:
                papers_with_ap += 1
        elif isinstance(aps, int):
            ap_total += aps
            if aps:
                papers_with_ap += 1

        review_t = []
        review_v = []
        review_gv = []
        revs = b3.get("reviews") or []
        if revs:
            papers_with_b3 += 1
        for rev in revs:
            review_total += 1
            finding_total += rev.get("total_findings") or len(rev.get("findings") or [])
            if rev.get("b3_truthfulness_rate") is not None:
                value = float(rev["b3_truthfulness_rate"])
                review_t.append(value)
                b3_t.append(value)
            if rev.get("b3_validity_rate") is not None:
                value = float(rev["b3_validity_rate"])
                review_v.append(value)
                b3_v.append(value)
            if rev.get("b3_grounded_valid_rate") is not None:
                value = float(rev["b3_grounded_valid_rate"])
                review_gv.append(value)
                b3_gv.append(value)
        paper_b3_t = mean(review_t) if review_t else None
        paper_b3_v = mean(review_v) if review_v else None
        paper_b3_gv = mean(review_gv) if review_gv else None

        claims = b5.get("claims") or []
        eb_n = 0
        if isinstance(claims, list):
            for c in claims:
                if c.get("claim_type") == "evidence_basis":
                    eb_n += 1
        claim_eb_total += eb_n
        if b5.get("evidence_basis_t_rate") is not None:
            papers_with_b5 += 1

        if b5.get("high_risk"):
            high_risk += 1
        if meta.get("clean_view_eligible"):
            clean += 1
        if meta.get("processing_time_sec") is not None:
            times.append(float(meta["processing_time_sec"]))

        flat.append(
            {
                "paper_id": r.get("paper_id"),
                "b1_exact_count": exact,
                "b1_partial_count": partial,
                "b1_fabricated_count": fab,
                "b1_exact_rate": round(exact_rate, 4) if exact_rate is not None else None,
                "b1_partial_rate": round(partial_rate, 4) if partial_rate is not None else None,
                "b1_fabricated_rate": round(fab_rate, 4) if fab_rate is not None else None,
                "b2_t_rate": b2.get("t_rate"),
                "b2_c_rate": b2.get("c_rate"),
                "b2_a_rate": b2.get("a_rate"),
                "b2_t_rate_technical": b2.get("t_rate_technical"),
                "b2_administrative_ap_count": b2.get("administrative_ap_count"),
                "b3_truthfulness_mean": round(paper_b3_t, 4) if paper_b3_t is not None else None,
                "b3_validity_mean": round(paper_b3_v, 4) if paper_b3_v is not None else None,
                "b3_grounded_valid_mean": round(paper_b3_gv, 4) if paper_b3_gv is not None else None,
                "b3_review_count": len(b3.get("reviews") or []),
                "b5_evidence_basis_t_rate": b5.get("evidence_basis_t_rate"),
                "b5_evidence_basis_c_rate": b5.get("evidence_basis_c_rate"),
                "b5_evidence_basis_a_rate": b5.get("evidence_basis_a_rate"),
                "b5_high_risk": bool(b5.get("high_risk")),
                "clean_view_eligible": bool(meta.get("clean_view_eligible")),
                "processing_time_sec": meta.get("processing_time_sec"),
                "nli_pairs_total": meta.get("nli_pairs_total"),
                "pdf_extracted": meta.get("pdf_extracted"),
            }
        )

    papers = pd.DataFrame(flat)
    ann_total = ann_exact + ann_partial + ann_fab
    t_mean = safe_mean(b2_t)
    c_mean = safe_mean(b2_c)
    a_mean = safe_mean(b2_a)
    b3_t_mean = safe_mean(b3_t)
    b3_v_mean = safe_mean(b3_v)
    b3_gv_mean = safe_mean(b3_gv)
    b5_t_mean = safe_mean(b5_t)
    b5_c_mean = safe_mean(b5_c)
    b5_a_mean = safe_mean(b5_a)

    def est(rate: float | None, total: int) -> int | None:
        if rate is None or total <= 0:
            return None
        return int(round(float(rate) * total))

    summary = {
        "workflow": "tca_benchmark",
        "paper_count": len(papers),
        "b1_exact_rate_mean": safe_mean(b1_exact_rates),
        "b1_partial_rate_mean": safe_mean(b1_partial_rates),
        "b1_fabricated_rate_mean": safe_mean(b1_fab_rates),
        "b1_exact_count": ann_exact,
        "b1_partial_count": ann_partial,
        "b1_fabricated_count": ann_fab,
        "b1_annotation_total": ann_total,
        "b1_paper_count": len(b1_exact_rates),
        "b2_t_rate_mean": t_mean,
        "b2_c_rate_mean": c_mean,
        "b2_a_rate_mean": a_mean,
        "b2_attention_point_total": ap_total,
        "b2_paper_count": papers_with_ap,
        "b2_reference_eligible_paper_count": len(b2_c),
        "b2_t_count_est": est(t_mean, ap_total),
        "b2_c_count_est": None,
        "b2_a_count_est": None,
        "b3_truthfulness_mean": b3_t_mean,
        "b3_validity_mean": b3_v_mean,
        "b3_grounded_valid_mean": b3_gv_mean,
        "b3_review_total": review_total,
        "b3_finding_total": finding_total,
        "b3_paper_count": papers_with_b3,
        "b3_t_count_est": None,
        "b3_v_count_est": None,
        "b3_gv_count_est": None,
        "b5_evidence_basis_t_rate_mean": b5_t_mean,
        "b5_evidence_basis_c_rate_mean": b5_c_mean,
        "b5_evidence_basis_a_rate_mean": b5_a_mean,
        "b5_claim_total": claim_eb_total,
        "b5_paper_count": papers_with_b5,
        "b5_reference_eligible_paper_count": len(b5_c),
        "b5_t_count_est": est(b5_t_mean, claim_eb_total),
        "b5_c_count_est": None,
        "b5_a_count_est": None,
        "high_risk_papers": high_risk,
        "clean_view_eligible_papers": clean,
        "processing_time_sec_mean": safe_mean(times),
        "processing_time_sec_median": safe_median(times),
        "source_file": "tca_benchmark_results (5).jsonl",
        "note": (
            "annotation grounding; attention points (truthfulness/coverage/additionality); "
            "review-quality findings; chair evidence-basis claims"
        ),
    }
    return papers, summary


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------


def _bar(
    ax,
    labels: list[str],
    values: list[float],
    colors: list[str] | str,
    ylabel: str,
    title: str,
    value_fmt: str = "{:.1f}",
    rotate: int = 20,
) -> None:
    x = range(len(labels))
    bars = ax.bar(x, values, color=colors, edgecolor="white", linewidth=0.6)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=rotate, ha="right", fontsize=9)
    ax.set_ylabel(ylabel, fontsize=10)
    ax.set_title(title, fontsize=12, fontweight="bold", color=PALETTE["ink"], pad=10)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            value_fmt.format(val),
            ha="center",
            va="bottom",
            fontsize=8,
            color=PALETTE["slate"],
        )


def save_fig(
    fig,
    name: str,
    *,
    figsize: tuple[float, float] = REPORT_FIGSIZE,
    layout_rect: tuple[float, float, float, float] = (0.04, 0.07, 0.98, 0.96),
    apply_tight_layout: bool = True,
) -> Path:
    """Save a sharp, fixed-size 3:2 PNG for stable LaTeX placement."""
    path = FIGURES / name
    fig.set_size_inches(*figsize, forward=True)
    if apply_tight_layout:
        fig.tight_layout(rect=layout_rect)
    fig.savefig(path, dpi=REPORT_DPI, facecolor="white")
    plt.close(fig)
    return path


def chart_overview_case_counts(
    *,
    runner_n: int,
    tca_n: int,
    track_n: int,
    rule_n: int,
    llm_n: int,
    chat_n: int,
) -> Path:
    """Sample-size overview aligned with Chapter 4 dual corpora.

    1.127 is the shared workflow-runner corpus (performance, tokens, outputs),
    not Submission Autofill alone. 1.097 is the TCA scoring subset.
    """
    _setup_vietnamese_font()
    labels = [
        "Bộ thực thi\nluồng AI",
        "Bộ đánh giá\nTCA",
        "Track\nRecommendation",
        "Submission Gating\n(Rule Check)",
        "Submission Gating\n(LLM Steering)",
        "Chatbot\nAgent",
    ]
    values = [runner_n, tca_n, track_n, rule_n, llm_n, chat_n]
    colors = [
        PALETTE["coral"],
        PALETTE["navy"],
        PALETTE["teal"],
        PALETTE["mint"],
        PALETTE["amber"],
        PALETTE["ink"],
    ]
    fig = plt.figure(figsize=REPORT_FIGSIZE_DENSE)
    # Chart occupies the upper half only; x-tick labels need free space below the axes.
    ax = fig.add_axes([0.09, 0.52, 0.86, 0.40])
    x = range(len(labels))
    bars = ax.bar(list(x), values, color=colors, edgecolor="white", linewidth=0.6, width=0.68)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, fontsize=11)
    ax.set_ylabel("Số đơn vị", fontsize=12)
    ax.tick_params(axis="y", labelsize=11)
    ax.tick_params(axis="x", pad=6)
    ax.set_title(
        "Quy mô các bộ đánh giá luồng AI",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=12,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ymax = max(values) * 1.14 if values else 1
    ax.set_ylim(0, ymax)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            _fmt_int(val),
            ha="center",
            va="bottom",
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
        )

    note_lines = [
        "• Bộ thực thi luồng AI (1.127 bài): tập đầu vào chung để chạy các luồng AI, ghi hiệu năng, tài nguyên (thời gian, token) và tạo đầu ra.",
        "• Bộ đánh giá TCA (1.097 gói): đánh giá chung độ chính xác, độ trung thực, độ đáng tin cậy và tiềm năng của các luồng AI.",
        "• Bộ đánh giá chuyên biệt: Track Recommendation, Submission Gating và Chatbot Agent — mẫu số riêng, không gộp với 1.127 bài.",
    ]
    # Note band sits well below the two-line x-tick labels (axes bottom = 0.52).
    fig.text(0.09, 0.28, "Chú giải", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    y = 0.22
    for line in note_lines:
        wrapped = "\n".join(textwrap.wrap(line, width=105))
        fig.text(0.09, y, wrapped, fontsize=10, color=PALETTE["slate"], linespacing=1.3, va="top")
        y -= 0.065

    return save_fig(
        fig,
        "fig01_overview_case_counts.png",
        figsize=REPORT_FIGSIZE_DENSE,
        apply_tight_layout=False,
    )


def chart_runner_by_conference(auto_df: pd.DataFrame) -> Path:
    """Distribution of the 1.127-paper runner corpus by conference/track."""
    _setup_vietnamese_font()
    col = "conference_year_track"
    if col not in auto_df.columns:
        raise KeyError(f"Missing column {col!r} in autofill/runner corpus")

    # Normalize raw track labels for report display (match Chapter 4 table)
    display = (
        auto_df[col]
        .fillna("(không xác định)")
        .astype(str)
        .str.replace("Short_Paper_Track", "Short Paper Track", regex=False)
        .str.replace("_", " ", regex=False)
    )
    counts = display.value_counts()
    # Largest at top (barh draws bottom-first)
    counts = counts.sort_values(ascending=True)
    total = int(counts.sum())
    labels = counts.index.tolist()
    values = [int(v) for v in counts.values.tolist()]
    pcts = [100.0 * v / total if total else 0.0 for v in values]

    # Highlight top-4 sources (Chapter 4: ~70.72% of papers)
    top4 = set(counts.sort_values(ascending=False).head(4).index.tolist())
    colors = [PALETTE["coral"] if lab in top4 else PALETTE["navy"] for lab in labels]

    fig = plt.figure(figsize=REPORT_FIGSIZE)
    ax = fig.add_axes([0.30, 0.14, 0.62, 0.74])
    y = range(len(labels))
    ax.barh(list(y), values, color=colors, edgecolor="white", linewidth=0.5, height=0.72)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels, fontsize=12)
    ax.set_xlabel("Số bài", fontsize=12)
    ax.tick_params(axis="x", labelsize=11)
    ax.set_title(
        "Phân bố 1.127 bài đầu vào theo hội nghị / track",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
        loc="left",
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.35)
    xmax = max(values) * 1.22 if values else 1
    ax.set_xlim(0, xmax)

    for yi, val, pct in zip(y, values, pcts):
        ax.text(
            val + xmax * 0.015,
            yi,
            f"{_fmt_int(val)}  ({pct:.1f}%)".replace(".", ","),
            va="center",
            ha="left",
            fontsize=11,
            color=PALETTE["ink"],
            fontweight="bold",
        )

    top4_share = sum(v for lab, v in zip(labels, values) if lab in top4)
    top4_pct = 100.0 * top4_share / total if total else 0.0
    fig.text(
        0.30,
        0.04,
        (
            f"Tổng: {_fmt_int(total)} bài  ·  "
            f"Bốn nguồn lớn nhất (cam): {_fmt_int(top4_share)} bài "
            f"({top4_pct:.2f}%)".replace(".", ",")
        ),
        fontsize=11,
        color=PALETTE["slate"],
    )

    return save_fig(
        fig,
        "fig01b_runner_by_conference.png",
        apply_tight_layout=False,
    )


def chart_gating_rule(rule_df: pd.DataFrame, rule_summary: dict[str, Any]) -> Path:
    labels = ["Verdict\nAccuracy", "Rule-ID\nRecall", "False Block\nCount"]
    values = [
        100.0 * float(rule_summary.get("blocking_verdict_accuracy") or 0),
        100.0 * float(rule_summary.get("rule_id_recall") or 0),
        float(rule_summary.get("false_block_count") or 0),
    ]
    colors = [PALETTE["pass"], PALETTE["mint"], PALETTE["fail"]]
    fig, axes = plt.subplots(1, 2, figsize=REPORT_FIGSIZE)
    _bar(axes[0], labels, values, colors, "Giá trị đo", "Submission Gating — Rule Check", "{:.0f}", 0)

    verdict_counts = rule_df["actual_verdict"].value_counts()
    order = [v for v in ["pass", "warn", "block"] if v in verdict_counts.index]
    display_labels = {"pass": "pass", "warn": "warn", "block": "block"}
    v_colors = {"pass": PALETTE["pass"], "warn": PALETTE["warn"], "block": PALETTE["block"]}
    axes[1].bar(
        [display_labels[v] for v in order],
        [int(verdict_counts[v]) for v in order],
        color=[v_colors[v] for v in order],
        edgecolor="white",
    )
    axes[1].set_title("Verdict Distribution", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[1].set_ylabel("Số trường hợp")
    axes[1].spines["top"].set_visible(False)
    axes[1].spines["right"].set_visible(False)
    for i, v in enumerate(order):
        axes[1].text(i, verdict_counts[v], str(int(verdict_counts[v])), ha="center", va="bottom", fontsize=9)
    return save_fig(fig, "fig02_gating_rule_metrics.png")


def chart_gating_llm(llm_df: pd.DataFrame, llm_summary: dict[str, Any]) -> Path:
    fig, axes = plt.subplots(1, 2, figsize=REPORT_FIGSIZE)
    verdict_counts = llm_df["verdict"].value_counts()
    order = [v for v in ["pass", "warn", "block"] if v in verdict_counts.index]
    display_labels = {"pass": "pass", "warn": "warn", "block": "block"}
    v_colors = {"pass": PALETTE["pass"], "warn": PALETTE["warn"], "block": PALETTE["block"]}
    axes[0].bar(
        [display_labels[v] for v in order],
        [int(verdict_counts[v]) for v in order],
        color=[v_colors[v] for v in order],
    )
    axes[0].set_title("Submission Gating — LLM Steering", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[0].set_ylabel("Số phát hiện")
    axes[0].spines["top"].set_visible(False)
    axes[0].spines["right"].set_visible(False)

    contract = int(llm_summary.get("llm_block_contract_violation_count") or 0)
    labels = ["No\nViolation", "Contract\nViolation"]
    values = [len(llm_df) - contract, contract]
    axes[1].bar(labels, values, color=[PALETTE["pass"], PALETTE["fail"]])
    axes[1].set_title("Output Contract", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    axes[1].set_ylabel("Số lượt")
    axes[1].spines["top"].set_visible(False)
    axes[1].spines["right"].set_visible(False)
    for i, v in enumerate(values):
        axes[1].text(i, v, str(v), ha="center", va="bottom", fontsize=9)
    return save_fig(fig, "fig03_gating_llm_verdicts.png")


def chart_track_by_conference(track_df: pd.DataFrame) -> Path:
    counts = track_df["conference"].value_counts().sort_values(ascending=True)
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    ax.barh(counts.index.tolist(), counts.values.tolist(), color=PALETTE["teal"])
    ax.set_xlabel("Số trường hợp")
    ax.set_title("Track Recommendation — phân bố theo hội nghị / track", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for y, v in enumerate(counts.values.tolist()):
        ax.text(v + 0.1, y, str(v), va="center", fontsize=8, color=PALETTE["slate"])
    return save_fig(fig, "fig04_track_by_conference.png")


def chart_chatbot_latency(by_scenario: pd.DataFrame) -> Path:
    labels = by_scenario["label_vi"].tolist()
    duration = by_scenario["avg_duration_s"].tolist()
    answer = by_scenario["avg_time_to_first_answer_s"].tolist()
    ttft = by_scenario["avg_ttft_s"].tolist()
    x = range(len(labels))
    width = 0.28
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    ax.bar([i - width for i in x], ttft, width, label="TTFT", color=PALETTE["mint"])
    ax.bar(list(x), answer, width, label="Time to First Answer Token", color=PALETTE["amber"])
    ax.bar([i + width for i in x], duration, width, label="Total Duration", color=PALETTE["navy"])
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=8)
    ax.set_ylabel("Giây")
    ax.set_title("Chatbot Agent — độ trễ trung bình theo kịch bản", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    ax.legend(frameon=False, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    return save_fig(fig, "fig05_chatbot_latency_by_scenario.png")


def chart_chatbot_tool_success(by_scenario: pd.DataFrame) -> Path:
    df = by_scenario.copy()
    # unsupported may have 0 tool calls
    rates = []
    labels = []
    for _, row in df.iterrows():
        labels.append(row["label_vi"])
        rates.append(row["tool_success_rate_pct"] if pd.notna(row["tool_success_rate_pct"]) else 0.0)
    colors = [PALETTE["pass"] if (r or 0) >= 80 else PALETTE["amber"] if (r or 0) >= 60 else PALETTE["fail"] for r in rates]
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    _bar(
        ax,
        labels,
        [float(r or 0) for r in rates],
        colors,
        "Tool-call Success Rate (%)",
        "Chatbot Agent — Tool-call Success Rate",
        "{:.1f}%",
        25,
    )
    ax.set_ylim(0, 110)
    return save_fig(fig, "fig06_chatbot_tool_success.png")


def chart_chatbot_manual_outcomes() -> Path:
    labels = ["Đạt", "Đạt một phần", "Không đạt"]
    values = [
        CHATBOT_MANUAL_OVERALL["pass"],
        CHATBOT_MANUAL_OVERALL["partial"],
        CHATBOT_MANUAL_OVERALL["fail"],
    ]
    colors = [PALETTE["pass"], PALETTE["partial"], PALETTE["fail"]]
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        colors=colors,
        autopct=lambda p: f"{p:.1f}%\n({int(round(p * sum(values) / 100))})",
        startangle=90,
        wedgeprops={"linewidth": 1, "edgecolor": "white"},
        textprops={"fontsize": 10},
    )
    for t in autotexts:
        t.set_fontsize(9)
        t.set_color("white")
        t.set_fontweight("bold")
    ax.set_title(
        f"Chatbot Agent — kết quả rà soát thủ công ({CHATBOT_MANUAL_OVERALL['trials']} lượt)",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
    )
    return save_fig(fig, "fig07_chatbot_manual_outcomes.png")


def chart_autofill_quality(autofill_df: pd.DataFrame) -> Path:
    metrics = {
        "Tiêu đề khớp\nchính xác": 100.0 * float(autofill_df["title_exact_match"].mean()),
        "Tóm tắt\nROUGE-L": 100.0 * float(autofill_df["abstract_rouge_l"].mean()),
        "Từ khóa\nF1": 100.0 * float(autofill_df["keyword_f1"].mean()),
        "Tác giả\nF1": 100.0 * float(autofill_df["author_f1"].mean()),
    }
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    _bar(
        ax,
        list(metrics.keys()),
        list(metrics.values()),
        [PALETTE["navy"], PALETTE["teal"], PALETTE["mint"], PALETTE["amber"]],
        "Điểm trung bình (%)",
        f"Submission Autofill — đối sánh metadata (n={len(autofill_df)})",
        "{:.1f}%",
        0,
    )
    ax.set_ylim(0, 110)
    return save_fig(fig, "fig08_autofill_quality.png")


def _setup_vietnamese_font() -> None:
    """Prefer a system font that renders Vietnamese diacritics."""
    from matplotlib import font_manager

    candidates = [
        "Segoe UI",
        "Arial",
        "Tahoma",
        "Calibri",
        "DejaVu Sans",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            plt.rcParams["axes.unicode_minus"] = False
            return


def _fmt_int(n: int | float | None) -> str:
    if n is None:
        return "—"
    return f"{int(n):,}".replace(",", ".")


def _draw_workflow_panel(
    *,
    filename: str,
    title: str,
    sample_line: str,
    bars: list[tuple[str, float | None, int | None, str]],
    legend_items: list[tuple[str, str, str]],
) -> Path:
    """One workflow chart: bars left, short legend right (no footnotes / how-to-read)."""
    _setup_vietnamese_font()

    fig = plt.figure(figsize=REPORT_FIGSIZE)
    ax = fig.add_axes([0.10, 0.18, 0.51, 0.66])

    labels = [b[0] for b in bars]
    values = [100.0 * float(b[1] or 0) for b in bars]
    counts = [b[2] for b in bars]
    colors = [PALETTE[b[3]] for b in bars]
    x = range(len(labels))
    bar_rects = ax.bar(list(x), values, color=colors, edgecolor="white", linewidth=0.6, width=0.62)

    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, fontsize=10.5)
    ax.set_ylabel("Tỷ lệ (%)", fontsize=12)
    ax.set_ylim(0, 120)
    ax.tick_params(axis="y", labelsize=11)
    display_title = title.replace(" — ", "\n", 1)
    fig.text(
        0.14,
        0.94,
        display_title,
        fontsize=13.5,
        fontweight="bold",
        color=PALETTE["ink"],
        ha="left",
        va="top",
        linespacing=1.05,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)

    for rect, pct_val, cnt in zip(bar_rects, values, counts):
        label = f"{pct_val:.2f}%"
        if cnt is not None:
            label += f"\n({_fmt_int(cnt)})"
        ax.text(
            rect.get_x() + rect.get_width() / 2,
            rect.get_height() + 1.8,
            label,
            ha="center",
            va="bottom",
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
            linespacing=1.2,
        )

    fig.text(
        0.10,
        0.055,
        sample_line,
        fontsize=12,
        color=PALETTE["slate"],
        transform=fig.transFigure,
    )

    # Right-side legend: short term + one-line gloss only
    legend_ax = fig.add_axes([0.65, 0.18, 0.32, 0.66])
    legend_ax.set_xlim(0, 1)
    legend_ax.set_ylim(0, 1)
    legend_ax.axis("off")
    legend_ax.add_patch(
        plt.Rectangle(
            (0.0, 0.0),
            1.0,
            1.0,
            transform=legend_ax.transAxes,
            facecolor="#F7F5F2",
            edgecolor="#D6D0C8",
            linewidth=1.0,
            zorder=0,
        )
    )
    legend_ax.text(
        0.08,
        0.90,
        "Chú giải",
        fontsize=13,
        fontweight="bold",
        color=PALETTE["ink"],
        va="top",
        transform=legend_ax.transAxes,
    )

    n_items = max(len(legend_items), 1)
    y = 0.76
    step = 0.62 / n_items
    for term, gloss, color_key in legend_items:
        legend_ax.plot(
            0.10,
            y - 0.01,
            marker="s",
            markersize=12,
            color=PALETTE[color_key],
            transform=legend_ax.transAxes,
            linestyle="None",
            clip_on=False,
        )
        legend_ax.text(
            0.18,
            y + 0.02,
            term,
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
            va="top",
            transform=legend_ax.transAxes,
        )
        wrapped = "\n".join(textwrap.wrap(gloss, width=28))
        legend_ax.text(
            0.18,
            y - 0.055,
            wrapped,
            fontsize=11,
            color=PALETTE["slate"],
            va="top",
            transform=legend_ax.transAxes,
            linespacing=1.25,
        )
        y -= step

    return save_fig(fig, filename, apply_tight_layout=False)


def chart_claim_source_reference_flow() -> Path:
    """Render the claim-to-source/reference decision flow as a report asset."""
    _setup_vietnamese_font()
    fig = plt.figure(figsize=REPORT_FIGSIZE)
    ax = fig.add_axes([0.02, 0.03, 0.96, 0.94])
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")

    def box(
        x: float,
        y: float,
        width: float,
        height: float,
        text: str,
        *,
        facecolor: str = "#F7F5F2",
        edgecolor: str = "#355566",
    ) -> None:
        patch = FancyBboxPatch(
            (x - width / 2, y - height / 2),
            width,
            height,
            boxstyle="round,pad=0.012,rounding_size=0.018",
            linewidth=1.5,
            edgecolor=edgecolor,
            facecolor=facecolor,
        )
        ax.add_patch(patch)
        ax.text(
            x,
            y,
            text,
            ha="center",
            va="center",
            fontsize=10.5,
            color=PALETTE["ink"],
            fontweight="bold",
            linespacing=1.15,
        )

    def diamond(x: float, y: float, width: float, height: float, text: str) -> None:
        patch = Polygon(
            [
                (x, y + height / 2),
                (x + width / 2, y),
                (x, y - height / 2),
                (x - width / 2, y),
            ],
            closed=True,
            linewidth=1.5,
            edgecolor="#355566",
            facecolor="#EEF3F4",
        )
        ax.add_patch(patch)
        ax.text(
            x,
            y,
            text,
            ha="center",
            va="center",
            fontsize=9.8,
            color=PALETTE["ink"],
            fontweight="bold",
            linespacing=1.1,
        )

    def arrow(start: tuple[float, float], end: tuple[float, float]) -> None:
        ax.annotate(
            "",
            xy=end,
            xytext=start,
            arrowprops={
                "arrowstyle": "-|>",
                "color": "#355566",
                "linewidth": 1.5,
                "shrinkA": 0,
                "shrinkB": 0,
            },
        )

    box(0.11, 0.86, 0.18, 0.11, "Mệnh đề\nđầu ra")
    box(0.37, 0.86, 0.22, 0.11, "Đối chiếu với\nnguồn dữ liệu")
    diamond(0.66, 0.86, 0.22, 0.16, "Nguồn dữ liệu\ncó hỗ trợ?")
    box(
        0.89,
        0.86,
        0.18,
        0.11,
        "Không vượt bước\nbám nguồn",
        facecolor="#FBECE8",
        edgecolor=PALETTE["coral"],
    )

    box(0.66, 0.68, 0.20, 0.11, "Mệnh đề\nbám nguồn", facecolor="#EAF5F3", edgecolor=PALETTE["teal"])
    diamond(0.66, 0.49, 0.24, 0.17, "Có văn bản tham chiếu\nphù hợp?")
    box(
        0.89,
        0.49,
        0.18,
        0.11,
        "Chỉ báo cáo\nmức bám nguồn",
        facecolor="#F4F1E8",
        edgecolor=PALETTE["amber"],
    )

    box(0.66, 0.29, 0.24, 0.11, "Đối chiếu với\nvăn bản tham chiếu")
    box(
        0.49,
        0.10,
        0.24,
        0.11,
        "Được văn bản tham chiếu\nghi nhận",
        facecolor="#EAF5F3",
        edgecolor=PALETTE["teal"],
    )
    box(
        0.82,
        0.10,
        0.25,
        0.11,
        "Không được văn bản tham chiếu\nghi nhận",
        facecolor="#FBECE8",
        edgecolor=PALETTE["coral"],
    )

    arrow((0.20, 0.86), (0.26, 0.86))
    arrow((0.48, 0.86), (0.55, 0.86))
    arrow((0.77, 0.86), (0.80, 0.86))
    ax.text(0.785, 0.89, "Không", fontsize=9.5, color=PALETTE["slate"], ha="center")
    arrow((0.66, 0.78), (0.66, 0.735))
    ax.text(0.68, 0.755, "Có", fontsize=9.5, color=PALETTE["slate"], va="center")
    arrow((0.66, 0.625), (0.66, 0.575))
    arrow((0.78, 0.49), (0.80, 0.49))
    ax.text(0.79, 0.52, "Không", fontsize=9.5, color=PALETTE["slate"], ha="center")
    arrow((0.66, 0.405), (0.66, 0.345))
    ax.text(0.68, 0.375, "Có", fontsize=9.5, color=PALETTE["slate"], va="center")
    arrow((0.61, 0.235), (0.52, 0.155))
    arrow((0.71, 0.235), (0.79, 0.155))

    return save_fig(
        fig,
        "fig09_claim_source_reference_flow.png",
        apply_tight_layout=False,
    )


def chart_tca_workflow_panels(tca_summary: dict[str, Any]) -> list[tuple[Path, str, str]]:
    """One figure per workflow aspect; returns (path, content, placement) for index."""
    s = tca_summary
    flow = chart_claim_source_reference_flow()
    outputs: list[tuple[Path, str, str]] = [
        (flow, "Hậu kiểm mệnh đề--nguồn", "Thiết lập benchmark")
    ]

    # 1) Phân tích ban đầu cho phản biện — annotation
    p = _draw_workflow_panel(
        filename="fig09a_reviewer_annotation_grounding.png",
        title="Reviewer Initial Analysis — Mức bám nguồn của chú thích",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b1_paper_count'))} bài  ·  "
            f"Tổng số chú thích: {_fmt_int(s.get('b1_annotation_total'))}"
        ),
        bars=[
            ("Khớp gần\nnguyên văn", s.get("b1_exact_rate_mean"), s.get("b1_exact_count"), "navy"),
            ("Khớp\nmột phần", s.get("b1_partial_rate_mean"), s.get("b1_partial_count"), "teal"),
            (
                "Không xác định\nđược trong bài",
                s.get("b1_fabricated_rate_mean"),
                s.get("b1_fabricated_count"),
                "coral",
            ),
        ],
        legend_items=[
            ("Khớp gần nguyên văn", "Trùng gần như nguyên văn với nội dung bài", "navy"),
            ("Khớp một phần", "Có cơ sở trong bài nhưng chỉ khớp một phần", "teal"),
            ("Không xác định được", "Không tìm thấy đoạn tương ứng trong bài", "coral"),
        ],
    )
    outputs.append((p, "Mức bám nguồn của chú thích", "Reviewer Initial Analysis"))

    # 2) Phân tích ban đầu — điểm cần chú ý
    p = _draw_workflow_panel(
        filename="fig09b_reviewer_attention_points.png",
        title="Reviewer Initial Analysis — Điểm cần chú ý khi đọc bài",
        sample_line=(
            f"Mức bám nguồn: {_fmt_int(s.get('b2_paper_count'))} bài  ·  "
            f"Đối chiếu bản phản biện: {_fmt_int(s.get('b2_reference_eligible_paper_count'))} bài"
        ),
        bars=[
            ("Bám nguồn\nbản thảo", s.get("b2_t_rate_mean"), s.get("b2_t_count_est"), "navy"),
            (
                "Được ghi nhận\ntrong phản biện",
                s.get("b2_c_rate_mean"),
                s.get("b2_c_count_est"),
                "teal",
            ),
            (
                "Không được ghi nhận\ntrong phản biện",
                s.get("b2_a_rate_mean"),
                s.get("b2_a_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            ("Bám nguồn bản thảo", "Điểm cần lưu ý được nội dung bản thảo hỗ trợ", "navy"),
            (
                "Được ghi nhận",
                "Mệnh đề bám nguồn cũng xuất hiện trong các bản phản biện",
                "teal",
            ),
            (
                "Không được ghi nhận",
                "Mệnh đề bám nguồn không xuất hiện trong các bản phản biện",
                "coral",
            ),
        ],
    )
    outputs.append((p, "Điểm cần chú ý khi đọc bài", "Reviewer Initial Analysis"))

    # 3) Kiểm tra chất lượng phản biện
    p = _draw_workflow_panel(
        filename="fig09c_review_quality_auditor.png",
        title="Review Quality Auditor — Phát hiện trên bản phản biện",
        sample_line=(
            f"Cỡ mẫu: {_fmt_int(s.get('b3_paper_count'))} bài  ·  "
            f"{_fmt_int(s.get('b3_review_total'))} bản phản biện  ·  "
            f"{_fmt_int(s.get('b3_finding_total'))} phát hiện"
        ),
        bars=[
            (
                "Bám nguồn\nbản phản biện",
                s.get("b3_truthfulness_mean"),
                s.get("b3_t_count_est"),
                "navy",
            ),
            (
                "Mã vấn đề\nhợp lệ",
                s.get("b3_validity_mean"),
                s.get("b3_v_count_est"),
                "teal",
            ),
            (
                "Đồng thời đạt\ncả hai",
                s.get("b3_grounded_valid_mean"),
                s.get("b3_gv_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            (
                "Bám nguồn bản phản biện",
                "Phần giải thích bám vào bản phản biện đang được kiểm tra",
                "navy",
            ),
            (
                "Mã vấn đề hợp lệ",
                "Mã được gán phù hợp với đặc điểm quan sát được",
                "teal",
            ),
            (
                "Đồng thời đạt cả hai",
                "Phát hiện đạt cả mức bám nguồn và tính hợp lệ của mã",
                "coral",
            ),
        ],
    )
    outputs.append((p, "Phát hiện trên bản phản biện", "Review Quality Auditor"))

    # 4) Hỗ trợ quyết định chủ tịch
    p = _draw_workflow_panel(
        filename="fig09d_chair_evidence_basis.png",
        title="Chair Decision Copilot — Cơ sở bằng chứng",
        sample_line=(
            f"Mức bám nguồn: {_fmt_int(s.get('b5_paper_count'))} bài  ·  "
            f"Đối chiếu metareview: {_fmt_int(s.get('b5_reference_eligible_paper_count'))} bài"
        ),
        bars=[
            (
                "Bám nguồn\ndữ liệu",
                s.get("b5_evidence_basis_t_rate_mean"),
                s.get("b5_t_count_est"),
                "navy",
            ),
            (
                "Được ghi nhận\ntrong metareview",
                s.get("b5_evidence_basis_c_rate_mean"),
                s.get("b5_c_count_est"),
                "teal",
            ),
            (
                "Không được ghi nhận\ntrong metareview",
                s.get("b5_evidence_basis_a_rate_mean"),
                s.get("b5_a_count_est"),
                "coral",
            ),
        ],
        legend_items=[
            ("Bám nguồn dữ liệu", "Nhận định bám dữ liệu nguồn có sẵn", "navy"),
            (
                "Được ghi nhận",
                "Mệnh đề bám nguồn cũng xuất hiện trong metareview",
                "teal",
            ),
            (
                "Không được ghi nhận",
                "Mệnh đề bám nguồn không xuất hiện trong metareview",
                "coral",
            ),
        ],
    )
    outputs.append((p, "Cơ sở bằng chứng", "Chair Decision Copilot"))

    return outputs


def chart_workflow_headline(summaries: dict[str, dict[str, Any]]) -> Path:
    """Vẽ các chỉ số chính dùng trong phần tổng hợp kết quả."""
    labels = [
        "Submission Gating\nRule Check",
        "Submission Gating\nLLM Steering",
        "Chatbot Agent\nmanual pass",
        "Chatbot Agent\ntool-call success",
        "Submission Autofill\nROUGE-L",
        "Chú thích\nkhớp nguyên văn",
    ]
    rule = summaries["rule"]
    llm = summaries["llm"]
    chat = summaries["chatbot"]
    auto = summaries["autofill"]
    tca = summaries["tca"]
    values = [
        100.0 * float(rule.get("blocking_verdict_accuracy") or 0),
        100.0
        * (
            1.0
            - float(llm.get("llm_block_contract_violation_count") or 0)
            / max(int(llm.get("review_rows") or 1), 1)
        ),
        float(chat.get("manual_pass_rate_pct") or 0),
        100.0 * float(chat.get("tool_call_success_rate") or 0),
        100.0 * float(auto.get("abstract_rouge_l_mean") or 0),
        100.0 * float(tca.get("b1_exact_rate_mean") or 0),
    ]
    colors = [PALETTE["pass"], PALETTE["mint"], PALETTE["amber"], PALETTE["teal"], PALETTE["navy"], PALETTE["ink"]]
    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    _bar(ax, labels, values, colors, "Tỷ lệ (%)", "Tóm tắt các chỉ số chính", "{:.1f}%", 0)
    ax.set_ylim(0, 115)
    return save_fig(fig, "fig10_headline_metrics.png")


def operational_resource_df() -> pd.DataFrame:
    """Return operational metrics with reader-facing Vietnamese column names."""
    return pd.DataFrame(OPERATIONAL_RESOURCE_ROWS).rename(
        columns={
            "workflow_id": "Mã luồng xử lý",
            "label_vi": "Luồng xử lý",
            "label_short": "Nhãn biểu đồ",
            "unit": "Đơn vị công việc",
            "latency_mean_s": "Độ trễ trung bình (giây)",
            "latency_median_s": "Độ trễ trung vị (giây)",
            "latency_max_s": "Độ trễ cao nhất (giây)",
            "token_mean": "Token trung bình",
            "ops_mode": "Mã cơ chế vận hành",
            "ops_mode_vi": "Cơ chế vận hành đề xuất",
            "source": "Nguồn",
            "note": "Ghi chú",
        }
    )


def _fmt_latency(val: float) -> str:
    if val < 1:
        return f"{val:.2f}"
    if val < 100:
        return f"{val:.1f}"
    return f"{val:.0f}"


def chart_resource_latency() -> Path:
    """Horizontal mean latency with max whisker and median marker (Table 4.7)."""
    _setup_vietnamese_font()
    rows = OPERATIONAL_RESOURCE_ROWS
    ordered = sorted(rows, key=lambda r: float(r["latency_mean_s"]))
    labels = [r.get("label_chart") or r["label_short"] for r in ordered]
    means = [float(r["latency_mean_s"]) for r in ordered]
    medians = [
        float(r["latency_median_s"]) if r["latency_median_s"] is not None else None for r in ordered
    ]
    maxes = [float(r["latency_max_s"]) for r in ordered]
    xmax = max(maxes)

    fig = plt.figure(figsize=REPORT_FIGSIZE_DENSE)
    # Data left, legend panel right — never overlay the bars.
    ax = fig.add_axes([0.24, 0.16, 0.50, 0.74])
    y = list(range(len(labels)))

    ax.barh(
        y,
        means,
        color=PALETTE["navy"],
        edgecolor="white",
        linewidth=0.5,
        height=0.62,
        zorder=2,
    )
    for yi, mean_v, max_v, med_v in zip(y, means, maxes, medians):
        ax.plot(
            [mean_v, max_v],
            [yi, yi],
            color=PALETTE["coral"],
            linewidth=1.6,
            solid_capstyle="butt",
            zorder=3,
        )
        ax.plot(
            max_v,
            yi,
            marker="|",
            color=PALETTE["coral"],
            markersize=14,
            markeredgewidth=1.8,
            zorder=4,
        )
        if med_v is not None:
            ax.plot(
                med_v,
                yi,
                marker="D",
                color=PALETTE["teal"],
                markersize=7,
                markeredgecolor="white",
                markeredgewidth=0.6,
                zorder=5,
            )

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=11)
    ax.set_xlabel("Độ trễ (giây)", fontsize=12)
    ax.set_title(
        "Độ trễ của từng luồng xử lý AI",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
        loc="left",
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
    ax.axvline(100.0, color=PALETTE["amber"], linestyle=":", linewidth=1.3, alpha=0.9, zorder=1)
    ax.set_xlim(0, xmax * 1.22)
    ax.set_ylim(-0.7, len(labels) - 0.3)

    for yi, mean_v, max_v in zip(y, means, maxes):
        ax.text(
            max_v + xmax * 0.018,
            yi,
            f"{_fmt_latency(mean_v)}  |  {_fmt_latency(max_v)}",
            va="center",
            ha="left",
            fontsize=9,
            color=PALETTE["ink"],
        )

    legend_ax = fig.add_axes([0.78, 0.34, 0.18, 0.40])
    legend_ax.set_xlim(0, 1)
    legend_ax.set_ylim(0, 1)
    legend_ax.axis("off")
    legend_ax.add_patch(
        plt.Rectangle(
            (0.0, 0.0),
            1.0,
            1.0,
            transform=legend_ax.transAxes,
            facecolor="#F7F5F2",
            edgecolor="#D6D0C8",
            linewidth=1.0,
            zorder=0,
        )
    )
    legend_ax.text(
        0.10,
        0.90,
        "Chú giải",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        va="top",
        transform=legend_ax.transAxes,
    )
    legend_items = [
        (PALETTE["navy"], "s", "Trung bình"),
        (PALETTE["teal"], "D", "Trung vị"),
        (PALETTE["coral"], "|", "Cao nhất"),
        (PALETTE["amber"], ":", "Ngưỡng 100 giây"),
    ]
    y_leg = 0.72
    for color, marker, lab in legend_items:
        if marker == ":":
            legend_ax.plot(
                [0.08, 0.18],
                [y_leg, y_leg],
                color=color,
                linestyle=":",
                linewidth=1.6,
                transform=legend_ax.transAxes,
                clip_on=False,
            )
        elif marker == "|":
            legend_ax.plot(
                [0.08, 0.18],
                [y_leg, y_leg],
                color=color,
                linewidth=1.6,
                transform=legend_ax.transAxes,
                clip_on=False,
            )
            legend_ax.plot(
                0.18,
                y_leg,
                marker="|",
                color=color,
                markersize=12,
                markeredgewidth=1.6,
                transform=legend_ax.transAxes,
                clip_on=False,
            )
        elif marker == "s":
            legend_ax.plot(
                0.13,
                y_leg,
                marker="s",
                color=color,
                markersize=11,
                transform=legend_ax.transAxes,
                linestyle="None",
                clip_on=False,
            )
        else:
            legend_ax.plot(
                0.13,
                y_leg,
                marker=marker,
                color=color,
                markersize=8,
                markeredgecolor="white",
                markeredgewidth=0.5,
                transform=legend_ax.transAxes,
                linestyle="None",
                clip_on=False,
            )
        legend_ax.text(
            0.26,
            y_leg,
            lab,
            fontsize=10,
            color=PALETTE["ink"],
            va="center",
            transform=legend_ax.transAxes,
        )
        y_leg -= 0.16

    fig.text(
        0.50,
        0.045,
        "Nhãn số: trung bình | cao nhất (giây). Đơn vị: một bài, một lượt kiểm tra hoặc một hội thoại.\n"
        "Chatbot Agent chưa có độ trễ trung vị trong bảng tổng hợp.",
        ha="center",
        fontsize=10,
        color=PALETTE["slate"],
        linespacing=1.35,
    )
    return save_fig(
        fig,
        "fig11_resource_latency.png",
        figsize=REPORT_FIGSIZE_DENSE,
        apply_tight_layout=False,
    )


def chart_resource_tokens() -> Path:
    """Vẽ số token trung bình trên mỗi đơn vị công việc có số liệu."""
    _setup_vietnamese_font()
    rows = [r for r in OPERATIONAL_RESOURCE_ROWS if r.get("token_mean") is not None]
    labels = [r.get("label_chart") or r["label_short"] for r in rows]
    tokens = [float(r["token_mean"]) for r in rows]
    units = [str(r["unit"]) for r in rows]
    colors = [PALETTE["navy"], PALETTE["teal"], PALETTE["mint"], PALETTE["amber"], PALETTE["ink"]]
    colors = colors[: len(labels)]

    fig, ax = plt.subplots(figsize=REPORT_FIGSIZE)
    x = list(range(len(labels)))
    bars = ax.bar(x, tokens, color=colors, edgecolor="white", linewidth=0.6, width=0.62)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=10)
    ax.set_ylabel("Token trung bình / đơn vị công việc", fontsize=11)
    ax.set_title(
        "Lượng token trung bình theo luồng xử lý AI",
        fontsize=13,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ymax = max(tokens) * 1.22 if tokens else 1
    ax.set_ylim(0, ymax)

    for bar, val, unit in zip(bars, tokens, units):
        if val >= 1000:
            txt = f"{val:,.0f}".replace(",", ".")
        else:
            txt = f"{val:.1f}"
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"{txt}\n/{unit}",
            ha="center",
            va="bottom",
            fontsize=8,
            color=PALETTE["slate"],
        )

    fig.text(
        0.5,
        0.04,
        "Lưu ý: mẫu số khác nhau (bài · lượt kiểm tra · hội thoại); không dùng token để xếp hạng chi phí giữa các luồng.\n"
        "Submission Gating và Track Recommendation chưa có số token trong bảng tổng hợp.",
        ha="center",
        fontsize=9,
        color=PALETTE["slate"],
        linespacing=1.25,
    )
    return save_fig(
        fig,
        "fig12_resource_tokens.png",
        layout_rect=(0.08, 0.14, 0.98, 0.96),
    )


def chart_resource_ops_mode() -> Path:
    """Mean latency colored by recommended ops mode; max as whisker; legend outside."""
    _setup_vietnamese_font()
    rows = OPERATIONAL_RESOURCE_ROWS
    mode_order = [
        ("sync", "Đồng bộ", PALETTE["pass"]),
        ("non_blocking", "Không chặn", PALETTE["mint"]),
        ("interactive_with_timeout", "Tương tác + giới hạn chờ", PALETTE["amber"]),
        ("streaming", "Phản hồi từng phần + trạng thái", PALETTE["teal"]),
        ("background", "Xử lý nền / chạy trước", PALETTE["coral"]),
    ]
    mode_color = {m[0]: m[2] for m in mode_order}

    ordered = sorted(rows, key=lambda r: float(r["latency_mean_s"]))
    labels = [r.get("label_chart") or r["label_short"] for r in ordered]
    means = [float(r["latency_mean_s"]) for r in ordered]
    colors = [mode_color[r["ops_mode"]] for r in ordered]
    maxes = [float(r["latency_max_s"]) for r in ordered]

    fig = plt.figure(figsize=REPORT_FIGSIZE_DENSE)
    # Chart left, legend panel right — same discipline as TCA panels / fig01.
    ax = fig.add_axes([0.24, 0.18, 0.48, 0.72])
    y = list(range(len(labels)))
    bars = ax.barh(y, means, color=colors, edgecolor="white", linewidth=0.5, height=0.62, zorder=2)
    for yi, mean_v, max_v in zip(y, means, maxes):
        ax.plot(
            [mean_v, max_v],
            [yi, yi],
            color=PALETTE["slate"],
            linewidth=1.2,
            alpha=0.7,
            zorder=3,
        )
        ax.plot(
            max_v,
            yi,
            marker="|",
            color=PALETTE["slate"],
            markersize=12,
            markeredgewidth=1.4,
            zorder=4,
        )

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=11)
    ax.set_xlabel("Độ trễ trung bình (giây)", fontsize=11)
    ax.set_title(
        "Khuyến nghị vận hành theo độ trễ luồng AI",
        fontsize=15,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
        loc="left",
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
    # Threshold lines only — no floating tags near the title.
    ax.axvline(5.0, color=PALETTE["pass"], linestyle=":", linewidth=1.1, alpha=0.75, zorder=1)
    ax.axvline(30.0, color=PALETTE["coral"], linestyle=":", linewidth=1.1, alpha=0.75, zorder=1)
    ax.set_xlim(0, max(maxes) * 1.12)
    ax.set_ylim(-0.7, len(labels) - 0.3)

    for bar, val in zip(bars, means):
        if val < 1:
            txt = f"{val:.2f}"
        else:
            txt = f"{val:.1f}"
        ax.text(
            bar.get_width() + max(maxes) * 0.02,
            bar.get_y() + bar.get_height() / 2,
            txt,
            va="center",
            fontsize=10,
            color=PALETTE["ink"],
            fontweight="bold",
        )

    used_modes = {r["ops_mode"] for r in rows}
    legend_ax = fig.add_axes([0.76, 0.30, 0.20, 0.48])
    legend_ax.set_xlim(0, 1)
    legend_ax.set_ylim(0, 1)
    legend_ax.axis("off")
    legend_ax.add_patch(
        plt.Rectangle(
            (0.0, 0.0),
            1.0,
            1.0,
            transform=legend_ax.transAxes,
            facecolor="#F7F5F2",
            edgecolor="#D6D0C8",
            linewidth=1.0,
            zorder=0,
        )
    )
    legend_ax.text(
        0.08,
        0.92,
        "Cách vận hành",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        va="top",
        transform=legend_ax.transAxes,
    )
    legend_items = [(lab, col) for mid, lab, col in mode_order if mid in used_modes]
    n_items = max(len(legend_items), 1)
    y_leg = 0.78
    step = 0.58 / n_items
    for lab, col in legend_items:
        legend_ax.plot(
            0.10,
            y_leg,
            marker="s",
            markersize=11,
            color=col,
            transform=legend_ax.transAxes,
            linestyle="None",
            clip_on=False,
        )
        legend_ax.text(
            0.20,
            y_leg,
            lab,
            fontsize=10,
            color=PALETTE["ink"],
            va="center",
            transform=legend_ax.transAxes,
        )
        y_leg -= step

    legend_ax.text(
        0.08,
        0.12,
        "Nét đứt: 5 giây và 30 giây\n"
        "Dấu | : độ trễ cao nhất",
        fontsize=9,
        color=PALETTE["slate"],
        va="bottom",
        transform=legend_ax.transAxes,
        linespacing=1.35,
    )

    fig.text(
        0.50,
        0.045,
        "Rule Check: đồng bộ · LLM Steering: không chặn · Autofill: tương tác có giới hạn chờ ·\n"
        "Reviewer / Quality / Chair: xử lý nền · Chatbot Agent: phản hồi từng phần và hiển thị trạng thái.",
        ha="center",
        fontsize=10,
        color=PALETTE["slate"],
        linespacing=1.35,
    )
    return save_fig(
        fig,
        "fig13_resource_ops_mode.png",
        figsize=REPORT_FIGSIZE_DENSE,
        apply_tight_layout=False,
    )


def chart_chatbot_response_stages() -> Path:
    """Vẽ các giai đoạn phản hồi và tỷ lệ tool-call thành công của Chatbot Agent."""
    _setup_vietnamese_font()
    s = CHATBOT_RESPONSE_STAGES
    labels = [
        "TTFT",
        "Time to First\nAnswer Token",
        "Stream Duration",
        "Total Duration",
    ]
    values = [
        s["ttft_s"],
        s["first_answer_token_s"],
        s["stream_duration_s"],
        s["total_duration_s"],
    ]
    colors = [PALETTE["mint"], PALETTE["amber"], PALETTE["teal"], PALETTE["navy"]]

    fig, axes = plt.subplots(
        1,
        2,
        figsize=REPORT_FIGSIZE,
        gridspec_kw={"width_ratios": [1.35, 1.0]},
    )

    ax = axes[0]
    x = list(range(len(labels)))
    bars = ax.bar(x, values, color=colors, edgecolor="white", linewidth=0.6, width=0.62)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel("Giây", fontsize=10)
    ax.set_title(
        "Chatbot Agent — các giai đoạn phản hồi",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=8,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"{val:.2f} giây",
            ha="center",
            va="bottom",
            fontsize=9,
            color=PALETTE["slate"],
        )

    ax2 = axes[1]
    tool_ok = s["tool_calls_total"] - s["tool_calls_failed"]
    tool_fail = s["tool_calls_failed"]
    wedges, _, autotexts = ax2.pie(
        [tool_ok, tool_fail],
        labels=["Thành công", "Thất bại"],
        colors=[PALETTE["pass"], PALETTE["fail"]],
        autopct=lambda p: f"{p:.1f}%",
        startangle=90,
        textprops={"fontsize": 9, "color": PALETTE["ink"]},
        wedgeprops={"edgecolor": "white", "linewidth": 1.2},
    )
    for t in autotexts:
        t.set_color("white")
        t.set_fontweight("bold")
        t.set_fontsize(9)
    ax2.set_title(
        f"Kết quả gọi công cụ\n({int(tool_ok)}/{int(s['tool_calls_total'])} thành công · "
        f"{int(tool_fail)} thất bại)",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=8,
    )

    fig.text(
        0.5,
        0.04,
        "TTFT trung bình đạt 2,36 giây, nhưng Time to First Answer Token trung bình là 23,02 giây.\n"
        "Giao diện cần hiển thị trạng thái đang tra cứu hoặc tổng hợp; hệ thống cần giảm 31 lượt gọi công cụ thất bại.",
        ha="center",
        fontsize=9,
        color=PALETTE["slate"],
        linespacing=1.25,
    )
    return save_fig(
        fig,
        "fig14_chatbot_response_stages.png",
        layout_rect=(0.08, 0.14, 0.98, 0.96),
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def build_overview_sheet(
    track_s: dict[str, Any],
    rule_s: dict[str, Any],
    llm_s: dict[str, Any],
    chat_s: dict[str, Any],
    auto_s: dict[str, Any],
    tca_s: dict[str, Any],
) -> pd.DataFrame:
    rows = [
        {
            "workflow": "track_recommendation",
            "n": track_s["case_count"],
            "primary_metric": "invalid_track_rate",
            "primary_value": track_s["invalid_track_rate"],
            "secondary_metric": "human_label_filled",
            "secondary_value": track_s["human_label_filled"],
            "status": "completed_no_human_accuracy",
            "limitation": track_s["limitation"],
        },
        {
            "workflow": "submission_gating_rule_check",
            "n": rule_s["case_count"],
            "primary_metric": "blocking_verdict_accuracy",
            "primary_value": rule_s["blocking_verdict_accuracy"],
            "secondary_metric": "rule_id_recall",
            "secondary_value": rule_s["rule_id_recall"],
            "status": "completed",
            "limitation": "",
        },
        {
            "workflow": "submission_gating_llm_steering",
            "n": llm_s["case_count"],
            "primary_metric": "llm_block_contract_violation_count",
            "primary_value": llm_s["llm_block_contract_violation_count"],
            "secondary_metric": "content_finding_count",
            "secondary_value": llm_s["content_finding_count"],
            "status": "completed_no_human_labels",
            "limitation": llm_s["limitation"],
        },
        {
            "workflow": "chatbot_agent",
            "n": chat_s["trial_count"],
            "primary_metric": "manual_pass_rate_pct",
            "primary_value": chat_s["manual_pass_rate_pct"],
            "secondary_metric": "tool_call_success_rate",
            "secondary_value": chat_s["tool_call_success_rate"],
            "status": "completed_with_manual_review",
            "limitation": "manual labels from report, not auto-scored quality",
        },
        {
            "workflow": "submission_autofill",
            "n": auto_s["case_count"],
            "primary_metric": "abstract_rouge_l_mean",
            "primary_value": auto_s["abstract_rouge_l_mean"],
            "secondary_metric": "title_exact_match_rate",
            "secondary_value": auto_s["title_exact_match_rate"],
            "status": "completed",
            "limitation": "",
        },
        {
            "workflow": "tca_benchmark",
            "n": tca_s["paper_count"],
            "primary_metric": "annotation_exact_match_rate",
            "primary_value": tca_s["b1_exact_rate_mean"],
            "secondary_metric": "auditor_grounded_valid_rate",
            "secondary_value": tca_s["b3_grounded_valid_mean"],
            "status": "completed",
            "limitation": tca_s.get("note", ""),
        },
    ]
    return pd.DataFrame(rows)


def write_figure_index(paths: list[tuple[str, str, str]]) -> None:
    lines = [
        "# Danh mục hình — kết quả đánh giá các luồng xử lý",
        "",
        "Sinh tự động từ `scripts/export_benchmark_to_excel.py`.",
        "Dùng các tệp PNG trong `exports/figures/` cho Chương 4 và Chương 5.",
        "",
        "| File | Nội dung | Gợi ý chỗ dùng |",
        "| --- | --- | --- |",
    ]
    for fname, content, placement in paths:
        lines.append(f"| `{fname}` | {content} | {placement} |")
    lines.extend(
        [
            "",
            "## Tệp Excel tổng hợp",
            "",
            "- `workflow_benchmark_results.xlsx`",
            "  - `Overview` — các chỉ số chính theo luồng xử lý",
            "  - `Track_Recommendation` — 48 trường hợp Track Recommendation",
            "  - `Gating_Rule_Check` — 8 trường hợp kiểm tra quy tắc cố định",
            "  - `Gating_LLM_Steering` — các phát hiện kiểm tra nội dung",
            "  - `Chatbot_Trials` — số liệu truyền tải của 40 lượt hội thoại",
            "  - `Chatbot_By_Scenario` — tổng hợp theo kịch bản và kết quả rà soát",
            "  - `Autofill_Summary` + `Autofill_Cases` — số liệu đối sánh tệp CSV",
            "  - `TCA_Summary` + `TCA_Papers` — tỷ lệ đánh giá TCA theo bài",
            "  - `Resource_Usage` — Bảng 4.7: độ trễ trung bình/trung vị/cao nhất, token và cách vận hành",
            "  - `Chatbot_Response_Stages` — thời gian phản hồi và các lượt gọi công cụ thất bại",
            "",
            "## Lưu ý khi trích dẫn vào báo cáo",
            "",
            "1. Track Recommendation: `human_label` trống → **không** suy ra Top-1 Accuracy từ tệp này.",
            "2. Submission Gating — LLM Steering: `grounded/actionable` trống → chỉ kết luận về Output Contract Violation.",
            "3. Chatbot Agent: số liệu thời gian lấy từ `run_summary.json`; kết quả đạt/đạt một phần/không đạt lấy từ báo cáo rà soát thủ công.",
            "4. Hình 09a–09d: mỗi luồng có một hình; nhãn gồm tỷ lệ và số lượng; cỡ mẫu đặt dưới trục; chú giải đặt bên phải.",
            "5. Hình 09a dùng số lượng tuyệt đối. Các hình còn lại ghi số lượng ước tính trong ngoặc theo tỷ lệ trên tổng đơn vị.",
            "6. Hình 11–14: số liệu khớp Bảng 4.7 và các báo cáo luồng; **không** xếp hạng chi phí bằng token "
            "vì mẫu số khác nhau (bài / lượt kiểm tra / hội thoại). Độ trễ cao nhất của Chatbot Agent là giá trị của nhóm kịch bản chậm nhất.",
            "7. Hình 13: cách vận hành là **khuyến nghị thiết kế** dựa trên độ trễ, không phải kết quả so sánh triển khai.",
            "",
        ]
    )
    (EXPORT / "FIGURE_INDEX.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    EXPORT.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)

    track_df, track_s = load_track()
    rule_df, rule_s = load_gating_rule()
    llm_df, llm_s = load_gating_llm()
    chat_trials, chat_by_sc, chat_s = load_chatbot()
    auto_df, auto_s = load_autofill()
    tca_df, tca_s = load_tca()

    overview = build_overview_sheet(track_s, rule_s, llm_s, chat_s, auto_s, tca_s)

    # Compact autofill for Excel (full 1127 rows is fine; keep numeric cols first)
    auto_cols = [
        c
        for c in [
            "paper_id",
            "conference_year_track",
            "title_exact_match",
            "abstract_rouge_1",
            "abstract_rouge_l",
            "keyword_f1",
            "author_f1",
            "time_taken_seconds",
            "total_inference_time",
            "input_tokens",
            "output_tokens",
            "total_tokens",
        ]
        if c in auto_df.columns
    ]
    auto_compact = auto_df[auto_cols].copy()

    xlsx_path = EXPORT / "workflow_benchmark_results.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        write_df(writer, overview, "Overview")
        write_df(writer, pd.DataFrame([track_s]), "Track_Summary")
        write_df(writer, track_df, "Track_Recommendation")
        write_df(writer, pd.DataFrame([rule_s]), "Gating_Rule_Summary")
        write_df(writer, rule_df, "Gating_Rule_Check")
        write_df(writer, pd.DataFrame([llm_s]), "Gating_LLM_Summary")
        write_df(writer, llm_df, "Gating_LLM_Steering")
        write_df(writer, pd.DataFrame([chat_s]), "Chatbot_Summary")
        write_df(writer, chat_by_sc, "Chatbot_By_Scenario")
        write_df(writer, chat_trials, "Chatbot_Trials")
        write_df(writer, pd.DataFrame([auto_s]), "Autofill_Summary")
        write_df(writer, auto_compact, "Autofill_Cases")
        write_df(writer, pd.DataFrame([tca_s]), "TCA_Summary")
        write_df(writer, tca_df, "TCA_Papers")
        write_df(writer, operational_resource_df(), "Resource_Usage")
        chatbot_stage_df = pd.DataFrame(
            [
                {
                    "TTFT (giây)": CHATBOT_RESPONSE_STAGES["ttft_s"],
                    "Time to First Answer Token (giây)": CHATBOT_RESPONSE_STAGES[
                        "first_answer_token_s"
                    ],
                    "Stream Duration (giây)": CHATBOT_RESPONSE_STAGES["stream_duration_s"],
                    "Total Duration (giây)": CHATBOT_RESPONSE_STAGES["total_duration_s"],
                    "Tổng số lượt gọi công cụ": CHATBOT_RESPONSE_STAGES["tool_calls_total"],
                    "Số lượt gọi công cụ thất bại": CHATBOT_RESPONSE_STAGES["tool_calls_failed"],
                    "Tỷ lệ gọi công cụ thành công (%)": CHATBOT_RESPONSE_STAGES[
                        "tool_success_rate_pct"
                    ],
                }
            ]
        )
        write_df(writer, chatbot_stage_df, "Chatbot_Response_Stages")

    # Charts
    fig_meta: list[tuple[str, str, str]] = []
    p = chart_overview_case_counts(
        runner_n=int(auto_s["case_count"]),  # shared runner corpus = 1.127 papers
        tca_n=int(tca_s["paper_count"]),
        track_n=int(track_s["case_count"]),
        rule_n=int(rule_s["case_count"]),
        llm_n=int(llm_s["case_count"]),
        chat_n=int(chat_s["trial_count"]),
    )
    fig_meta.append(
        (
            p.name,
            "Quy mô các bộ dữ liệu đánh giá: bộ thực thi 1.127, bộ TCA 1.097 và các bộ chuyên biệt",
            "Chương 4 — thiết lập đánh giá",
        )
    )

    p = chart_runner_by_conference(auto_df)
    fig_meta.append(
        (
            p.name,
            "Phân bố 1.127 bài đầu vào theo hội nghị / track",
            "Chương 4 — thiết lập dữ liệu",
        )
    )

    p = chart_gating_rule(rule_df, rule_s)
    fig_meta.append((p.name, "Verdict Accuracy, Rule-ID Recall, False Block Count và Verdict Distribution", "Submission Gating — Rule Check"))

    p = chart_gating_llm(llm_df, llm_s)
    fig_meta.append((p.name, "Verdict Distribution và Output Contract Violation", "Submission Gating — LLM Steering"))

    p = chart_track_by_conference(track_df)
    fig_meta.append((p.name, "Phân bố trường hợp Track Recommendation theo hội nghị / track", "Submission Autofill — Track Recommendation"))

    p = chart_chatbot_latency(chat_by_sc)
    fig_meta.append((p.name, "TTFT, Time to First Answer Token và Total Duration theo kịch bản", "Chatbot Agent — độ trễ"))

    p = chart_chatbot_tool_success(chat_by_sc)
    fig_meta.append((p.name, "Tool-call Success Rate theo kịch bản", "Chatbot Agent — tool reliability"))

    p = chart_chatbot_manual_outcomes()
    fig_meta.append((p.name, "Phân bố kết quả rà soát thủ công trong 40 lượt", "Chatbot Agent — chất lượng"))

    p = chart_autofill_quality(auto_df)
    fig_meta.append((p.name, "Exact Match, ROUGE-L và F1 của các trường metadata", "Submission Autofill — trích xuất metadata"))

    for path, content, placement in chart_tca_workflow_panels(tca_s):
        fig_meta.append((path.name, content, placement))

    p = chart_workflow_headline(
        {"rule": rule_s, "llm": llm_s, "chatbot": chat_s, "autofill": auto_s, "tca": tca_s}
    )
    fig_meta.append((p.name, "Các chỉ số chính dùng cho báo cáo", "Chương 5 — tóm tắt kết quả"))

    p = chart_resource_latency()
    fig_meta.append(
        (
            p.name,
            "Độ trễ trung bình, trung vị và cao nhất theo luồng (Bảng 4.7), kèm ngưỡng 100 giây",
            "Chương 4 — tính khả thi vận hành",
        )
    )

    p = chart_resource_tokens()
    fig_meta.append(
        (
            p.name,
            "Token trung bình trên mỗi đơn vị công việc (mẫu số khác nhau — không xếp hạng chi phí)",
            "Chương 4 — tính khả thi vận hành",
        )
    )

    p = chart_resource_ops_mode()
    fig_meta.append(
        (
            p.name,
            "Khuyến nghị cách vận hành (đồng bộ / không chặn / phản hồi từng phần / xử lý nền)",
            "Chương 4 — tính khả thi vận hành",
        )
    )

    p = chart_chatbot_response_stages()
    fig_meta.append(
        (
            p.name,
            "Chatbot Agent: TTFT, Time to First Answer Token, Stream Duration, Total Duration và Tool-call Success Rate",
            "Chương 4 — Chatbot Agent / vận hành",
        )
    )

    write_figure_index(fig_meta)

    # Also drop a compact JSON for thesis tooling / appendix.
    resource_export = {
        "table": "ch4-ai-latency-token",
        "rows": OPERATIONAL_RESOURCE_ROWS,
        "chatbot_response_stages": CHATBOT_RESPONSE_STAGES,
        "figures": [
            "fig11_resource_latency.png",
            "fig12_resource_tokens.png",
            "fig13_resource_ops_mode.png",
            "fig14_chatbot_response_stages.png",
        ],
            "caveats": [
            "Mẫu số token khác nhau giữa các luồng; không xếp hạng chi phí chỉ bằng số token.",
            "Độ trễ cao nhất của Chatbot Agent là giá trị của nhóm kịch bản chậm nhất, không nhất thiết là một hội thoại riêng lẻ.",
            "Biểu đồ cách vận hành là khuyến nghị thiết kế dựa trên độ trễ, chưa phải kết quả so sánh triển khai.",
            "Chưa kiểm tra trường hợp dịch vụ AI hết thời gian chờ hoặc ngừng hoạt động trên toàn bộ luồng.",
        ],
    }
    (EXPORT / "resource_usage_metrics.json").write_text(
        json.dumps(resource_export, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print("Wrote:", xlsx_path)
    print("Figures:", FIGURES)
    print("Sheets:", "Overview + detail sheets + Resource_Usage")
    print("Charts:", len(fig_meta))
    print(
        "Headline:",
        {
            "gating_rule_acc": rule_s["blocking_verdict_accuracy"],
            "llm_contract_violations": llm_s["llm_block_contract_violation_count"],
            "chatbot_manual_pass_pct": chat_s["manual_pass_rate_pct"],
            "chatbot_tool_success": chat_s["tool_call_success_rate"],
            "autofill_rouge_l": auto_s["abstract_rouge_l_mean"],
            "tca_b1_exact": tca_s["b1_exact_rate_mean"],
            "track_invalid_rate": track_s["invalid_track_rate"],
            "track_human_labels": track_s["human_label_filled"],
            "tca_ann_total": tca_s.get("b1_annotation_total"),
            "tca_ap_total": tca_s.get("b2_attention_point_total"),
            "tca_findings": tca_s.get("b3_finding_total"),
            "tca_claims": tca_s.get("b5_claim_total"),
            "resource_rows": len(OPERATIONAL_RESOURCE_ROWS),
        },
    )


if __name__ == "__main__":
    main()
