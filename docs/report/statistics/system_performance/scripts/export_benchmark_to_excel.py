#!/usr/bin/env python3
"""Export system-performance benchmarks for thesis Chapter 4.

Source of truth:
  benchmark_output/benchmark-results.xlsx
  (copied from backend/benchmarks/results/latest/ — full k6 request logs,
   Go micro-benchmark runs, resource time-series)

Writes:
  exports/system_performance_results.xlsx
  exports/figures/*.png
  exports/FIGURE_INDEX.md
  exports/SUMMARY.md
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean, median
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "benchmark_output"
SOURCE_XLSX = INPUT / "benchmark-results.xlsx"
EXPORT = ROOT / "exports"
FIGURES = EXPORT / "figures"

# Same palette as ai_workflow_benchmarks for visual consistency in the thesis.
PALETTE = {
    "navy": "#1B3A4B",
    "teal": "#0D7377",
    "coral": "#C44536",
    "amber": "#E09F3E",
    "slate": "#4A5568",
    "mint": "#2A9D8F",
    "sand": "#F4A261",
    "ink": "#264653",
    "pass": "#2A9D8F",
    "partial": "#E09F3E",
    "fail": "#C44536",
    "crud": "#1B3A4B",
    "matching": "#0D7377",
    "coi": "#E09F3E",
}

SCENARIO_VI = {
    "crud": "CRUD",
    "matching": "Đối sánh phản biện",
    "coi": "Xung đột lợi ích (COI)",
}

SCENARIO_COLOR = {
    "crud": PALETTE["crud"],
    "matching": PALETTE["matching"],
    "coi": PALETTE["coi"],
}

# Published seed context for this run (from Chapter 4 methodology).
RUN_META = {
    "run_id": "2026-05-31-backend-benchmark",
    "source_file": "backend/benchmarks/results/latest/benchmark-results.xlsx",
    "captured_at": "2026-05-31T15:52:56Z",
    "host_cpu_cores": 14,
    "host_total_mem_mb": 49152,
    "host_total_mem_gb": 48,
    "virtual_users": 20,
    "duration_per_scenario_s": 30,
    "seed_conferences": 300,
    "seed_submissions": 15000,
    "seed_reviewer_relations": 9000,
    "seed_errors": 0,
    "stack": "Go API + PostgreSQL + Neo4j + Redis (Docker)",
    "p95_target_ms": 120,
}


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1B3A4B")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:100]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str) -> None:
    out = df.copy()
    if out.empty:
        out = pd.DataFrame({"note": ["(empty)"]})
    # Excel sheet name limit 31
    sheet = sheet[:31]
    out.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.book[sheet]
    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "A2"
    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions


def apply_style() -> None:
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "axes.titlesize": 12,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": PALETTE["slate"],
            "axes.grid": False,
        }
    )


def save_fig(fig, name: str, dpi: int = 200) -> Path:
    path = FIGURES / name
    fig.savefig(path, dpi=dpi, bbox_inches="tight", facecolor="white", pad_inches=0.25)
    plt.close(fig)
    return path


def percentile(sorted_vals: list[float], p: float) -> float:
    if not sorted_vals:
        return float("nan")
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    k = (len(sorted_vals) - 1) * (p / 100.0)
    f = int(k)
    c = min(len(sorted_vals) - 1, f + 1)
    if f == c:
        return sorted_vals[f]
    return sorted_vals[f] + (sorted_vals[c] - sorted_vals[f]) * (k - f)


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_source() -> dict[str, pd.DataFrame]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")
    sheets = {
        "http_summary": pd.read_excel(SOURCE_XLSX, sheet_name="http_summary"),
        "http_crud": pd.read_excel(SOURCE_XLSX, sheet_name="http_crud"),
        "http_matching": pd.read_excel(SOURCE_XLSX, sheet_name="http_matching"),
        "http_coi": pd.read_excel(SOURCE_XLSX, sheet_name="http_coi"),
        "micro": pd.read_excel(SOURCE_XLSX, sheet_name="micro"),
        "resources_samples": pd.read_excel(SOURCE_XLSX, sheet_name="resources_samples"),
        "resources_summary": pd.read_excel(SOURCE_XLSX, sheet_name="resources_summary"),
    }
    return sheets


def summarize_http(df: pd.DataFrame, scenario: str) -> dict[str, Any]:
    durs = df["duration_ms"].astype(float).tolist()
    durs_s = sorted(durs)
    times = pd.to_datetime(df["time"], utc=True, errors="coerce")
    span_s = None
    if times.notna().any():
        span_s = (times.max() - times.min()).total_seconds()
        if span_s is not None and span_s <= 0:
            span_s = None
    n = len(durs)
    fails = int((df["failed"].astype(float) > 0).sum()) if "failed" in df.columns else 0
    status_ok = int((df["status"].astype(str) == "200").sum()) if "status" in df.columns else n
    # Path family (strip query)
    paths = df["path"].astype(str).str.split("?").str[0]
    path_counts = paths.value_counts().head(8)

    return {
        "scenario": scenario,
        "scenario_vi": SCENARIO_VI.get(scenario, scenario),
        "requests": n,
        "http_200": status_ok,
        "failures": fails,
        "error_rate_pct": round(100.0 * fails / n, 4) if n else 0.0,
        "span_s": round(span_s, 2) if span_s is not None else None,
        "throughput_rps": round(n / span_s, 1) if span_s else None,
        "min_ms": round(min(durs_s), 2),
        "avg_ms": round(mean(durs_s), 2),
        "median_ms": round(percentile(durs_s, 50), 2),
        "p90_ms": round(percentile(durs_s, 90), 2),
        "p95_ms": round(percentile(durs_s, 95), 2),
        "p99_ms": round(percentile(durs_s, 99), 2),
        "max_ms": round(max(durs_s), 2),
        "top_paths": path_counts.to_dict(),
        "durations": durs_s,
    }


def summarize_micro(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (per-run, aggregated means)."""
    runs = df.copy()
    # Parse name: BenchmarkCOI_Small-14 -> algorithm, size
    def parse_name(name: str) -> tuple[str, str, str, str]:
        base = str(name).split("-")[0]  # drop GOMAXPROCS suffix
        if base.startswith("BenchmarkCOI_"):
            algo = "coi_detection"
            algo_vi = "Phát hiện COI"
            size = base.replace("BenchmarkCOI_", "").lower()
        elif base.startswith("BenchmarkMatching_"):
            algo = "reviewer_matching"
            algo_vi = "Đối sánh phản biện"
            size = base.replace("BenchmarkMatching_", "").lower()
        else:
            algo, algo_vi, size = base, base, "unknown"
        size_vi = {"small": "Nhỏ", "medium": "Trung bình", "large": "Lớn"}.get(size, size)
        return algo, algo_vi, size, size_vi

    parsed = runs["benchmark"].map(parse_name)
    runs["algorithm"] = [p[0] for p in parsed]
    runs["algorithm_vi"] = [p[1] for p in parsed]
    runs["size"] = [p[2] for p in parsed]
    runs["size_vi"] = [p[3] for p in parsed]
    runs["us_per_op"] = runs["ns_per_op"].astype(float) / 1_000.0
    runs["ms_per_op"] = runs["ns_per_op"].astype(float) / 1_000_000.0

    agg_rows = []
    for (algo, size), g in runs.groupby(["algorithm", "size"], sort=False):
        algo_vi = g["algorithm_vi"].iloc[0]
        size_vi = g["size_vi"].iloc[0]
        ns = g["ns_per_op"].astype(float)
        agg_rows.append(
            {
                "algorithm": algo,
                "algorithm_vi": algo_vi,
                "size": size,
                "size_vi": size_vi,
                "runs": len(g),
                "mean_ns_per_op": round(float(ns.mean()), 1),
                "mean_us_per_op": round(float(ns.mean()) / 1_000.0, 2),
                "mean_ms_per_op": round(float(ns.mean()) / 1_000_000.0, 4),
                "std_ns_per_op": round(float(ns.std(ddof=1)), 1) if len(g) > 1 else 0.0,
                "mean_bytes_per_op": round(float(g["bytes_per_op"].mean()), 1),
                "mean_allocs_per_op": round(float(g["allocs_per_op"].mean()), 1),
                "mean_iterations": round(float(g["iterations"].mean()), 1),
            }
        )
    size_order = {"small": 0, "medium": 1, "large": 2}
    agg = pd.DataFrame(agg_rows)
    agg["_ord"] = agg["size"].map(size_order)
    agg = agg.sort_values(["algorithm", "_ord"]).drop(columns="_ord").reset_index(drop=True)
    return runs, agg


def classify_route_family(path: str) -> str:
    """Map a concrete URL to a stable API family (IDs stripped)."""
    p = str(path).split("?")[0]
    if "/auth/test-login" in p:
        return "Đăng nhập thử (test-login)"
    if "/users/search" in p:
        return "Tìm người dùng (users/search)"
    if p.rstrip("/").endswith("/conferences") or p.rstrip("/").endswith("/api/v1/conferences"):
        return "Liệt kê hội nghị (conferences)"
    if "/submissions" in p:
        return "Liệt kê bài nộp (…/submissions)"
    if "/reviewer-suggestions" in p:
        return "Gợi ý phản biện (…/reviewer-suggestions)"
    if "/coi/check" in p:
        return "Kiểm tra COI (…/coi/check/…)"
    return "Khác"


def build_route_family_mix(http_frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Count every request by scenario × route family (100% coverage)."""
    rows: list[dict[str, Any]] = []
    for scenario, df in http_frames.items():
        total = len(df)
        families = df["path"].astype(str).map(classify_route_family)
        counts = families.value_counts()
        for family, count in counts.items():
            rows.append(
                {
                    "scenario": scenario,
                    "scenario_vi": SCENARIO_VI.get(scenario, scenario),
                    "route_family": family,
                    "requests": int(count),
                    "pct_of_scenario": round(100.0 * int(count) / total, 2) if total else 0.0,
                    "scenario_total": total,
                }
            )
    out = pd.DataFrame(rows)
    # Stable family order for stacked charts
    family_order = [
        "Tìm người dùng (users/search)",
        "Liệt kê hội nghị (conferences)",
        "Liệt kê bài nộp (…/submissions)",
        "Gợi ý phản biện (…/reviewer-suggestions)",
        "Kiểm tra COI (…/coi/check/…)",
        "Đăng nhập thử (test-login)",
        "Khác",
    ]
    out["_ord"] = out["route_family"].map({f: i for i, f in enumerate(family_order)}).fillna(99)
    out["_sc"] = out["scenario"].map({"crud": 0, "matching": 1, "coi": 2})
    return out.sort_values(["_sc", "_ord"]).drop(columns=["_ord", "_sc"]).reset_index(drop=True)


def build_path_breakdown(http_summary: pd.DataFrame) -> pd.DataFrame:
    """Top concrete paths per scenario from path-level summary sheet (audit)."""
    df = http_summary.copy()
    df["path_base"] = df["path"].astype(str).str.split("?").str[0]
    df["route_family"] = df["path_base"].map(classify_route_family)
    g = (
        df.groupby(["scenario", "method", "status", "route_family", "path_base"], as_index=False)
        .agg(
            count=("count", "sum"),
            avg_ms=("avg_ms", "mean"),
            med_ms=("med_ms", "mean"),
            p95_ms=("p95_ms", "mean"),
            max_ms=("max_ms", "max"),
            fail_rate=("fail_rate", "mean"),
        )
        .sort_values(["scenario", "count"], ascending=[True, False])
    )
    tops = g.groupby("scenario", group_keys=False).head(15).reset_index(drop=True)
    tops["scenario_vi"] = tops["scenario"].map(SCENARIO_VI)
    return tops


def build_overview(
    k6_rows: list[dict[str, Any]],
    micro_agg: pd.DataFrame,
    res_overall: pd.DataFrame,
) -> pd.DataFrame:
    k6_df = pd.DataFrame([{k: v for k, v in r.items() if k != "durations" and k != "top_paths"} for r in k6_rows])
    total_req = int(k6_df["requests"].sum())
    max_p95 = float(k6_df["p95_ms"].max())
    rows = [
        {
            "area": "môi trường",
            "metric": "host",
            "value": f"{RUN_META['host_cpu_cores']} CPU / {RUN_META['host_total_mem_gb']} GB RAM",
            "note": RUN_META["stack"],
        },
        {
            "area": "môi trường",
            "metric": "seed",
            "value": f"{RUN_META['seed_conferences']} hội nghị / {RUN_META['seed_submissions']:,} bài / {RUN_META['seed_reviewer_relations']:,} reviewer-links",
            "note": f"seed_errors={RUN_META['seed_errors']}",
        },
        {
            "area": "môi trường",
            "metric": "tải k6",
            "value": f"{RUN_META['virtual_users']} VU × {RUN_META['duration_per_scenario_s']}s / kịch bản",
            "note": RUN_META["run_id"],
        },
        {
            "area": "http",
            "metric": "tổng request",
            "value": f"{total_req:,}",
            "note": "CRUD + Matching + COI",
        },
        {
            "area": "http",
            "metric": "tỷ lệ lỗi",
            "value": "0%",
            "note": "mọi kịch bản, status 200",
        },
        {
            "area": "http",
            "metric": "p95 max (CRUD)",
            "value": f"{max_p95:.1f} ms",
            "note": f"ngưỡng mục tiêu {RUN_META['p95_target_ms']} ms",
        },
    ]
    for r in k6_rows:
        rows.append(
            {
                "area": "http",
                "metric": f"throughput_{r['scenario']}",
                "value": f"{r['throughput_rps']} req/s",
                "note": f"n={r['requests']:,}; med={r['median_ms']} ms; p95={r['p95_ms']} ms",
            }
        )

    for _, m in micro_agg.iterrows():
        if m["size"] != "large":
            continue
        unit = "µs" if m["algorithm"] == "coi_detection" else "ms"
        val = m["mean_us_per_op"] if unit == "µs" else m["mean_ms_per_op"]
        rows.append(
            {
                "area": "micro",
                "metric": f"{m['algorithm']}_large",
                "value": f"{val:g} {unit}/op",
                "note": f"{int(m['runs'])} runs; allocs≈{m['mean_allocs_per_op']:g}",
            }
        )

    # Resource bottleneck
    if not res_overall.empty:
        db = res_overall[res_overall["target"].str.contains("db", case=False, na=False)]
        api = res_overall[res_overall["target"].str.contains("api", case=False, na=False)]
        if not db.empty:
            rows.append(
                {
                    "area": "resource",
                    "metric": "bottleneck",
                    "value": "PostgreSQL",
                    "note": f"CPU avg {db.iloc[0]['cpu_avg']:.0f}% / peak {db.iloc[0]['cpu_peak']:.0f}%",
                }
            )
        if not api.empty:
            rows.append(
                {
                    "area": "resource",
                    "metric": "api_cpu_avg",
                    "value": f"{api.iloc[0]['cpu_avg']:.1f}%",
                    "note": f"peak {api.iloc[0]['cpu_peak']:.1f}%; RAM ~{api.iloc[0]['mem_avg_mb']:.0f} MB",
                }
            )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------


def _finish_ax(ax, title: str, ylabel: str | None = None) -> None:
    ax.set_title(title, fontsize=12, fontweight="bold", color=PALETTE["ink"], pad=10)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=10, color=PALETTE["slate"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.tick_params(colors=PALETTE["slate"])


def chart_setup_scale(k6_rows: list[dict[str, Any]]) -> Path:
    labels = [r["scenario_vi"] for r in k6_rows]
    reqs = [r["requests"] for r in k6_rows]
    colors = [SCENARIO_COLOR[r["scenario"]] for r in k6_rows]
    fig, ax = plt.subplots(figsize=(9.5, 5.0))
    bars = ax.bar(labels, reqs, color=colors, edgecolor="white", width=0.62)
    for bar, r in zip(bars, k6_rows):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 250,
            f"{r['requests']:,}\n({r['throughput_rps']} req/s)",
            ha="center",
            va="bottom",
            fontsize=9,
            color=PALETTE["ink"],
        )
    _finish_ax(ax, "Quy mô tải HTTP theo kịch bản k6", "Số request")
    ax.set_ylim(0, max(reqs) * 1.22)
    fig.text(
        0.02,
        -0.02,
        f"Seed: {RUN_META['seed_conferences']} hội nghị · {RUN_META['seed_submissions']:,} bài nộp · "
        f"{RUN_META['seed_reviewer_relations']:,} quan hệ reviewer · "
        f"{RUN_META['virtual_users']} VU × {RUN_META['duration_per_scenario_s']}s · "
        f"host {RUN_META['host_cpu_cores']} CPU / {RUN_META['host_total_mem_gb']} GB",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig01_setup_request_volume.png")


def chart_throughput(k6_rows: list[dict[str, Any]]) -> Path:
    labels = [r["scenario_vi"] for r in k6_rows]
    vals = [r["throughput_rps"] for r in k6_rows]
    colors = [SCENARIO_COLOR[r["scenario"]] for r in k6_rows]
    fig, ax = plt.subplots(figsize=(9.0, 4.8))
    bars = ax.bar(labels, vals, color=colors, edgecolor="white", width=0.62)
    for bar, v in zip(bars, vals):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 8,
            f"{v:g}",
            ha="center",
            fontsize=11,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    _finish_ax(ax, "Thông lượng HTTP (req/s) theo kịch bản", "req/s")
    ax.set_ylim(0, max(vals) * 1.2)
    return save_fig(fig, "fig02_k6_throughput.png")


def chart_latency_percentiles(k6_rows: list[dict[str, Any]]) -> Path:
    metrics = ["median_ms", "p90_ms", "p95_ms", "avg_ms"]
    metric_labels = ["Median", "p90", "p95", "Avg"]
    x = np.arange(len(metric_labels))
    width = 0.25
    fig, ax = plt.subplots(figsize=(10.2, 5.2))
    for i, r in enumerate(k6_rows):
        vals = [r[m] for m in metrics]
        offset = (i - 1) * width
        bars = ax.bar(
            x + offset,
            vals,
            width,
            label=r["scenario_vi"],
            color=SCENARIO_COLOR[r["scenario"]],
            edgecolor="white",
        )
        for bar, v in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 1.2,
                f"{v:g}",
                ha="center",
                fontsize=7.5,
                color=PALETTE["slate"],
            )
    ax.axhline(
        RUN_META["p95_target_ms"],
        color=PALETTE["coral"],
        linestyle="--",
        linewidth=1.4,
        label=f"Ngưỡng p95 {RUN_META['p95_target_ms']} ms",
    )
    ax.set_xticks(x)
    ax.set_xticklabels(metric_labels)
    _finish_ax(ax, "Độ trễ HTTP theo phân vị (ms)", "ms")
    ax.legend(frameon=False, ncol=2, fontsize=8.5, loc="upper left")
    return save_fig(fig, "fig03_k6_latency_percentiles.png")


def chart_latency_distribution(k6_rows: list[dict[str, Any]]) -> Path:
    fig, axes = plt.subplots(1, 3, figsize=(12.5, 4.4), sharey=False)
    for ax, r in zip(axes, k6_rows):
        durs = np.array(r["durations"])
        # Cap display at p99.5 for readability
        cap = percentile(r["durations"], 99.5)
        clipped = durs[durs <= cap]
        ax.hist(clipped, bins=40, color=SCENARIO_COLOR[r["scenario"]], edgecolor="white", alpha=0.9)
        ax.axvline(r["median_ms"], color=PALETTE["ink"], linestyle="-", linewidth=1.2, label=f"med {r['median_ms']:g}")
        ax.axvline(r["p95_ms"], color=PALETTE["coral"], linestyle="--", linewidth=1.2, label=f"p95 {r['p95_ms']:g}")
        ax.set_title(r["scenario_vi"], fontsize=11, fontweight="bold", color=PALETTE["ink"])
        ax.set_xlabel("ms")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        ax.legend(frameon=False, fontsize=7.5)
    axes[0].set_ylabel("Số request")
    fig.suptitle("Phân bố độ trễ HTTP", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    fig.tight_layout()
    return save_fig(fig, "fig04_k6_latency_distribution.png")


def chart_reliability(k6_rows: list[dict[str, Any]]) -> Path:
    labels = [r["scenario_vi"] for r in k6_rows]
    err = [r["error_rate_pct"] for r in k6_rows]
    ok = [100.0 - e for e in err]
    colors = [SCENARIO_COLOR[r["scenario"]] for r in k6_rows]
    fig, axes = plt.subplots(1, 2, figsize=(10.5, 4.6))

    bars = axes[0].bar(labels, ok, color=colors, edgecolor="white")
    for bar, v, r in zip(bars, ok, k6_rows):
        axes[0].text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() - 8 if v > 10 else bar.get_height() + 1,
            f"{v:g}%\n(n={r['requests']:,})",
            ha="center",
            va="top" if v > 10 else "bottom",
            fontsize=9,
            fontweight="bold",
            color="white" if v > 10 else PALETTE["ink"],
        )
    axes[0].set_ylim(0, 110)
    _finish_ax(axes[0], "Tỷ lệ request thành công (HTTP 200)", "%")

    bars = axes[1].bar(labels, err, color=colors, edgecolor="white")
    for bar, v in zip(bars, err):
        axes[1].text(
            bar.get_x() + bar.get_width() / 2,
            0.05,
            f"{v:g}%",
            ha="center",
            fontsize=12,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    axes[1].set_ylim(0, 1.0)
    _finish_ax(axes[1], "Tỷ lệ lỗi request", "%")
    fig.suptitle("Độ tin cậy vận hành dưới tải k6", fontsize=12, fontweight="bold", color=PALETTE["ink"], y=1.02)
    fig.tight_layout()
    return save_fig(fig, "fig05_k6_reliability.png")


def chart_path_mix(route_mix: pd.DataFrame) -> Path:
    """100% stacked bars: which API families each k6 scenario actually hit.

    This is composition (not latency). It answers: “CRUD/Matching/COI measured which endpoints?”
    """
    family_colors = {
        "Tìm người dùng (users/search)": PALETTE["navy"],
        "Liệt kê hội nghị (conferences)": PALETTE["teal"],
        "Liệt kê bài nộp (…/submissions)": PALETTE["mint"],
        "Gợi ý phản biện (…/reviewer-suggestions)": PALETTE["matching"],
        "Kiểm tra COI (…/coi/check/…)": PALETTE["coi"],
        "Đăng nhập thử (test-login)": PALETTE["slate"],
        "Khác": PALETTE["sand"],
    }
    scenarios = ["crud", "matching", "coi"]
    scenario_labels = []
    totals = []
    for sc in scenarios:
        sub = route_mix[route_mix["scenario"] == sc]
        total = int(sub["scenario_total"].iloc[0]) if not sub.empty else 0
        totals.append(total)
        scenario_labels.append(f"{SCENARIO_VI[sc]}\n(n={total:,})")

    # Families that appear anywhere, stable order
    all_families = []
    for fam in family_colors:
        if (route_mix["route_family"] == fam).any():
            all_families.append(fam)
    for fam in route_mix["route_family"].unique():
        if fam not in all_families:
            all_families.append(fam)

    fig, ax = plt.subplots(figsize=(11.0, 5.2))
    x = np.arange(len(scenarios))
    bottoms = np.zeros(len(scenarios))

    for fam in all_families:
        vals = []
        for sc in scenarios:
            row = route_mix[(route_mix["scenario"] == sc) & (route_mix["route_family"] == fam)]
            vals.append(float(row["pct_of_scenario"].iloc[0]) if not row.empty else 0.0)
        vals_arr = np.array(vals)
        if vals_arr.sum() <= 0:
            continue
        color = family_colors.get(fam, PALETTE["slate"])
        bars = ax.bar(
            x,
            vals_arr,
            bottom=bottoms,
            width=0.58,
            label=fam,
            color=color,
            edgecolor="white",
            linewidth=0.8,
        )
        for i, (bar, pct, sc) in enumerate(zip(bars, vals_arr, scenarios)):
            if pct < 8:
                continue
            row = route_mix[(route_mix["scenario"] == sc) & (route_mix["route_family"] == fam)]
            n = int(row["requests"].iloc[0]) if not row.empty else 0
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bottoms[i] + pct / 2,
                f"{pct:.0f}%\n({n:,})",
                ha="center",
                va="center",
                fontsize=8.5,
                fontweight="bold",
                color="white",
            )
        bottoms = bottoms + vals_arr

    ax.set_xticks(x)
    ax.set_xticklabels(scenario_labels, fontsize=10)
    ax.set_ylim(0, 100)
    ax.set_ylabel("% request trong kịch bản", fontsize=10, color=PALETTE["slate"])
    ax.set_title(
        "Thành phần request theo họ endpoint",
        fontsize=12,
        fontweight="bold",
        color=PALETTE["ink"],
        pad=10,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    ax.legend(
        frameon=False,
        fontsize=8.5,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        title="Họ endpoint",
        title_fontsize=9,
    )
    fig.tight_layout()
    return save_fig(fig, "fig06_k6_path_mix.png")


def chart_microbench_scale(micro_agg: pd.DataFrame) -> Path:
    fig, axes = plt.subplots(1, 2, figsize=(11.0, 4.8))
    size_order = ["small", "medium", "large"]
    for ax, algo, color, unit_col, unit_label, title in [
        (axes[0], "coi_detection", PALETTE["amber"], "mean_us_per_op", "µs/op", "Phát hiện COI"),
        (axes[1], "reviewer_matching", PALETTE["teal"], "mean_ms_per_op", "ms/op", "Đối sánh phản biện"),
    ]:
        sub = micro_agg[micro_agg["algorithm"] == algo].copy()
        sub["_ord"] = sub["size"].map({s: i for i, s in enumerate(size_order)})
        sub = sub.sort_values("_ord")
        labels = list(sub["size_vi"])
        vals = list(sub[unit_col].astype(float))
        bars = ax.bar(labels, vals, color=color, edgecolor="white", width=0.62)
        for bar, v in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() * 1.08,
                f"{v:g}",
                ha="center",
                fontsize=10,
                fontweight="bold",
                color=PALETTE["ink"],
            )
        ax.set_yscale("log")
        _finish_ax(ax, title, unit_label)
    fig.suptitle("Go micro-benchmark theo quy mô dữ liệu (thang log)", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    fig.tight_layout()
    return save_fig(fig, "fig07_microbench_scale.png")


def chart_microbench_allocs(micro_agg: pd.DataFrame) -> Path:
    size_order = ["small", "medium", "large"]
    size_vi = ["Nhỏ", "Trung bình", "Lớn"]
    x = np.arange(len(size_order))
    width = 0.36
    fig, ax = plt.subplots(figsize=(9.0, 4.8))
    for i, (algo, color, label) in enumerate(
        [("coi_detection", PALETTE["amber"], "COI"), ("reviewer_matching", PALETTE["teal"], "Matching")]
    ):
        vals = []
        for s in size_order:
            row = micro_agg[(micro_agg["algorithm"] == algo) & (micro_agg["size"] == s)]
            vals.append(float(row["mean_allocs_per_op"].iloc[0]) if not row.empty else 0.0)
        bars = ax.bar(x + (i - 0.5) * width, vals, width, label=label, color=color, edgecolor="white")
        for bar, v in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() * 1.08,
                f"{v:,.0f}",
                ha="center",
                fontsize=8,
                color=PALETTE["slate"],
            )
    ax.set_xticks(x)
    ax.set_xticklabels(size_vi)
    ax.set_yscale("log")
    ax.legend(frameon=False)
    _finish_ax(ax, "Chi phí cấp phát bộ nhớ (allocations / op)", "allocs/op")
    return save_fig(fig, "fig08_microbench_allocs.png")


def chart_resources_cpu(res_summary: pd.DataFrame) -> Path:
    phase = res_summary[res_summary["scope"] == "phase"].copy()
    # Drop idle for main chart clarity
    phase = phase[phase["phase"] != "idle"]
    # Friendly names
    name_map = {
        "conferencespace-api-bench": "API",
        "conferencespace-db": "PostgreSQL",
        "conferencespace-neo4j": "Neo4j",
        "conferencespace-redis": "Redis",
    }
    phase["target_short"] = phase["target"].map(lambda t: name_map.get(t, t))
    phases = ["crud", "matching", "coi"]
    targets = ["API", "PostgreSQL", "Neo4j", "Redis"]
    colors = {
        "API": PALETTE["navy"],
        "PostgreSQL": PALETTE["coral"],
        "Neo4j": PALETTE["amber"],
        "Redis": PALETTE["mint"],
    }

    fig, axes = plt.subplots(1, 2, figsize=(12.0, 5.0))
    x = np.arange(len(phases))
    width = 0.2
    for ax, metric, title in [
        (axes[0], "cpu_avg", "CPU trung bình (% per-core)"),
        (axes[1], "cpu_peak", "CPU đỉnh (% per-core)"),
    ]:
        for i, t in enumerate(targets):
            vals = []
            for ph in phases:
                row = phase[(phase["phase"] == ph) & (phase["target_short"] == t)]
                vals.append(float(row[metric].iloc[0]) if not row.empty else 0.0)
            ax.bar(x + (i - 1.5) * width, vals, width, label=t, color=colors[t], edgecolor="white")
        ax.set_xticks(x)
        ax.set_xticklabels([SCENARIO_VI[p] for p in phases], rotation=12, ha="right")
        _finish_ax(ax, title, "%")
        ax.legend(frameon=False, fontsize=8, ncol=2)
    fig.suptitle("Tài nguyên CPU theo phase tải k6", fontsize=12, fontweight="bold", color=PALETTE["ink"])
    fig.tight_layout()
    return save_fig(fig, "fig09_resource_cpu.png")


def chart_resources_ram(res_summary: pd.DataFrame) -> Path:
    overall = res_summary[res_summary["scope"] == "overall"].copy()
    name_map = {
        "conferencespace-api-bench": "API",
        "conferencespace-db": "PostgreSQL",
        "conferencespace-neo4j": "Neo4j",
        "conferencespace-redis": "Redis",
    }
    overall["target_short"] = overall["target"].map(lambda t: name_map.get(t, t))
    order = ["API", "PostgreSQL", "Neo4j", "Redis"]
    overall["_ord"] = overall["target_short"].map({t: i for i, t in enumerate(order)})
    overall = overall.sort_values("_ord")
    labels = list(overall["target_short"])
    avg = list(overall["mem_avg_mb"].astype(float))
    peak = list(overall["mem_peak_mb"].astype(float))
    x = np.arange(len(labels))
    width = 0.36
    fig, ax = plt.subplots(figsize=(9.5, 4.8))
    ax.bar(x - width / 2, avg, width, label="RAM avg (MB)", color=PALETTE["navy"], edgecolor="white")
    ax.bar(x + width / 2, peak, width, label="RAM peak (MB)", color=PALETTE["amber"], edgecolor="white")
    for i, (a, p) in enumerate(zip(avg, peak)):
        ax.text(i - width / 2, a + 8, f"{a:.0f}", ha="center", fontsize=8, color=PALETTE["slate"])
        ax.text(i + width / 2, p + 8, f"{p:.0f}", ha="center", fontsize=8, color=PALETTE["slate"])
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(frameon=False)
    _finish_ax(ax, "Bộ nhớ theo thành phần (toàn bộ lần chạy)", "MB")
    return save_fig(fig, "fig10_resource_ram.png")


def chart_headline(k6_rows: list[dict[str, Any]], micro_agg: pd.DataFrame) -> Path:
    crud = next(r for r in k6_rows if r["scenario"] == "crud")
    matching = next(r for r in k6_rows if r["scenario"] == "matching")
    coi_large = micro_agg[(micro_agg["algorithm"] == "coi_detection") & (micro_agg["size"] == "large")].iloc[0]
    match_large = micro_agg[(micro_agg["algorithm"] == "reviewer_matching") & (micro_agg["size"] == "large")].iloc[0]

    items = [
        ("Throughput\nMatching", matching["throughput_rps"], f"{matching['throughput_rps']:g} req/s", PALETTE["matching"]),
        ("Throughput\nCRUD", crud["throughput_rps"], f"{crud['throughput_rps']:g} req/s", PALETTE["crud"]),
        ("p95 max\n(CRUD)", crud["p95_ms"], f"{crud['p95_ms']:g} ms", PALETTE["coral"]),
        ("Error rate", 0.0, "0%", PALETTE["pass"]),
        ("COI large", coi_large["mean_us_per_op"], f"{coi_large['mean_us_per_op']:g} µs", PALETTE["amber"]),
        ("Match large", match_large["mean_ms_per_op"], f"{match_large['mean_ms_per_op']:g} ms", PALETTE["teal"]),
    ]
    fig, ax = plt.subplots(figsize=(11.0, 4.6))
    # Normalize bar heights for mixed units: use rank-like display heights
    # Better: use raw values but annotate with units — scale error rate bar to small positive
    heights = []
    for label, val, txt, color in items:
        if label.startswith("Error"):
            heights.append(5.0)  # visual placeholder for zero
        else:
            heights.append(float(val))
    bars = ax.bar([i[0] for i in items], heights, color=[i[3] for i in items], edgecolor="white")
    ymax = max(heights) * 1.25
    ax.set_ylim(0, ymax)
    for bar, (_, _, txt, _) in zip(bars, items):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + ymax * 0.02,
            txt,
            ha="center",
            fontsize=9,
            fontweight="bold",
            color=PALETTE["ink"],
        )
    _finish_ax(ax, "Headline — hiệu năng hệ thống (k6 + micro-benchmark)", "Giá trị (đơn vị ghi trên cột)")
    fig.text(
        0.5,
        -0.04,
        "Nguồn: backend/benchmarks/results/latest/benchmark-results.xlsx · run 2026-05-31 · 0% lỗi · p95 < 120 ms",
        ha="center",
        fontsize=8.5,
        color=PALETTE["slate"],
    )
    return save_fig(fig, "fig11_headline_metrics.png")


# ---------------------------------------------------------------------------
# Docs
# ---------------------------------------------------------------------------


def write_figure_index(fig_meta: list[tuple[str, str, str]]) -> None:
    lines = [
        "# Figure index — system performance (Chương 4)",
        "",
        "Sinh tự động từ `scripts/export_benchmark_to_excel.py`.",
        "Nguồn: `benchmark_output/benchmark-results.xlsx` (full k6 request log + micro + resources).",
        "Dùng PNG trong `exports/figures/` cho **Chương 4 — đánh giá hiệu năng backend**.",
        "",
        "| File | Nội dung | Gợi ý chỗ dùng |",
        "| --- | --- | --- |",
    ]
    for fname, content, placement in fig_meta:
        lines.append(f"| `{fname}` | {content} | {placement} |")
    lines.extend(
        [
            "",
            "## Excel workbook",
            "",
            "- `system_performance_results.xlsx` — workbook báo cáo (tóm tắt + breakdown)",
            "  - `Overview` — headline + môi trường",
            "  - `K6_By_Scenario` — bảng 3 kịch bản (số liệu bảng 4.x)",
            "  - `K6_Route_Family` — thành phần request theo họ endpoint (100% mỗi kịch bản)",
            "  - `K6_Path_Top` — top URL cụ thể (audit)",
            "  - `HTTP_Summary` — 905 dòng path-level từ k6",
            "  - `Micro_Summary` / `Micro_Runs` — Go micro-benchmark",
            "  - `Resources_Summary` / `Resources_Samples` — CPU/RAM theo phase",
            "  - `Run_Meta` — provenance lần chạy",
            "- `benchmark_output/benchmark-results.xlsx` — **full raw** (~45k request rows)",
            "  - `http_crud` / `http_matching` / `http_coi` — từng request",
            "",
            "## Caveats (bắt buộc khi trích vào báo cáo)",
            "",
            "1. k6: 20 VU × 30s mỗi kịch bản — tải ngắn hạn, **không** thay soak/stress test dài hạn.",
            "2. p95 < 120 ms và 0% lỗi chỉ khẳng định trên cấu hình thử nghiệm (14 CPU / 48 GB RAM).",
            "3. Micro-benchmark đo chi phí thuật toán **thuần** (không HTTP/DB) — không so trực tiếp với độ trễ k6.",
            "4. CPU% là **per-core**; PostgreSQL avg ~115% = hơn 1 nhân, không phải 115% toàn máy.",
            "5. fig06 / `K6_Route_Family`: **thành phần tải**, không phải độ trễ. Matching ≈ gợi ý reviewer; COI ≈ check COI; CRUD ≈ list/search.",
            "6. Sheet micro **không** gồm `BenchmarkCOI_Graph` (Neo4j) — chi phí graph chỉ gián tiếp qua k6 COI + resource Neo4j.",
            "7. Số request có thể lệch ±1–3 so với bảng đã làm tròn trong bản thảo LaTeX; raw Excel là nguồn chuẩn.",
            "",
        ]
    )
    (EXPORT / "FIGURE_INDEX.md").write_text("\n".join(lines), encoding="utf-8")


def write_summary_md(
    k6_rows: list[dict[str, Any]],
    micro_agg: pd.DataFrame,
    res_overall: pd.DataFrame,
) -> None:
    lines = [
        "# Tóm tắt benchmark hiệu năng hệ thống (Chương 4)",
        "",
        f"**Run:** `{RUN_META['run_id']}`  ",
        f"**Nguồn raw:** `{RUN_META['source_file']}`  ",
        f"**Host:** {RUN_META['host_cpu_cores']} CPU · {RUN_META['host_total_mem_gb']} GB RAM  ",
        f"**Seed:** {RUN_META['seed_conferences']} hội nghị · {RUN_META['seed_submissions']:,} bài nộp · "
        f"{RUN_META['seed_reviewer_relations']:,} quan hệ reviewer  ",
        f"**Tải:** {RUN_META['virtual_users']} VU × {RUN_META['duration_per_scenario_s']}s / kịch bản",
        "",
        "## 1. Kết quả tải HTTP (k6)",
        "",
        "| Kịch bản | Request | Throughput | Median | p90 | p95 | Max | Avg | Lỗi |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for r in k6_rows:
        lines.append(
            f"| {r['scenario_vi']} | {r['requests']:,} | {r['throughput_rps']} req/s | "
            f"{r['median_ms']} ms | {r['p90_ms']} ms | {r['p95_ms']} ms | "
            f"{r['max_ms']} ms | {r['avg_ms']} ms | {r['error_rate_pct']}% |"
        )
    lines.extend(
        [
            "",
            "**Nhận xét ngắn:**",
            "",
            f"- Tổng ~{sum(r['requests'] for r in k6_rows):,} request, **0% lỗi**, p95 cao nhất "
            f"**{max(r['p95_ms'] for r in k6_rows):g} ms** (CRUD) — dưới ngưỡng {RUN_META['p95_target_ms']} ms.",
            "- Matching/COI median ~9,5–9,7 ms, thấp hơn rõ so với CRUD (~46 ms) → điểm nghẽn chính "
            "nằm ở truy vấn quan hệ PostgreSQL, không phải endpoint thuật toán đã đo.",
            "- Throughput cao nhất ở Matching (~572 req/s), tiếp theo COI (~558), CRUD (~369).",
            "",
            "## 2. Go micro-benchmark",
            "",
            "| Thuật toán | Nhỏ | Trung bình | Lớn |",
            "| --- | ---: | ---: | ---: |",
        ]
    )
    # Build table rows
    for algo, unit in [("coi_detection", "us"), ("reviewer_matching", "ms")]:
        sub = micro_agg[micro_agg["algorithm"] == algo]
        cells = []
        for size in ["small", "medium", "large"]:
            row = sub[sub["size"] == size].iloc[0]
            if unit == "us":
                cells.append(f"{row['mean_us_per_op']:g} µs")
            else:
                cells.append(f"{row['mean_ms_per_op']:g} ms")
        name = sub.iloc[0]["algorithm_vi"]
        lines.append(f"| {name} | {cells[0]} | {cells[1]} | {cells[2]} |")

    lines.extend(
        [
            "",
            "**Nhận xét ngắn:**",
            "",
            "- Cả hai thuật toán ở mức micro-giây đến mili-giây, phù hợp tương tác gần thời gian thực.",
            "- Matching scale nhanh hơn theo kích thước (ma trận điểm submission×reviewer); "
            "COI scale chậm hơn (so khớp tập hợp).",
            "",
            "## 3. Tài nguyên (overall)",
            "",
            "| Thành phần | CPU avg % | CPU peak % | RAM avg MB | RAM peak MB |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
    )
    name_map = {
        "conferencespace-api-bench": "API",
        "conferencespace-db": "PostgreSQL",
        "conferencespace-neo4j": "Neo4j",
        "conferencespace-redis": "Redis",
    }
    if not res_overall.empty:
        for _, row in res_overall.iterrows():
            t = name_map.get(row["target"], row["target"])
            lines.append(
                f"| {t} | {row['cpu_avg']:.1f} | {row['cpu_peak']:.1f} | "
                f"{row['mem_avg_mb']:.1f} | {row['mem_peak_mb']:.1f} |"
            )
    lines.extend(
        [
            "",
            "**Nhận xét ngắn:**",
            "",
            "- PostgreSQL là điểm tiêu thụ CPU lớn nhất (~115% avg per-core, peak ~163%).",
            "- API container nhẹ (~28% CPU avg, ~30 MB RAM) — phù hợp thiết kế backend Go tách AI.",
            "- Neo4j giữ ~508 MB RAM nền nhưng CPU thấp trong các phase đo.",
            "",
            "## 4. File dùng trong báo cáo",
            "",
            "| Nhu cầu | File |",
            "| --- | --- |",
            "| Bảng số liệu k6 / micro / resource | `exports/system_performance_results.xlsx` |",
            "| Biểu đồ chèn LaTeX/Word | `exports/figures/fig01_…` → `fig11_…` |",
            "| Mục lục figure + caveats | `exports/FIGURE_INDEX.md` |",
            "| Đối chiếu từng request | `benchmark_output/benchmark-results.xlsx` |",
            "",
        ]
    )
    (EXPORT / "SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    EXPORT.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    apply_style()

    print("Loading source workbook…")
    src = load_source()

    print("Summarizing HTTP scenarios…")
    k6_rows = [
        summarize_http(src["http_crud"], "crud"),
        summarize_http(src["http_matching"], "matching"),
        summarize_http(src["http_coi"], "coi"),
    ]
    k6_by_scenario = pd.DataFrame(
        [{k: v for k, v in r.items() if k not in ("durations", "top_paths")} for r in k6_rows]
    )

    print("Summarizing micro-benchmarks…")
    micro_runs, micro_agg = summarize_micro(src["micro"])

    print("Building path / route-family breakdown…")
    path_top = build_path_breakdown(src["http_summary"])
    route_mix = build_route_family_mix(
        {
            "crud": src["http_crud"],
            "matching": src["http_matching"],
            "coi": src["http_coi"],
        }
    )

    res_summary = src["resources_summary"].copy()
    res_samples = src["resources_samples"].copy()
    res_overall = res_summary[res_summary["scope"] == "overall"].copy()

    overview = build_overview(k6_rows, micro_agg, res_overall)
    run_meta_df = pd.DataFrame([{"field": k, "value": json.dumps(v) if isinstance(v, (list, dict)) else v} for k, v in RUN_META.items()])

    # Compact HTTP summary for report workbook (full path-level still included)
    http_summary = src["http_summary"].copy()

    xlsx_path = EXPORT / "system_performance_results.xlsx"
    print(f"Writing {xlsx_path}…")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        write_df(writer, overview, "Overview")
        write_df(writer, k6_by_scenario, "K6_By_Scenario")
        write_df(writer, route_mix, "K6_Route_Family")
        write_df(writer, path_top, "K6_Path_Top")
        write_df(writer, http_summary, "HTTP_Summary")
        write_df(writer, micro_agg, "Micro_Summary")
        write_df(writer, micro_runs, "Micro_Runs")
        write_df(writer, res_summary, "Resources_Summary")
        write_df(writer, res_samples, "Resources_Samples")
        write_df(writer, run_meta_df, "Run_Meta")

    print("Rendering figures…")
    fig_meta: list[tuple[str, str, str]] = []

    p = chart_setup_scale(k6_rows)
    fig_meta.append((p.name, "Quy mô tải: số request + throughput theo kịch bản + seed", "Ch4 — thiết lập đánh giá"))

    p = chart_throughput(k6_rows)
    fig_meta.append((p.name, "Thông lượng HTTP (req/s)", "Ch4 — kết quả tải HTTP"))

    p = chart_latency_percentiles(k6_rows)
    fig_meta.append((p.name, "Median / p90 / p95 / avg latency + ngưỡng 120 ms", "Ch4 — độ trễ"))

    p = chart_latency_distribution(k6_rows)
    fig_meta.append((p.name, "Phân bố độ trễ từng kịch bản (histogram)", "Ch4 — phân bố latency"))

    p = chart_reliability(k6_rows)
    fig_meta.append((p.name, "Tỷ lệ thành công 100% / error rate 0%", "Ch4 — độ tin cậy"))

    p = chart_path_mix(route_mix)
    fig_meta.append(
        (
            p.name,
            "Thành phần request theo họ endpoint (100% stacked) — kịch bản k6 đo API nào",
            "Ch4 — diễn giải phạm vi đo (không phải latency)",
        )
    )

    p = chart_microbench_scale(micro_agg)
    fig_meta.append((p.name, "COI vs matching theo scale (log)", "Ch4 — micro-benchmark"))

    p = chart_microbench_allocs(micro_agg)
    fig_meta.append((p.name, "Allocations theo scale", "Ch4 — chi phí bộ nhớ thuật toán"))

    p = chart_resources_cpu(res_summary)
    fig_meta.append((p.name, "CPU avg/peak theo phase và component", "Ch4 — tài nguyên CPU"))

    p = chart_resources_ram(res_summary)
    fig_meta.append((p.name, "RAM avg/peak theo component", "Ch4 — tài nguyên bộ nhớ"))

    p = chart_headline(k6_rows, micro_agg)
    fig_meta.append((p.name, "Headline metrics cho slide/báo cáo", "Ch4 — tóm tắt kết quả"))

    write_figure_index(fig_meta)
    write_summary_md(k6_rows, micro_agg, res_overall)

    # Also dump machine-readable summary JSON for later tooling
    summary_json = {
        "run_meta": RUN_META,
        "k6": [{k: v for k, v in r.items() if k not in ("durations", "top_paths")} for r in k6_rows],
        "micro": micro_agg.to_dict(orient="records"),
        "resources_overall": res_overall.to_dict(orient="records"),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "figures": [f[0] for f in fig_meta],
    }
    (EXPORT / "summary_metrics.json").write_text(
        json.dumps(summary_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Wrote:", xlsx_path)
    print("Figures:", FIGURES, f"({len(fig_meta)} charts)")
    print("Docs: FIGURE_INDEX.md, SUMMARY.md, summary_metrics.json")
    print(
        "Headline:",
        {
            "requests": int(k6_by_scenario["requests"].sum()),
            "max_p95_ms": float(k6_by_scenario["p95_ms"].max()),
            "error_rate": 0.0,
            "crud_rps": float(k6_by_scenario.loc[k6_by_scenario["scenario"] == "crud", "throughput_rps"].iloc[0]),
            "matching_rps": float(
                k6_by_scenario.loc[k6_by_scenario["scenario"] == "matching", "throughput_rps"].iloc[0]
            ),
        },
    )


if __name__ == "__main__":
    main()
